import os
import html
import datetime
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from config import EBV_VERSION

from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet

# Aktualisierte, detailgetreue Fußnoten aus den juristischen Vorgaben
FUSSNOTEN_TEXTE = [
    "1: Die Materialwerte gelten für Bodenmaterial und Baggergut mit bis zu 10 Volumenprozent (BM und BG) oder bis zu 50 Volumenprozent (BM-F und BG-F) mineralischer Fremdbestandteile im Sinne von § 2 Nummer 8 der Bundes-Bodenschutz- und Altlastenverordnung mit nur vernachlässigbaren Anteilen an Störstoffen im Sinne von § 2 Nummer 9 der Bundes-Bodenschutz- und Altlastenverordnung. Bodenmaterial der Klasse BM-0 und Baggergut der Klasse BG-0 erfüllen die wertebezogenen Anforderungen an das Auf- oder Einbringen gemäß § 7 Absatz 3 der Bundes-Bodenschutz- und Altlastenverordnung. Bodenmaterial der Klasse BM-0 und Baggergut der Klasse BG-0 Sand erfüllen die wertebezogenen Anforderungen an das Auf- oder Einbringen gemäß § 8 Absatz 2 der Bundes-Bodenschutz- und Altlastenverordnung; Bodenmaterial der Klasse BM-0* und Baggergut der Klasse BG-0* erfüllen die wertebezogenen Anforderungen an das Auf- oder Einbringen gemäß § 8 Absatz 3 Nummer 1 der Bundes-Bodenschutz- und Altlastenverordnung.",
    "2: Bodenarten-Hauptgruppen gemäß Bodenkundlicher Kartieranleitung, 5. Auflage, Hannover 2005 (KA5); stark schluffige Sande, lehmig-schluffige Sande und stark lehmige Sande sowie Materialien, die nicht bodenartspezifisch zugeordnet werden können, sind entsprechend der Bodenart Lehm, Schluff zu bewerten.",
    "3: Die Eluatwerte in Spalte 6 sind mit Ausnahme des Eluatwertes für Sulfat nur maßgeblich, wenn für den betreffenden Stoff der jeweilige Feststoffwert nach Spalte 3 bis 5 überschritten wird. Der Eluatwert für PAK15 und Napthalin und Methylnaphtaline, gesamt, ist maßgeblich, wenn der Feststoffwert für PAK16 nach Spalte 3 bis 5 überschritten wird. Die in Klammern genannten Werte gelten jeweils bei einem TOC-Gehalt von ≥ 0,5 %.",
    "4: Stoffspezifischer Orientierungswert; bei Abweichungen ist die Ursache zu prüfen.",
    "5: Bei Überschreitung des Wertes ist die Ursache zu prüfen. Handelt es sich um naturbedingt erhöhte Sulfatkonzentrationen, ist eine Verwertung innerhalb der betroffenen Gebiete möglich. Außerhalb dieser Gebiete ist über die Verwertungseignung im Einzelfall und in Abstimmung mit der zuständigen Behörde zu entscheiden.",
    "6: Der Wert 1 mg/kg gilt für Sand und Lehm/Schluff. Für Ton gilt der Wert 1,5 mg/kg.",
    "7: Bodenmaterialspezifischer Orientierungswert. Bei heterogenen Bodenverhältnissen mineralischer Böden kann der TOC-Gehalt der Masse des anfallenden Materials als maßgeblich bei Verwertung im Umfeld des anfallenden Materials und Verwendung unter gleichen Bedingungen herangezogen werden. Beim Einbau sind Volumenbeständigkeit und Setzungsprozesse sowie die Vorgaben von § 6 Absatz 11 Satz 2 und 3 der Bundes-Bodenschutz- und Altlastenverordnung zu berücksichtigen.",
    "8: Die angegebenen Werte gelten für Kohlenwasserstoffverbindungen mit einer Kettenlänge von C10 bis C22. Der Gesamtgehalt bestimmt nach der DIN EN 14039, „Charakterisierung von Abfällen – Bestimmung des Gehalts an Kohlenwasserstoffen von C10 bis C40 mittels Gaschromatographie“, Ausgabe Januar 2005 darf insgesamt den in Klammern genannten Wert nicht überschreiten.",
    "9: PAK15: PAK16 ohne Naphthalin und Methylnaphthalin.",
    "10: PAK16: stellvertretend für die Gruppe der polyzyklischen aromatischen Kohlenwasserstoffe (PAK) werden nach der Liste der US-amerikanischen Umweltbehörde, Environmental Protection Agency (EPA), 16 ausgewählte PAK untersucht: Acenaphthen, Acenaphthylen, Anthracen, Benzo[a]anthracen, Benzo[a]pyren, Benzo[b]fluoranthen, Benzo[g,h,i]perylen, Benzo[k]fluoranthen, Chrysen, Dibenzo[a,h]anthracen, Fluoranthen, Fluoren, Indeno[1,2,3- cd]pyren, Naphthalin, Phenanthren und Pyren.",
    "11: Bei Überschreitung der Werte sind die Materialien auf fallspezifische Belastungen zu untersuchen.",
    "12: Bei Quecksilber und Thallium ist für die Klassifizierung (BM-F0* bis BM-F3) der Gesamtgehalt maßgeblich. Der Eluatwert für BM-0* ist einzuhalten."
]

CLASS_RANKS = {"BM-0": 0, "BM-0 (Eluat n. maßgeblich)": 0, "BM-0*": 1, "> BM-0* (Eluat; für BM-F nur Feststoff maßgeblich)": 1, "BM-F0*": 2, "BM-F1": 3, "BM-F2": 4, "BM-F3": 5, "> BM-F3 (Deponie!)": 6, "Kein Messwert": -1, "Kein Messwert (< BG)": -1}

def clean_float_string(val):
    if pd.isna(val) or val is None: return ""
    val_str = str(val)
    def round_match(match):
        num = float(match.group(0))
        return f"{num:.4f}".rstrip('0').rstrip('.')
    cleaned = re.sub(r'\d+\.\d{5,}', round_match, val_str)
    return html.escape(cleaned)

def create_combined_report(sheet_dict, output_dir, original_filename, bodenart):
    if not os.path.exists(output_dir): os.makedirs(output_dir)
        
    base_name = os.path.splitext(original_filename)[0]
    
    # 1. SETUP EXCEL
    excel_path = os.path.join(output_dir, f"Auswertung_{base_name}.xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # 2. SETUP PDF
    pdf_path = os.path.join(output_dir, f"Auswertung_{base_name}.pdf")
    doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A3), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    pdf_elements = []
    
    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]
    style_normal.fontSize = 10
    style_normal.leading = 12
    style_bold = styles["Normal"].clone("Bold")
    style_bold.fontName = "Helvetica-Bold"
    style_bold.fontSize = 10
    style_title = styles["Heading3"].clone("Title")
    style_title.textColor = colors.HexColor("#9C0006")
    style_title.fontSize = 11
    style_title.leading = 13
    style_fn = styles["Normal"].clone("Footnote")
    style_fn.fontSize = 8
    style_fn.leading = 10
    
    # 3. SETUP HTML
    html_path = os.path.join(output_dir, f"Auswertung_{base_name}.html")
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <meta charset="utf-8"/>
    <style>
        body {{ font-family: Helvetica, Arial, sans-serif; font-size: 12px; color: #000; line-height: 1.3; margin: 20px; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 15px; margin-bottom: 30px; }}
        th, td {{ border: 1px solid #a0a0a0; padding: 6px; text-align: left; vertical-align: middle; }}
        th {{ background-color: #e0e0e0; font-weight: bold; border-bottom: 2px solid #555; }}
        .header-red {{ color: #9C0006; font-weight: bold; margin-top: 5px; margin-bottom: 10px; }}
        .header-bold {{ font-weight: bold; font-size: 16px; margin-top: 40px; padding-bottom: 5px; border-bottom: 2px solid #000; }}
        .row-green {{ background-color: #C6EFCE; color: #006100; }}
        .row-blue {{ background-color: #DDEBF7; color: #2E75B6; }}
        .row-yellow {{ background-color: #FFEB9C; color: #9C5700; }}
        .row-red {{ background-color: #FFC7CE; color: #9C0006; }}
        .row-gray {{ background-color: #F2F2F2; color: #7A7A7A; }}
        .summary {{ font-size: 15px; font-weight: bold; margin-top: 15px; margin-bottom: 15px; }}
        .footnotes {{ font-size: 10px; line-height: 1.4; color: #333; margin-bottom: 50px; }}
        .page-break {{ page-break-after: always; }}
    </style>
    </head>
    <body>
    """
    
    sheet_count = 0
    for sheet_name, df in sheet_dict.items():
        sheet_count += 1
        max_rank = -1
        worst_class = "BM-0"
        
        for idx, row in df.iterrows():
            val = str(row.get("Eingestufte Klasse", ""))
            param = str(row.get("Parameter", ""))
            rank = CLASS_RANKS.get(val, -1)
            is_orientierungswert = param in ["pH-Wert", "Elektrische Leitfähigkeit", "TOC"]
            if rank > max_rank and not is_orientierungswert:
                max_rank = rank
                worst_class = val
                
        # --- EXCEL ---
        df_excel = df.drop(columns=["Format_Italic", "Format_Bold_Fn"], errors='ignore')
        ws = wb.create_sheet(title=str(sheet_name)[:31])
        
        ws.cell(row=1, column=1).value = f"PROBE: {sheet_name} | Referenz-Bodenart: '{bodenart}' | Datum: {datetime.date.today().isoformat()}"
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        ws.cell(row=2, column=1).value = f"RECHTLICHER HINWEIS: Automatisierte Vorprüfung. Ersetzt keine gutachterliche Freigabe. Grundlage: {EBV_VERSION['gesetz']}."
        ws.cell(row=2, column=1).font = Font(bold=True, color="9C0006")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        
        for r_idx, row_data in enumerate(dataframe_to_rows(df_excel, index=False, header=True), 4):
            for c_idx, value in enumerate(row_data, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
                
        for row_num in range(5, ws.max_row + 1):
            val = str(ws.cell(row=row_num, column=4).value)
            
            df_idx = row_num - 5
            if df_idx < len(df):
                is_italic = bool(df.iloc[df_idx].get("Format_Italic", False))
                is_bold_fn = bool(df.iloc[df_idx].get("Format_Bold_Fn", False))
            else:
                is_italic, is_bold_fn = False, False
                
            fill_to_use = None
            font_color = "000000"
            if "> BM-F3" in val: fill_to_use, font_color = fill_red, "9C0006"
            elif "BM-F" in val: fill_to_use, font_color = fill_yellow, "9C5700"
            elif "> BM-0*" in val or "BM-0*" in val: fill_to_use, font_color = fill_blue, "2E75B6"
            elif "BM-0" in val: fill_to_use, font_color = fill_green, "006100"
            elif "Nicht in EBV" in val or "Kein" in val: fill_to_use, font_color = fill_gray, "7A7A7A"
                
            if fill_to_use:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_num, column=col).fill = fill_to_use
                    is_bold = True if (col == 6 and is_bold_fn) else False
                    ws.cell(row=row_num, column=col).font = Font(color=font_color, italic=is_italic, bold=is_bold)

        summary_row = ws.max_row + 2
        ws.cell(row=summary_row, column=1).value = "GESAMTEINSTUFUNG DER PROBE (Worst-Case):"
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=4).value = worst_class
        ws.cell(row=summary_row, column=4).font = Font(bold=True, size=12)
        
        if max_rank <= 0: ws.cell(row=summary_row, column=4).fill = fill_green
        elif max_rank == 1: ws.cell(row=summary_row, column=4).fill = fill_blue
        elif max_rank <= 5: ws.cell(row=summary_row, column=4).fill = fill_yellow
        elif max_rank == 6: ws.cell(row=summary_row, column=4).fill = fill_red

        start_row = summary_row + 3
        ws.cell(row=start_row, column=1).value = "Regelbezüge & Fußnoten (Anlage 1 Tabelle 3 EBV):"
        ws.cell(row=start_row, column=1).font = Font(bold=True)
        for i, text in enumerate(FUSSNOTEN_TEXTE):
            cell = ws.cell(row=start_row + 1 + i, column=1)
            cell.value = text
            ws.merge_cells(start_row=start_row + 1 + i, start_column=1, end_row=start_row + 1 + i, end_column=6)
            
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 25

        # --- PDF & HTML Datenaufbereitung ---
        if sheet_count > 1:
            pdf_elements.append(PageBreak())
            html_content += "<div class='page-break'></div>"
            
        pdf_elements.append(Paragraph(f"<b>PROBE: {sheet_name} | Referenz-Bodenart: '{bodenart}' | Datum: {datetime.date.today().isoformat()}</b>", style_bold))
        pdf_elements.append(Paragraph(f"RECHTLICHER HINWEIS: Automatisierte Vorprüfung. Ersetzt keine gutachterliche Freigabe. Grundlage: {EBV_VERSION['gesetz']}.", style_title))
        pdf_elements.append(Spacer(1, 10))
        
        html_content += f"""
            <div class="header-bold">PROBE: {sheet_name} | Referenz-Bodenart: '{bodenart}' | Datum: {datetime.date.today().isoformat()}</div>
            <div class="header-red">RECHTLICHER HINWEIS: Automatisierte Vorprüfung. Ersetzt keine gutachterliche Freigabe. Grundlage: {EBV_VERSION['gesetz']}.</div>
            <table>
                <thead>
                    <tr>
                        <th width="30%">Parameter</th><th width="12%">Einheit</th><th width="15%">Messwert</th><th width="20%">Eingestufte Klasse</th><th width="15%">Maßgeblicher GW</th><th width="8%">Fußnote</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        pdf_data = [[Paragraph("<b>Parameter</b>", style_bold), Paragraph("<b>Einheit</b>", style_bold), Paragraph("<b>Messwert</b>", style_bold), Paragraph("<b>Eingestufte Klasse</b>", style_bold), Paragraph("<b>Maßgeblicher GW</b>", style_bold), Paragraph("<b>Fußnote</b>", style_bold)]]
        bg_colors = []
        
        for idx, row in df.iterrows():
            val = str(row.get("Eingestufte Klasse", ""))
            is_italic = bool(row.get("Format_Italic", False))
            is_bold_fn = bool(row.get("Format_Bold_Fn", False))
            
            row_color = colors.white
            cls = ""
            if "> BM-F3" in val: row_color, cls = colors.HexColor("#FFC7CE"), "row-red"
            elif "BM-F" in val: row_color, cls = colors.HexColor("#FFEB9C"), "row-yellow"
            elif "> BM-0*" in val or "BM-0*" in val: row_color, cls = colors.HexColor("#DDEBF7"), "row-blue"
            elif "BM-0" in val: row_color, cls = colors.HexColor("#C6EFCE"), "row-green"
            elif "Nicht in EBV" in val or "Kein" in val: row_color, cls = colors.HexColor("#F2F2F2"), "row-gray"
            bg_colors.append(row_color)

            p_val = clean_float_string(row.get('Parameter',''))
            e_val = clean_float_string(row.get('Einheit',''))
            m_val = clean_float_string(row.get('Messwert',''))
            k_val = clean_float_string(val)
            gw_val = clean_float_string(row.get('Maßgeblicher GW',''))
            fn_val = clean_float_string(row.get('Fußnote',''))

            # Wende HTML-Tags an, die von Browser und PDF-Renderer verstanden werden
            if is_italic:
                p_val, e_val, m_val, k_val, gw_val = f"<i>{p_val}</i>", f"<i>{e_val}</i>", f"<i>{m_val}</i>", f"<i>{k_val}</i>", f"<i>{gw_val}</i>"
            if is_bold_fn:
                fn_val = f"<b>{fn_val}</b>"
            elif is_italic:
                fn_val = f"<i>{fn_val}</i>"

            pdf_data.append([
                Paragraph(p_val, style_normal), Paragraph(e_val, style_normal), Paragraph(m_val, style_normal),
                Paragraph(k_val, style_normal), Paragraph(gw_val, style_normal), Paragraph(fn_val, style_normal)
            ])
            
            html_content += f"<tr class='{cls}'><td>{p_val}</td><td>{e_val}</td><td>{m_val}</td><td>{k_val}</td><td>{gw_val}</td><td>{fn_val}</td></tr>"
            
        t = Table(pdf_data, colWidths=[350, 120, 140, 250, 150, 100], repeatRows=1)
        t_style = [('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#e0e0e0")), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('BOTTOMPADDING', (0, 0), (-1, -1), 2), ('TOPPADDING', (0, 0), (-1, -1), 2)]
        for i, color in enumerate(bg_colors): t_style.append(('BACKGROUND', (0, i+1), (-1, i+1), color))
        t.setStyle(TableStyle(t_style))
        pdf_elements.append(t)
        pdf_elements.append(Spacer(1, 10))
        
        if max_rank <= 0: sum_color, sum_color_cls = colors.HexColor("#C6EFCE"), "row-green"
        elif max_rank == 1: sum_color, sum_color_cls = colors.HexColor("#DDEBF7"), "row-blue"
        elif max_rank <= 5: sum_color, sum_color_cls = colors.HexColor("#FFEB9C"), "row-yellow"
        else: sum_color, sum_color_cls = colors.HexColor("#FFC7CE"), "row-red"
        
        sum_data = [[Paragraph("<b>GESAMTEINSTUFUNG DER PROBE (Worst-Case):</b>", style_bold), Paragraph(f"<b>{worst_class}</b>", style_bold)]]
        sum_t = Table(sum_data, colWidths=[300, 150])
        sum_t.setStyle(TableStyle([('BACKGROUND', (1,0), (1,0), sum_color), ('BOX', (1,0), (1,0), 1, colors.grey), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        pdf_elements.append(sum_t)
        
        pdf_elements.append(Spacer(1, 10))
        pdf_elements.append(Paragraph("<b>Regelbezüge & Fußnoten (Anlage 1 Tabelle 3 EBV):</b>", style_bold))
        pdf_elements.append(Paragraph(" | ".join(FUSSNOTEN_TEXTE), style_fn))
        
        html_content += f"""
                </tbody>
            </table>
            <div class="summary">
                GESAMTEINSTUFUNG DER PROBE (Worst-Case): <span class="{sum_color_cls}">{worst_class}</span>
            </div>
            <div class="footnotes">
                <b>Regelbezüge & Fußnoten (Anlage 1 Tabelle 3 EBV):</b><br/>
                {" | ".join(FUSSNOTEN_TEXTE)}
            </div>
        """

    wb.save(excel_path)
    
    html_content += "</body></html>"
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)
        
    try:
        doc.build(pdf_elements)
        print(f"  -> BERICHTE GENERIERT IN: {output_dir}")
        print(f"     - Auswertung_Alle_Proben.xlsx (Alle Proben in Tabs)")
        print(f"     - Auswertung_Alle_Proben.pdf (Alle Proben fortlaufend)")
        print(f"     - Auswertung_Alle_Proben.html")
    except Exception as e:
        print(f"  -> FEHLER beim PDF-Export: {e}")