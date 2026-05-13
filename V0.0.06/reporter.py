"""
EBV Tool v05 — unified reporter.

Produces two output sets in a single call to :func:`create_combined_report`:

1. **Legacy v04 outputs** (renamed with ``_OLD_Design`` suffix):
   - ``Evaluation_<base>_OLD_Design.xlsx``
   - ``Evaluation_<base>_OLD_Design.pdf``
   - ``Evaluation_<base>_OLD_Design.html``

2. **New v05 output** matching the company Mantelverordnung template:
   - ``Evaluation_<base>.pdf`` — Feststoff + Eluat sheets only,
     visually identical to printing those tabs from the company
     Excel template.

The new format requires project metadata + per-sample metadata that
``step1_extraktion.py`` writes into the validation Excel
(``_Project`` sheet + per-sample columns ``Probenbezeichnung``,
``Petrographische_Beschreibung``, ``Stratigraphie``, ``Labor_Nummer``).
When metadata is missing the new PDF is still produced; missing
fields render as empty cells.
"""
from __future__ import annotations

import datetime
import html
import os
import re
import shutil
import subprocess
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from config import EBV_VERSION, ebv_tabelle_3

# ---------------------------------------------------------------------------
# Legal footnotes (unchanged from v04)
# ---------------------------------------------------------------------------

FOOTNOTES_EN: List[str] = [
    "1: Die Materialwerte gelten für Bodenmaterial und Baggergut mit bis zu 10 Volumenprozent (BM und BG) oder bis zu 50 Volumenprozent (BM-F und BG-F) mineralischer Fremdbestandteile im Sinne von § 2 Nummer 8 der Bundes-Bodenschutz- und Altlastenverordnung mit nur vernachlässigbaren Anteilen an Störstoffen im Sinne von § 2 Nummer 9 der Bundes-Bodenschutz- und Altlastenverordnung. Bodenmaterial der Klasse BM-0 und Baggergut der Klasse BG-0 erfüllen die wertebezogenen Anforderungen an das Auf- oder Einbringen gemäß § 7 Absatz 3 der Bundes-Bodenschutz- und Altlastenverordnung. Bodenmaterial der Klasse BM-0 und Baggergut der Klasse BG-0 Sand erfüllen die wertebezogenen Anforderungen an das Auf- oder Einbringen gemäß § 8 Absatz 2 der Bundes-Bodenschutz- und Altlastenverordnung; Bodenmaterial der Klasse BM-0* und Baggergut der Klasse BG-0* erfüllen die wertebezogenen Anforderungen an das Auf- oder Einbringen gemäß § 8 Absatz 3 Nummer 1 der Bundes-Bodenschutz- und Altlastenverordnung.",
    "2: Bodenarten-Hauptgruppen gemäß Bodenkundlicher Kartieranleitung, 5. Auflage, Hannover 2005 (KA5); stark schluffige Sande, lehmig-schluffige Sande und stark lehmige Sande sowie Materialien, die nicht bodenartspezifisch zugeordnet werden können, sind entsprechend der Bodenart Lehm, Schluff zu bewerten.",
    "3: Die Eluatwerte in Spalte 6 sind mit Ausnahme des Eluatwertes für Sulfat nur maßgeblich, wenn für den betreffenden Stoff der jeweilige Feststoffwert nach Spalte 3 bis 5 überschritten wird. Der Eluatwert für PAK15 und Napthalin und Methylnaphtaline, gesamt, ist maßgeblich, wenn der Feststoffwert für PAK16 nach Spalte 3 bis 5 überschritten wird. Die in Klammern genannten Werte gelten jeweils bei einem TOC-Gehalt von ≥ 0,5 %.",
    "4: Stoffspezifischer Orientierungswert; bei Abweichungen ist die Ursache zu prüfen.",
    "5: Bei Überschreitung des Wertes ist die Ursache zu prüfen. Handelt es sich um naturbedingt erhöhte Sulfatkonzentrationen, ist eine Verwertung innerhalb der betroffenen Gebiete möglich. Außerhalb dieser Gebiete ist über die Verwertungseignung im Einzelfall und in Abstimmung mit der zuständigen Behörde zu entscheiden.",
    "6: Der Wert 1 mg/kg gilt für Sand und Lehm/Schluff. Für Ton gilt der Wert 1,5 mg/kg.",
    "7: Bodenmaterialspezifischer Orientierungswert. Bei heterogenen Bodenverhältnissen mineralischer Böden kann der TOC-Gehalt der Masse des anfallenden Materials als maßgeblich bei Verwertung im Umfeld des anfallenden Materials und Verwendung unter gleichen Bedingungen herangezogen werden. Beim Einbau sind Volumenbeständigkeit und Setzungsprozesse sowie die Vorgaben von § 6 Absatz 11 Satz 2 und 3 der Bundes-Bodenschutz- und Altlastenverordnung zu berücksichtigen.",
    "8: Die angegebenen Werte gelten für Kohlenwasserstoffverbindungen mit einer Kettenlänge von C10 bis C22. Der Gesamtgehalt bestimmt nach der DIN EN 14039, „Charakterisierung von Abfällen – Bestimmung des Gehalts an Kohlenwasserstoffen von C10 bis C40 mittels Gaschromatographie\", Ausgabe Januar 2005 darf insgesamt den in Klammern genannten Wert nicht überschreiten.",
    "9: PAK15: PAK16 ohne Naphthalin und Methylnaphthalin.",
    "10: PAK16: stellvertretend für die Gruppe der polyzyklischen aromatischen Kohlenwasserstoffe (PAK) werden nach der Liste der US-amerikanischen Umweltbehörde, Environmental Protection Agency (EPA), 16 ausgewählte PAK untersucht: Acenaphthen, Acenaphthylen, Anthracen, Benzo[a]anthracen, Benzo[a]pyren, Benzo[b]fluoranthen, Benzo[g,h,i]perylen, Benzo[k]fluoranthen, Chrysen, Dibenzo[a,h]anthracen, Fluoranthen, Fluoren, Indeno[1,2,3- cd]pyren, Naphthalin, Phenanthren und Pyren.",
    "11: Bei Überschreitung der Werte sind die Materialien auf fallspezifische Belastungen zu untersuchen.",
    "12: Bei Quecksilber und Thallium ist für die Klassifizierung (BM-F0* bis BM-F3) der Gesamtgehalt maßgeblich. Der Eluatwert für BM-0* ist einzuhalten.",
]

CLASS_RANKS: Dict[str, int] = {
    "BM-0": 0,
    "BM-0 (Eluat n. maßgeblich)": 0,
    "BM-0*": 1,
    "> BM-0* (Eluat; für BM-F nur Feststoff maßgeblich)": 1,
    "> BM-0* (höhere Klassen nicht definiert; übergeordneter Parameter maßgeblich)": 2,
    "BM-F0* (Eluat cap; PAK16 Feststoff für höhere Klassen maßgeblich)": 2,
    "BM-F0*": 2,
    "BM-F1": 3,
    "BM-F2": 4,
    "BM-F3": 5,
    "> BM-F3 (Landfill!)": 6,
    "Not in EBV": -1,
    "No Value": -1,
    "No Value (< LOQ)": -1,
}

# Inverse map for rendering: rank -> short label used in company template.
RANK_TO_LABEL: Dict[int, str] = {
    0: "BM-0",
    1: "BM-0*",
    2: "BM-F0*",
    3: "BM-F1",
    4: "BM-F2",
    5: "BM-F3",
    6: ">BM-F3",
}

# ---------------------------------------------------------------------------
# Company-template column maps (v05)
# ---------------------------------------------------------------------------

FIRST_SAMPLE_ROW: int = 43

#: Feststoff sheet: EBV parameter (Matrix) -> Excel column letter.
FESTSTOFF_PARAM_TO_COL: Dict[Tuple[str, str], str] = {
    ("Mineralische Fremdbestandteile", "Feststoff"): "J",
    ("TOC", "Feststoff"): "K",
    ("pH-Wert", "Eluat"): "L",  # template shows pH on both sheets
    # M = Cyanide -- not in EBV config, left blank
    ("EOX", "Feststoff"): "N",
    # O = paired KW-Index (C10-C22 / C10-C40) -- handled specially
    ("PAK16", "Feststoff"): "P",
    ("Benzo(a)pyren", "Feststoff"): "Q",
    # R = BTEX, S = LHKW -- not in EBV config
    ("PCB6 und PCB-118", "Feststoff"): "T",
    ("Arsen", "Feststoff"): "U",
    ("Blei", "Feststoff"): "V",
    ("Cadmium", "Feststoff"): "W",
    ("Chrom, gesamt", "Feststoff"): "X",
    ("Kupfer", "Feststoff"): "Y",
    ("Nickel", "Feststoff"): "Z",
    ("Quecksilber", "Feststoff"): "AA",
    ("Thallium", "Feststoff"): "AB",
    ("Zink", "Feststoff"): "AC",
    # AD = Tributylzinn-Kationen -- not in EBV config
}

#: Eluat sheet: EBV parameter (Matrix) -> Excel column letter.
#: Layout mirrors the user's master Mantelverordnung template — TOC
#: column inserted at K so reviewers can see which samples are above
#: the 0.5 % TOC band that flips the Klammerwerte thresholds.
ELUAT_PARAM_TO_COL: Dict[Tuple[str, str], str] = {
    ("Mineralische Fremdbestandteile", "Feststoff"): "J",
    ("TOC", "Feststoff"): "K",                   # visualisation aid (TOC band)
    ("pH-Wert", "Eluat"): "L",
    ("Elektrische Leitfähigkeit", "Eluat"): "M",
    ("Sulfat", "Eluat"): "N",
    ("PAK15", "Eluat"): "O",
    ("Arsen", "Eluat"): "P",
    ("Blei", "Eluat"): "Q",
    ("Cadmium", "Eluat"): "R",
    ("Chrom, gesamt", "Eluat"): "S",
    ("Kupfer", "Eluat"): "T",
    ("Nickel", "Eluat"): "U",
    ("Quecksilber", "Eluat"): "V",
    ("Thallium", "Eluat"): "W",
    ("Zink", "Eluat"): "X",
    ("Naphthalin und Methylnaphthaline, gesamt", "Eluat"): "Y",
    ("PCB6 und PCB-118", "Eluat"): "Z",
    # AA=Fluorid, AB=DOC, AC=Chlorid, AD..AS=Tabelle-4 params -- not in EBV config
}

#: Column O on Feststoff sheet = "KW-Index (MKW)" — paired C10-C22 (C10-C40).
FESTSTOFF_MKW_COL: str = "O"

# Mapping from internal class names (with underscores) to template labels.
BM_SUBTYPE_LABEL: Dict[str, str] = {
    "BM_0_Sand": "BM-0 Sand",
    "BM_0_Lehm_Schluff": "BM-0 Lehm, Schluff",
    "BM_0_Ton": "BM-0 Ton",
}

LIBREOFFICE_BIN: str = "soffice"


# ---------------------------------------------------------------------------
# Public data contracts (v05 metadata)
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ProjectMeta:
    """Project-level header data read from the validation Excel ``_Project`` sheet."""

    projektnummer: str = ""
    bauvorhaben: str = ""
    los: str = ""
    bauwerk: str = ""


@dataclass(frozen=True)
class SampleMeta:
    """Per-sample metadata read from the validation Excel sample sheet."""

    probenbezeichnung: str = ""
    petrographische_beschreibung: str = ""
    stratigraphie: str = ""
    labor_nummer: str = ""


# ---------------------------------------------------------------------------
# Utilities used by both legacy and new outputs
# ---------------------------------------------------------------------------


def clean_float_string(val: Any, allow_html: bool = False) -> str:
    """Clean floating-point formatting bugs and process HTML tags safely.

    Args:
        val: input value (may be float, str, None, or NaN).
        allow_html: if True, preserve ``<b>...</b>`` tags after escaping.

    Returns:
        Cleaned string suitable for table cells.
    """
    if pd.isna(val) or val is None:
        return ""
    val_str = str(val)

    def round_match(match: re.Match) -> str:
        return f"{float(match.group(0)):.4f}".rstrip("0").rstrip(".")

    cleaned = re.sub(r"\d+\.\d{5,}", round_match, val_str)

    if allow_html:
        escaped = html.escape(cleaned)
        return escaped.replace("&lt;b&gt;", "<b>").replace("&lt;/b&gt;", "</b>")
    # Strip HTML tags BEFORE escaping so plain-text cells don't leak
    # "&lt;b&gt;" into the workbook (the previous order escaped the tags
    # first, then the replace looked for unescaped "<b>" and found none).
    stripped = re.sub(r"</?b>", "", cleaned)
    return html.escape(stripped)


def _de_decimal(s: str) -> str:
    """Convert dot decimal separator to German comma, preserving non-numeric chars."""
    return s.replace(".", ",")


def _strip_html(s: str) -> str:
    """Remove HTML tags from a string (for the v05 plain-text rendering)."""
    return re.sub(r"<[^>]+>", "", s)


def _format_messwert_for_template(messwert_raw: Any) -> str:
    """Format an evaluator Messwert string for company-template display.

    Conversions applied:
        * NaN / empty -> ""
        * ``"< 50"`` (operator + value) -> ``"<50"`` (no space)
        * ``"< BG"`` -> ``"n.n."`` (template convention for non-detect)
        * Float precision artifacts (≥5 decimals) -> 4 sig figs
        * Decimal dot -> German comma
        * Strips HTML

    Args:
        messwert_raw: raw value from the evaluator output column ``Messwert``.

    Returns:
        Display-ready string.
    """
    if messwert_raw is None or (isinstance(messwert_raw, float) and pd.isna(messwert_raw)):
        return ""
    s = str(messwert_raw).strip()
    if not s:
        return ""
    s = _strip_html(s)
    if s.lower() == "< bg" or s.lower() == "<bg":
        return "n.n."

    # Round long-precision floats: "0.030000000000000002" -> "0.03"
    def _round_match(match: re.Match) -> str:
        return f"{float(match.group(0)):.4f}".rstrip("0").rstrip(".")

    s = re.sub(r"\d+\.\d{5,}", _round_match, s)

    # Compact "< 50" -> "<50", "> 50" -> ">50"
    s = re.sub(r"^([<>])\s+", r"\1", s)
    s = _de_decimal(s)
    return s


def _is_writable(cell: Any) -> bool:
    """Return True if cell is writable (skip merged anchors)."""
    return not isinstance(cell, MergedCell)


def _set_cell(ws: Worksheet, coord: str, value: Any) -> None:
    """Assign value to cell at coord, skipping merged anchors."""
    cell = ws[coord]
    if not _is_writable(cell):
        return
    cell.value = value


def _set_class_label_cell(ws: Worksheet, coord: str, klasse_template: str) -> None:
    """Write a class label with the footnote-2 superscript suffix on BM-0
    soil-subtype classes (BM-0 Sand / Lehm, Schluff / Ton).

    The user's manual workbook formats these labels as rich text — e.g.
    ``"BM-0 Lehm, Schluff²"`` — with footnote 2 (which describes the
    soil-type-aware threshold convention). Other classes (BM-0*, BM-F0*…
    BM-F3) carry no footnote suffix.

    Args:
        ws: target worksheet.
        coord: cell coordinate (e.g. ``"H44"``).
        klasse_template: pre-formatted joint label, e.g.
            ``"BM-0 Lehm, Schluff, BG-0 Lehm, Schluff"``.
    """
    cell = ws[coord]
    if not _is_writable(cell):
        return
    label = klasse_template or ""
    # Only BM-0 soil-subtype labels carry footnote 2. Other classes pass
    # through as plain text.
    if "BM-0 Sand" not in label and "BM-0 Lehm" not in label and "BM-0 Ton" not in label:
        cell.value = label
        return
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        # The joint format is "BM-0 <Soil>, BG-0 <Soil>". Append a
        # superscript "2" after EACH "BM-0 <Soil>" and "BG-0 <Soil>" chunk
        # (matches the manual). Split on ", " to find both occurrences,
        # then re-join with the superscript inserted between each chunk
        # and its trailing comma.
        normal = InlineFont(rFont="Calibri", sz=8.0)
        sup = InlineFont(rFont="Arial", sz=8.0, vertAlign="superscript")
        parts: list = []
        chunks = label.split(", BG-0 ", 1)
        if len(chunks) == 2:
            bm_part = chunks[0]
            bg_part = "BG-0 " + chunks[1]
            parts.append(TextBlock(font=normal, text=bm_part))
            parts.append(TextBlock(font=sup, text="2"))
            parts.append(TextBlock(font=normal, text=", " + bg_part))
            parts.append(TextBlock(font=sup, text="2"))
        else:
            parts.append(TextBlock(font=normal, text=label))
            parts.append(TextBlock(font=sup, text="2"))
        cell.value = CellRichText(parts)
    except Exception:
        # If rich-text support is unavailable for any reason, fall back
        # to a plain-text "Label²" approximation.
        cell.value = label.replace(", BG-0 ", "², BG-0 ") + "²"


#: Classification -> hex fill color for the H ("Zuordnung") cell.
#: Palette taken from the user's master Mantelverordnung workbook
#: (Distelhausen reference, 2026-05-13). Green->blue->purple gradient
#: with lighter shades for lower-risk classes.
CLASS_FILL_HEX: Dict[str, str] = {
    "bm0_sand":  "D6E3BC",   # BM-0 Sand
    "bm0_lehm":  "C2D69B",   # BM-0 Lehm, Schluff
    "bm0_ton":   "76923C",   # BM-0 Ton
    "bm0star":   "DAEEF3",   # BM-0*, BG-0*3
    "bmf0star":  "95B3D7",   # BM-F0*
    "bmf1":      "CCC0D9",   # BM-F1
    "bmf2":      "B2A1C7",   # BM-F2
    "bmf3":      "5F497A",   # BM-F3 (and >BM-F3)
    "white":     "FFFFFF",
}


def _classification_color_key(klasse_template: str) -> str:
    """Pick fill-color key from a template/evaluator class label.

    Resolves via CLASS_RANKS first so suffixed labels ("> BM-0* (höhere
    Klassen nicht definiert...)" / "BM-0 (Eluat n. maßgeblich)" /
    "BM-F0* (Eluat cap; ...)") map to the right colour band rather than
    catching on a generic ">" or "BM-F3" substring.
    """
    s = klasse_template or ""
    rank = CLASS_RANKS.get(s, None)
    if rank is None:
        # Try without parenthetical suffix
        head = s.split("(", 1)[0].strip().rstrip(",")
        rank = CLASS_RANKS.get(head, None)
    if rank is not None:
        if rank >= 6: return "bmf3"
        if rank == 5: return "bmf3"
        if rank == 4: return "bmf2"
        if rank == 3: return "bmf1"
        if rank == 2: return "bmf0star"
        if rank == 1: return "bm0star"
        if rank == 0:
            if "BM-0 Ton" in s: return "bm0_ton"
            if "BM-0 Lehm" in s: return "bm0_lehm"
            return "bm0_sand"
    # Fallback substring resolution for joint labels ("BM-F2, BG-F2" etc.)
    if "BM-F3" in s: return "bmf3"
    if "BM-F2" in s: return "bmf2"
    if "BM-F1" in s: return "bmf1"
    if "BM-F0*" in s: return "bmf0star"
    if "BM-0*" in s or "BG-0*" in s: return "bm0star"
    if "BM-0 Ton" in s: return "bm0_ton"
    if "BM-0 Lehm" in s: return "bm0_lehm"
    if "BM-0" in s: return "bm0_sand"
    if s.startswith(">"): return "bmf3"
    return "white"


def _copy_style_from_reference(
    ws: Worksheet,
    target_row: int,
    reference_row: int,
    last_col_letter: str,
) -> None:
    """Copy font / alignment / border / number format from a reference row.

    This is used to give every sample row the same look as the first
    sample slot (row 43) in the skeleton, which has the company's
    intended sample-row styling. The cell *fill* is intentionally NOT
    copied — we want a clean canvas before reapplying the classification
    color on the H cell.

    Args:
        ws: target worksheet.
        target_row: row index to style.
        reference_row: row index whose style is copied (typically ``FIRST_SAMPLE_ROW``).
        last_col_letter: last column letter to style (e.g. ``"AD"`` or ``"AS"``).
    """
    from copy import copy as _copy
    from openpyxl.utils import column_index_from_string

    last_col_idx = column_index_from_string(last_col_letter)
    for col_idx in range(1, last_col_idx + 1):
        col_letter = get_column_letter(col_idx)
        ref_cell = ws[f"{col_letter}{reference_row}"]
        tgt_cell = ws[f"{col_letter}{target_row}"]
        if isinstance(tgt_cell, MergedCell) or isinstance(ref_cell, MergedCell):
            continue
        if ref_cell.has_style:
            tgt_cell.font = _copy(ref_cell.font)
            tgt_cell.alignment = _copy(ref_cell.alignment)
            tgt_cell.border = _copy(ref_cell.border)
            tgt_cell.number_format = ref_cell.number_format
            # Intentionally skip fill — sample rows start with no fill, classification color applied separately
            tgt_cell.fill = PatternFill(fill_type=None)
            tgt_cell.protection = _copy(ref_cell.protection)

    # Mirror row height too so MP05 doesn't stand out
    if reference_row in ws.row_dimensions and ws.row_dimensions[reference_row].height:
        ws.row_dimensions[target_row].height = ws.row_dimensions[reference_row].height


def _apply_classification_fill(ws: Worksheet, row_idx: int, klasse_template: str) -> None:
    """Color the H cell (Zuordnung) by classification.

    Args:
        ws: target worksheet.
        row_idx: sample row.
        klasse_template: pre-formatted class label e.g. ``"BM-F2, BG-F2"``.
    """
    color_key = _classification_color_key(klasse_template)
    hex_color = CLASS_FILL_HEX[color_key]
    cell = ws[f"H{row_idx}"]
    if not _is_writable(cell):
        return
    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


# ---------------------------------------------------------------------------
# Split-class computation
# ---------------------------------------------------------------------------


def _classify_split(
    df: pd.DataFrame,
    bodenart_internal: str,
) -> Tuple[str, str, str]:
    """Compute Feststoff, Eluat, and Gesamt class labels for one sample.

    Args:
        df: evaluator output DataFrame (long format, one row per parameter).
        bodenart_internal: internal soil-subtype name from config
            (``"BM_0_Sand"`` / ``"BM_0_Lehm_Schluff"`` / ``"BM_0_Ton"``).

    Returns:
        Tuple ``(feststoff_label, eluat_label, gesamt_label)`` with template
        labels formatted as ``"BM-F2, BG-F2"`` (joint BM/BG) or
        ``">BM-F3, BG-F3"`` etc., with BM-0 soil-subtype suffix when applicable.
    """
    # Build lookup of EBV parameter -> Matrix to partition rows correctly
    matrix_lookup: Dict[str, List[str]] = {}
    for item in ebv_tabelle_3:
        matrix_lookup.setdefault(item["parameter"], []).append(item["typ"])

    # Eluat → Feststoff cross-reference for FN3 evaluation. Most Eluat
    # parameters share the name with their Feststoff counterpart; PAK15
    # and Naphthalin map to PAK16 (per ELUAT_FESTSTOFF_XREF).
    try:
        from evaluator import ELUAT_FESTSTOFF_XREF as _XREF
    except Exception:
        _XREF = {"PAK15": "PAK16", "Naphthalin und Methylnaphthaline, gesamt": "PAK16"}

    # First pass: per-param Feststoff status (needed for FN3 at Gesamt).
    feststoff_status: Dict[str, str] = {}
    for _, row in df.iterrows():
        param = str(row.get("Parameter", ""))
        einheit = str(row.get("Einheit", ""))
        klasse = str(row.get("Eingestufte Klasse", ""))
        u = einheit.lower()
        is_feststoff = ("kg" in u) or ("vol" in u) or ("m%" in u) or (
            "%" in u and "kg" not in u and "/" not in u
        )
        if is_feststoff:
            feststoff_status[param] = klasse

    feststoff_rank: int = -1
    eluat_rank: int = -1
    eluat_gesamt_rank: int = -1  # Eluat rank that ACTUALLY drives Gesamt (FN3 + orientation filter)

    # Orientation parameters: classify strictly per-cell, but their rank
    # is NOT decisive for the Gesamt class. pH, el. Leitfähigkeit and
    # TOC are all Hilfsparameter — the user reviews the cell-colour
    # diagnostic and decides whether to bracket.
    ORIENTATION_PARAMS = {"pH-Wert", "Elektrische Leitfähigkeit", "TOC"}

    # Feststoff-side paired parameters: C10-C40 is the "Gesamtgehalt" of
    # the KW-Index family. When the C10-C22 partner (the un-bracketed
    # value per FN8) is BM-0, the C10-C40 Klammerwert is NOT decisive
    # for class escalation. Mirrors the FN3 logic but for a Feststoff/
    # Feststoff pair.
    FESTSTOFF_PAIRED: Dict[str, str] = {
        "Kohlenwasserstoffe (C10-C40)": "Kohlenwasserstoffe (C10-C22)",
    }

    for _, row in df.iterrows():
        param = str(row.get("Parameter", ""))
        einheit = str(row.get("Einheit", ""))
        klasse = str(row.get("Eingestufte Klasse", ""))
        rank = CLASS_RANKS.get(klasse, -1)

        u = einheit.lower()
        is_feststoff = ("kg" in u) or ("vol" in u) or ("m%" in u) or (
            "%" in u and "kg" not in u and "/" not in u
        )

        if is_feststoff:
            # Feststoff rank: skip orientation params AND skip the C10-C40
            # Klammerwert when the C10-C22 partner sits at BM-0 (per user
            # convention: brackets-only doesn't push class).
            if param in ORIENTATION_PARAMS:
                continue
            paired = FESTSTOFF_PAIRED.get(param)
            if paired and feststoff_status.get(paired) in ("BM-0", "No Value (< LOQ)"):
                continue
            if rank > feststoff_rank:
                feststoff_rank = rank
        else:
            # Eluat sheet rank includes orientation drivers (pH, Leitf.)
            # so the H cell on the Eluat sheet shows the strict threshold
            # class (e.g. Leitf 966 → BM-F3).
            if rank > eluat_rank:
                eluat_rank = rank
            # Gesamt-side Eluat rank: skip orientation drivers AND apply
            # FN3 — if the corresponding Feststoff parameter is BM-0
            # (or No Value (< LOQ)), this Eluat value is "nicht
            # maßgeblich" for the Gesamt class. Sulfat is exempt per FN3.
            if param in ORIENTATION_PARAMS:
                continue
            xref = _XREF.get(param, param)
            if param != "Sulfat" and feststoff_status.get(xref) in ("BM-0", "No Value (< LOQ)"):
                continue
            if rank > eluat_gesamt_rank:
                eluat_gesamt_rank = rank

    gesamt_rank = max(feststoff_rank, eluat_gesamt_rank)

    def rank_to_template_label(rank: int) -> str:
        if rank <= 0:
            # BM-0 carries soil-subtype suffix
            return BM_SUBTYPE_LABEL.get(bodenart_internal, "BM-0 Sand")
        return RANK_TO_LABEL.get(rank, "BM-0")

    def joint_label(rank: int) -> str:
        """Append BG twin (joint BM/BG label as in company template)."""
        bm_label = rank_to_template_label(rank)
        # BM-0 Sand -> "BM-0 Sand, BG-0 Sand"; BM-F2 -> "BM-F2, BG-F2"; ">BM-F3" -> ">BM-F3, BG-F3"
        if bm_label.startswith(">"):
            return f"{bm_label}, {bm_label.replace('BM', 'BG', 1)[1:].lstrip()}".rstrip()  # not used
        if bm_label.startswith("BM-0"):
            twin = bm_label.replace("BM-0", "BG-0", 1)
            return f"{bm_label}, {twin}"
        twin = bm_label.replace("BM", "BG", 1)
        if bm_label.startswith(">"):
            twin = bm_label[1:].replace("BM", "BG", 1)
            return f"{bm_label}, >{twin}".replace(">>", ">")
        return f"{bm_label}, {twin}"

    # Special-case > BM-F3 to keep '>' on BG twin too
    def joint_label_v2(rank: int) -> str:
        if rank == 6:
            return ">BM-F3, >BG-F3"
        return joint_label(rank)

    return (
        joint_label_v2(feststoff_rank if feststoff_rank >= 0 else 0),
        joint_label_v2(eluat_rank if eluat_rank >= 0 else 0),
        joint_label_v2(gesamt_rank if gesamt_rank >= 0 else 0),
    )


# ---------------------------------------------------------------------------
# Legacy v04 output (renamed to _OLD_Design)
# ---------------------------------------------------------------------------


def _generate_old_design_reports(
    sheet_dict: Dict[str, pd.DataFrame],
    output_dir: str,
    base_name: str,
    bodenart: str,
) -> None:
    """Generate the v04-style Excel + PDF + HTML with ``_OLD_Design`` suffix.

    Layout, styling, footnote text are byte-identical to v04 output.
    Only the output filenames carry the new suffix.
    """
    excel_path = os.path.join(output_dir, f"Evaluation_{base_name}_OLD_Design.xlsx")
    pdf_path = os.path.join(output_dir, f"Evaluation_{base_name}_OLD_Design.pdf")
    html_path = os.path.join(output_dir, f"Evaluation_{base_name}_OLD_Design.html")

    wb = Workbook()
    wb.remove(wb.active)

    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A3),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20,
    )
    pdf_elements: List[Any] = []
    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]
    style_normal.fontSize, style_normal.leading = 9.5, 11
    style_bold = styles["Normal"].clone("Bold")
    style_bold.fontName, style_bold.fontSize = "Helvetica-Bold", 9.5
    style_title = styles["Heading3"].clone("Title")
    style_title.textColor, style_title.fontSize, style_title.leading = (
        colors.HexColor("#9C0006"),
        10.5,
        12,
    )
    style_fn = styles["Normal"].clone("Footnote")
    style_fn.fontSize, style_fn.leading = 7.5, 9

    html_content = f"""<!DOCTYPE html><html><head><meta charset="utf-8"/><style>
        body {{ font-family: Helvetica, Arial, sans-serif; font-size: 12px; line-height: 1.3; margin: 20px; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 10px; margin-bottom: 10px; }}
        th, td {{ border: 1px solid #a0a0a0; padding: 4px; text-align: left; vertical-align: middle; }}
        th {{ background-color: #e0e0e0; font-weight: bold; border-bottom: 2px solid #555; }}
        .header-red {{ color: #9C0006; font-weight: bold; margin-bottom: 10px; }}
        .header-bold {{ font-weight: bold; font-size: 16px; margin-top: 20px; border-bottom: 2px solid #000; }}
        .row-green {{ background-color: #C6EFCE; color: #006100; }}
        .row-blue {{ background-color: #DDEBF7; color: #2E75B6; }}
        .row-yellow {{ background-color: #FFEB9C; color: #9C5700; }}
        .row-red {{ background-color: #FFC7CE; color: #9C0006; }}
        .row-gray {{ background-color: #F2F2F2; color: #7A7A7A; }}
        .summary {{ font-size: 14px; font-weight: bold; margin-top: 10px; }}
        .disclaimer {{ font-size: 11px; color: #9C0006; margin-bottom: 10px; font-style: italic; }}
        .footnotes {{ font-size: 10px; color: #333; margin-bottom: 30px; }}
        .page-break {{ page-break-after: always; }}
    </style></head><body>"""

    sheet_count = 0
    for sheet_name, df in sheet_dict.items():
        sheet_count += 1
        max_rank = -1
        worst_class = "BM-0"
        for _, row in df.iterrows():
            val = str(row.get("Eingestufte Klasse", ""))
            param = str(row.get("Parameter", ""))
            rank = CLASS_RANKS.get(val, -1)
            if rank > max_rank and param not in {"pH-Wert", "Elektrische Leitfähigkeit"}:
                max_rank = rank
                worst_class = val

        df_excel = df.drop(columns=["Format_Italic"], errors="ignore")
        df_excel = df_excel.rename(
            columns={
                "Einheit": "Unit",
                "Messwert": "Measured Value",
                "Eingestufte Klasse": "Classified Category",
                "Maßgeblicher GW": "Applicable Limit",
                "Fußnote": "Footnote",
            }
        )

        ws = wb.create_sheet(title=str(sheet_name)[:31])
        ws.cell(row=1, column=1).value = (
            f"SAMPLE: {sheet_name} | Reference Soil Type: '{bodenart}' | "
            f"Date: {datetime.date.today().isoformat()}"
        )
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        ws.cell(row=2, column=1).value = (
            f"LEGAL DISCLAIMER: Automated preliminary check. Does not replace expert "
            f"approval. Basis: {EBV_VERSION['law']}."
        )
        ws.cell(row=2, column=1).font = Font(bold=True, color="9C0006")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

        for r_idx, row_data in enumerate(dataframe_to_rows(df_excel, index=False, header=True), 4):
            for c_idx, value in enumerate(row_data, 1):
                clean_val = clean_float_string(value, allow_html=False) if isinstance(value, str) else value
                ws.cell(row=r_idx, column=c_idx, value=clean_val)

        for row_num in range(5, ws.max_row + 1):
            val = str(ws.cell(row=row_num, column=4).value)
            df_idx = row_num - 5
            is_italic = bool(df.iloc[df_idx].get("Format_Italic", False)) if df_idx < len(df) else False
            fill_to_use, f_color = None, "000000"
            # Use the user-template colour palette per the parameter's own
            # class. Orientation params (pH, el. Leitfähigkeit) still
            # carry the class colour of THEIR own row so reviewers can see
            # the threshold breach; the orientation status is enforced
            # at the Zusammenfassung Gesamt stage, not by hiding cells.
            if "> BM-F3" in val or "BM-F3" in val:
                fill_to_use, f_color = PatternFill(start_color=CLASS_FILL_HEX["bmf3"], end_color=CLASS_FILL_HEX["bmf3"], fill_type="solid"), "FFFFFF"
            elif "BM-F2" in val:
                fill_to_use, f_color = PatternFill(start_color=CLASS_FILL_HEX["bmf2"], end_color=CLASS_FILL_HEX["bmf2"], fill_type="solid"), "000000"
            elif "BM-F1" in val:
                fill_to_use, f_color = PatternFill(start_color=CLASS_FILL_HEX["bmf1"], end_color=CLASS_FILL_HEX["bmf1"], fill_type="solid"), "000000"
            elif "BM-F0*" in val:
                fill_to_use, f_color = PatternFill(start_color=CLASS_FILL_HEX["bmf0star"], end_color=CLASS_FILL_HEX["bmf0star"], fill_type="solid"), "000000"
            elif "> BM-0*" in val or "BM-0*" in val:
                fill_to_use, f_color = PatternFill(start_color=CLASS_FILL_HEX["bm0star"], end_color=CLASS_FILL_HEX["bm0star"], fill_type="solid"), "000000"
            elif "BM-0 Ton" in val:
                fill_to_use, f_color = PatternFill(start_color=CLASS_FILL_HEX["bm0_ton"], end_color=CLASS_FILL_HEX["bm0_ton"], fill_type="solid"), "FFFFFF"
            elif "BM-0 Lehm" in val:
                fill_to_use, f_color = PatternFill(start_color=CLASS_FILL_HEX["bm0_lehm"], end_color=CLASS_FILL_HEX["bm0_lehm"], fill_type="solid"), "000000"
            elif "BM-0" in val:
                fill_to_use, f_color = PatternFill(start_color=CLASS_FILL_HEX["bm0_sand"], end_color=CLASS_FILL_HEX["bm0_sand"], fill_type="solid"), "000000"
            elif "Not in EBV" in val or "No Value" in val:
                fill_to_use, f_color = fill_gray, "7A7A7A"

            if fill_to_use:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_num, column=col).fill = fill_to_use
                    ws.cell(row=row_num, column=col).font = Font(color=f_color, italic=is_italic)

        summary_row = ws.max_row + 2
        ws.cell(row=summary_row, column=1).value = "OVERALL SAMPLE CLASSIFICATION (Worst-Case):"
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=summary_row, column=4).value = worst_class
        ws.cell(row=summary_row, column=4).font = Font(bold=True, size=12)

        if max_rank <= 0:
            ws.cell(row=summary_row, column=4).fill = fill_green
        elif max_rank == 1:
            ws.cell(row=summary_row, column=4).fill = fill_blue
        elif max_rank <= 5:
            ws.cell(row=summary_row, column=4).fill = fill_yellow
        elif max_rank == 6:
            ws.cell(row=summary_row, column=4).fill = fill_red

        ws.cell(row=summary_row + 1, column=1).value = (
            "NOTE: Orientation values (pH, Electrical Conductivity, TOC) do not affect "
            "the automated worst-case classification."
        )
        ws.cell(row=summary_row + 1, column=1).font = Font(italic=True, color="9C0006")

        ws.cell(row=summary_row + 3, column=1).value = "Regulations & Footnotes (Appendix 1 Table 3 EBV):"
        ws.cell(row=summary_row + 3, column=1).font = Font(bold=True)
        for i, text in enumerate(FOOTNOTES_EN):
            ws.cell(row=summary_row + 4 + i, column=1).value = text
            ws.merge_cells(
                start_row=summary_row + 4 + i,
                start_column=1,
                end_row=summary_row + 4 + i,
                end_column=6,
            )

        # Per-sample sheet column widths tuned so the wide text fits without
        # truncation: Parameter (long names like "Naphthalin und
        # Methylnaphthaline, gesamt"), Classified Category (which now also
        # carries the "Eluat cap" suffix), and Applicable Limit (5+
        # threshold values separated by " / ") all need extra room.
        _column_widths = {1: 48, 2: 20, 3: 22, 4: 38, 5: 38, 6: 14}
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = (
                _column_widths.get(col_idx, 25)
            )
        # Wrap text on header + body rows so long classification labels
        # display on multiple lines rather than spilling into neighbour cells.
        from openpyxl.styles import Alignment as _Align
        for r in range(4, ws.max_row + 1):
            ws.row_dimensions[r].height = 22
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cur = cell.alignment
                cell.alignment = _Align(
                    horizontal=cur.horizontal or "left",
                    vertical="center",
                    wrap_text=True,
                )

        if sheet_count > 1:
            pdf_elements.append(PageBreak())
            html_content += "<div class='page-break'></div>"

        pdf_elements.append(
            Paragraph(
                f"<b>SAMPLE: {sheet_name} | Reference Soil Type: '{bodenart}' | "
                f"Date: {datetime.date.today().isoformat()}</b>",
                style_bold,
            )
        )
        pdf_elements.append(
            Paragraph(
                f"LEGAL DISCLAIMER: Automated preliminary check. Does not replace "
                f"expert approval. Basis: {EBV_VERSION['law']}.",
                style_title,
            )
        )
        pdf_elements.append(Spacer(1, 5))

        pdf_data: List[List[Paragraph]] = [
            [
                Paragraph("<b>Parameter</b>", style_bold),
                Paragraph("<b>Unit</b>", style_bold),
                Paragraph("<b>Measured Value</b>", style_bold),
                Paragraph("<b>Classified Category</b>", style_bold),
                Paragraph("<b>Applicable Limit</b>", style_bold),
                Paragraph("<b>Footnote</b>", style_bold),
            ]
        ]
        bg_colors: List[Any] = []

        html_content += f"""
            <div class="header-bold">SAMPLE: {sheet_name} | Reference Soil Type: '{bodenart}' | Date: {datetime.date.today().isoformat()}</div>
            <div class="header-red">LEGAL DISCLAIMER: Automated preliminary check. Does not replace expert approval. Basis: {EBV_VERSION['law']}.</div>
            <table><thead><tr>
            <th width="30%">Parameter</th><th width="12%">Unit</th><th width="15%">Measured Value</th><th width="20%">Classified Category</th><th width="15%">Applicable Limit</th><th width="8%">Footnote</th>
            </tr></thead><tbody>"""

        for _, row in df.iterrows():
            val = str(row.get("Eingestufte Klasse", ""))
            is_italic = bool(row.get("Format_Italic", False))

            row_color, cls = colors.white, ""
            # Match the user's Mantelverordnung palette. Orientation params
            # stay coloured by their own class; the orientation filter is
            # applied at the Zusammenfassung-Gesamt stage.
            if "> BM-F3" in val or "BM-F3" in val:
                row_color, cls = colors.HexColor("#5F497A"), "row-bmf3"
            elif "BM-F2" in val:
                row_color, cls = colors.HexColor("#B2A1C7"), "row-bmf2"
            elif "BM-F1" in val:
                row_color, cls = colors.HexColor("#CCC0D9"), "row-bmf1"
            elif "BM-F0*" in val:
                row_color, cls = colors.HexColor("#95B3D7"), "row-bmf0star"
            elif "> BM-0*" in val or "BM-0*" in val:
                row_color, cls = colors.HexColor("#DAEEF3"), "row-bm0star"
            elif "BM-0 Ton" in val:
                row_color, cls = colors.HexColor("#76923C"), "row-bm0ton"
            elif "BM-0 Lehm" in val:
                row_color, cls = colors.HexColor("#C2D69B"), "row-bm0lehm"
            elif "BM-0" in val:
                row_color, cls = colors.HexColor("#D6E3BC"), "row-bm0sand"
            elif "Not in EBV" in val or "No Value" in val:
                row_color, cls = colors.HexColor("#F2F2F2"), "row-gray"
            bg_colors.append(row_color)

            p_val = clean_float_string(row.get("Parameter", ""), allow_html=True)
            e_val = clean_float_string(row.get("Einheit", ""), allow_html=True)
            m_val = clean_float_string(row.get("Messwert", ""), allow_html=True)
            k_val = clean_float_string(val, allow_html=True)
            gw_val = clean_float_string(row.get("Maßgeblicher GW", ""), allow_html=True)
            fn_val = clean_float_string(row.get("Fußnote", ""), allow_html=True)

            if is_italic:
                p_val = f"<i>{p_val}</i>"
                e_val = f"<i>{e_val}</i>"
                m_val = f"<i>{m_val}</i>"
                k_val = f"<i>{k_val}</i>"
                gw_val = f"<i>{gw_val}</i>"
                fn_val = f"<i>{fn_val}</i>"

            pdf_data.append(
                [
                    Paragraph(p_val, style_normal),
                    Paragraph(e_val, style_normal),
                    Paragraph(m_val, style_normal),
                    Paragraph(k_val, style_normal),
                    Paragraph(gw_val, style_normal),
                    Paragraph(fn_val, style_normal),
                ]
            )
            html_content += (
                f"<tr class='{cls}'><td>{p_val}</td><td>{e_val}</td><td>{m_val}</td>"
                f"<td>{k_val}</td><td>{gw_val}</td><td>{fn_val}</td></tr>"
            )

        t = Table(pdf_data, colWidths=[270, 100, 120, 220, 300, 80], repeatRows=1)
        t_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0e0e0")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
            ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ]
        for i, color in enumerate(bg_colors):
            t_style.append(("BACKGROUND", (0, i + 1), (-1, i + 1), color))
        t.setStyle(TableStyle(t_style))
        pdf_elements.append(t)
        pdf_elements.append(Spacer(1, 5))

        if max_rank <= 0:
            sum_color, sum_color_cls = colors.HexColor("#C6EFCE"), "row-green"
        elif max_rank == 1:
            sum_color, sum_color_cls = colors.HexColor("#DDEBF7"), "row-blue"
        elif max_rank <= 5:
            sum_color, sum_color_cls = colors.HexColor("#FFEB9C"), "row-yellow"
        else:
            sum_color, sum_color_cls = colors.HexColor("#FFC7CE"), "row-red"

        sum_data = [
            [
                Paragraph("<b>OVERALL SAMPLE CLASSIFICATION (Worst-Case):</b>", style_bold),
                Paragraph(f"<b>{worst_class}</b>", style_bold),
            ]
        ]
        sum_t = Table(sum_data, colWidths=[300, 150])
        sum_t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (1, 0), (1, 0), sum_color),
                    ("BOX", (1, 0), (1, 0), 1, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        pdf_elements.append(sum_t)

        style_disclaimer = styles["Normal"].clone("Disclaimer")
        style_disclaimer.textColor = colors.HexColor("#9C0006")
        style_disclaimer.fontName = "Helvetica-Oblique"
        pdf_elements.append(
            Paragraph(
                "NOTE: Orientation values (pH, Electrical Conductivity, TOC) do not "
                "affect the automated worst-case classification.",
                style_disclaimer,
            )
        )
        pdf_elements.append(Spacer(1, 5))

        pdf_elements.append(Paragraph("<b>Regulations & Footnotes (Appendix 1 Table 3 EBV):</b>", style_bold))
        pdf_elements.append(Paragraph(" | ".join(FOOTNOTES_EN), style_fn))

        html_content += f"""</tbody></table>
            <div class="summary">OVERALL SAMPLE CLASSIFICATION (Worst-Case): <span class="{sum_color_cls}">{worst_class}</span></div>
            <div class="disclaimer">NOTE: Orientation values (pH, Electrical Conductivity, TOC) do not affect the automated worst-case classification according to EBV.</div>
            <div class="footnotes"><b>Regulations & Footnotes (Appendix 1 Table 3 EBV):</b><br/>{" | ".join(FOOTNOTES_EN)}</div>
        """

    wb.save(excel_path)
    html_content += "</body></html>"
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    try:
        doc.build(pdf_elements)
    except Exception as e:
        print(f"  -> ERROR during OLD-design PDF Export: {e}")


# ---------------------------------------------------------------------------
# New v05 output (company Mantelverordnung format)
# ---------------------------------------------------------------------------


def _build_messwert_lookup(df: pd.DataFrame) -> Dict[Tuple[str, str], str]:
    """Build a (parameter, matrix) -> formatted-Messwert lookup for one sample.

    Matrix is inferred from the unit string:
        * contains 'kg', 'vol', 'm%' (without '/') -> Feststoff
        * otherwise -> Eluat
    """
    out: Dict[Tuple[str, str], str] = {}
    for _, row in df.iterrows():
        param = str(row.get("Parameter", "")).strip()
        einheit = str(row.get("Einheit", "")).lower()
        if "kg" in einheit or "vol" in einheit or einheit.strip() in {"m%", "%"}:
            matrix = "Feststoff"
        else:
            matrix = "Eluat"
        out[(param, matrix)] = _format_messwert_for_template(row.get("Messwert", ""))
    return out


def _mineralische_fb_label(klasse_template: str) -> str:
    """Map a Feststoff template-class label to the Mineralische FB display."""
    return "bis 50" if "BM-F" in klasse_template or "BG-F" in klasse_template else "bis 10"


def _write_sample_row_v05(
    ws: Worksheet,
    row_idx: int,
    sample_meta: SampleMeta,
    klasse_template: str,
    messwerte: Dict[Tuple[str, str], str],
    column_map: Dict[Tuple[str, str], str],
    is_feststoff_sheet: bool,
) -> None:
    """Write one sample row into a Feststoff or Eluat sheet of the skeleton.

    Args:
        ws: target worksheet.
        row_idx: 1-indexed row to write.
        sample_meta: per-sample metadata.
        klasse_template: pre-formatted class label e.g. ``"BM-F2, BG-F2"``.
        messwerte: ``(parameter, matrix) -> display-formatted value``.
        column_map: ``(parameter, matrix) -> column letter``.
        is_feststoff_sheet: True for Feststoff (handles paired MKW), False for Eluat.
    """
    _set_cell(ws, f"A{row_idx}", sample_meta.probenbezeichnung)
    _set_cell(ws, f"B{row_idx}", sample_meta.petrographische_beschreibung)
    _set_cell(ws, f"C{row_idx}", sample_meta.stratigraphie)
    _set_cell(ws, f"D{row_idx}", sample_meta.labor_nummer)
    # Render the class label with the footnote-2 superscript suffix that
    # the manual workbook uses on BM-0 Sand/Lehm/Ton rows. Footnote 2
    # documents the soil-type-aware threshold system, so it only applies
    # to the BM-0 + soil-subtype classes (NOT to BM-0* / BM-F*).
    _set_class_label_cell(ws, f"H{row_idx}", klasse_template)
    _set_cell(ws, f"J{row_idx}", _mineralische_fb_label(klasse_template))

    for (param, matrix), col in column_map.items():
        if param == "Mineralische Fremdbestandteile":
            continue  # already written above
        value = messwerte.get((param, matrix), "")
        if value:
            _set_cell(ws, f"{col}{row_idx}", value)

    # Special handling: Feststoff sheet column O = "MKW" = paired C10-C22 (C10-C40)
    if is_feststoff_sheet:
        c22 = messwerte.get(("Kohlenwasserstoffe (C10-C22)", "Feststoff"), "")
        c40 = messwerte.get(("Kohlenwasserstoffe (C10-C40)", "Feststoff"), "")
        if c22 or c40:
            paired = f"{c22} ({c40})" if (c22 and c40) else (c22 or c40)
            _set_cell(ws, f"{FESTSTOFF_MKW_COL}{row_idx}", paired)


def _replace_header_tokens(ws: Worksheet, project: ProjectMeta) -> None:
    """Replace ``{{...}}`` placeholders in the skeleton header area.

    Beside the simple ``{{PROJEKTNUMMER}}`` / ``{{BAUVORHABEN}}`` /
    ``{{LOS}}`` / ``{{BAUWERK}}`` substitutions, the company-template
    header has two composite cells that bundle multiple fields:

      * ``{{LOS}}`` (row 4) renders as ``"LOS X"`` (or empty when no LOS).
      * ``{{BAUVORHABEN_LINE}}`` (row 5) renders as
        ``"<Bauvorhaben> (BW <Bauwerk>)"``.
    """
    los_disp = f"LOS {project.los}" if project.los and project.los != "No Value" else (project.los or "")
    if project.bauvorhaben and project.bauwerk and project.bauwerk != "No Value":
        bauline = f"{project.bauvorhaben} (BW {project.bauwerk})"
    elif project.bauvorhaben:
        bauline = project.bauvorhaben
    else:
        bauline = ""
    mapping: Dict[str, str] = {
        "{{PROJEKTNUMMER}}": project.projektnummer or "",
        "{{BAUVORHABEN}}": project.bauvorhaben or "",
        "{{LOS}}": los_disp,
        "{{BAUWERK}}": project.bauwerk or "",
        "{{BAUVORHABEN_LINE}}": bauline,
    }
    # Header cells in the company template: A2, A3, A4, A5 plus a few sheet-specific spots
    for row in ws.iter_rows(min_row=1, max_row=8):
        for cell in row:
            if not _is_writable(cell):
                continue
            if isinstance(cell.value, str) and cell.value in mapping:
                cell.value = mapping[cell.value]



def _drivers_for_split(
    df: pd.DataFrame,
) -> Tuple[List[str], List[str], int, int]:
    """Identify driving parameters for Feststoff and Eluat worst-case ranks.

    A driver is a parameter whose individual class rank equals the matrix's
    max rank — i.e. the parameter that pushed the sample to its worst-case
    classification. Orientation parameters (pH, Leitfähigkeit, TOC) are
    excluded.

    Args:
        df: evaluator output DataFrame (long format).

    Returns:
        ``(feststoff_drivers, eluat_drivers, feststoff_rank, eluat_rank)``
        with the ranks of each matrix. Ranks of -1 mean nothing exceeded
        BM-0 in that matrix.
    """
    feststoff_rank: int = -1
    eluat_rank: int = -1

    for _, row in df.iterrows():
        param = str(row.get("Parameter", ""))
        einheit = str(row.get("Einheit", ""))
        klasse = str(row.get("Eingestufte Klasse", ""))
        if param in {"pH-Wert", "Elektrische Leitfähigkeit"}:
            continue
        rank = CLASS_RANKS.get(klasse, -1)
        u = einheit.lower()
        is_feststoff = ("kg" in u) or ("vol" in u) or ("m%" in u) or ("%" in u and "kg" not in u and "/" not in u)
        if is_feststoff:
            if rank > feststoff_rank:
                feststoff_rank = rank
        else:
            if rank > eluat_rank:
                eluat_rank = rank

    feststoff_drivers: List[str] = []
    eluat_drivers: List[str] = []
    for _, row in df.iterrows():
        param = str(row.get("Parameter", ""))
        einheit = str(row.get("Einheit", ""))
        klasse = str(row.get("Eingestufte Klasse", ""))
        if param in {"pH-Wert", "Elektrische Leitfähigkeit"}:
            continue
        rank = CLASS_RANKS.get(klasse, -1)
        if rank <= 0:
            continue
        u = einheit.lower()
        is_feststoff = ("kg" in u) or ("vol" in u) or ("m%" in u) or ("%" in u and "kg" not in u and "/" not in u)
        if is_feststoff and rank == feststoff_rank:
            feststoff_drivers.append(param)
        elif not is_feststoff and rank == eluat_rank:
            eluat_drivers.append(param)

    return feststoff_drivers, eluat_drivers, feststoff_rank, eluat_rank


def _apply_per_cell_class_fills(
    ws: Worksheet,
    row_idx: int,
    df: pd.DataFrame,
    col_map: Dict[Tuple[str, str], str],
    is_feststoff_sheet: bool,
) -> None:
    """Colour every parameter cell on the sample row with the fill that
    matches its OWN per-parameter class — green for BM-0, light-blue for
    BM-0*, purple shades for BM-F1..F3 etc. — matching the user's master
    Mantelverordnung workbook.

    The H-cell class is still applied separately via
    :func:`_apply_classification_fill`; this function paints the parameter
    cells so reviewers can see at a glance WHICH parameter drove the row
    into a given class (including orientation parameters like
    Leitfähigkeit, which contribute to per-cell colour but are excluded
    from the Zusammenfassung Gesamt).

    Args:
        ws: target worksheet.
        row_idx: sample row.
        df: evaluator output (long format, one row per parameter).
        col_map: param→column letter mapping for this sheet.
        is_feststoff_sheet: True for Feststoff (handles paired MKW column).
    """
    # Build a (param_name, matrix) → classification class label lookup so
    # we can colour each cell by ITS OWN class. Matrix is derived from
    # the unit string, matching the convention used elsewhere in the
    # reporter ("kg"/"vol"/"m%" → Feststoff, otherwise Eluat).
    klassen: Dict[Tuple[str, str], str] = {}
    for _, row in df.iterrows():
        param = str(row.get("Parameter", ""))
        einheit = str(row.get("Einheit", ""))
        klasse = str(row.get("Eingestufte Klasse", ""))
        u = einheit.lower()
        is_feststoff_row = ("kg" in u) or ("vol" in u) or ("m%" in u) or (
            "%" in u and "kg" not in u and "/" not in u
        )
        matrix = "Feststoff" if is_feststoff_row else "Eluat"
        klassen[(param, matrix)] = klasse

    for (param, matrix), col in col_map.items():
        klasse = klassen.get((param, matrix), "")
        if not klasse or klasse in ("Not in EBV", "No Value"):
            continue
        # Only paint BM-class cells. "BM-0 (Eluat n. maßgeblich)" and
        # similar suffixed labels still resolve cleanly via the colour-key
        # function below.
        color_key = _classification_color_key(klasse)
        if color_key == "white":
            continue
        hex_color = CLASS_FILL_HEX[color_key]
        cell = ws[f"{col}{row_idx}"]
        if not _is_writable(cell):
            continue
        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    # Paired KW-Index column on Feststoff: colour by the higher of the
    # C10-C22 / C10-C40 row classes.
    if is_feststoff_sheet:
        ranks = []
        for p_name in ("Kohlenwasserstoffe (C10-C22)", "Kohlenwasserstoffe (C10-C40)"):
            kk = klassen.get((p_name, "Feststoff"), "")
            r = CLASS_RANKS.get(kk, -1)
            if r >= 0:
                ranks.append((r, kk))
        if ranks:
            ranks.sort(reverse=True)
            worst_klasse = ranks[0][1]
            color_key = _classification_color_key(worst_klasse)
            if color_key != "white":
                hex_color = CLASS_FILL_HEX[color_key]
                cell = ws[f"{FESTSTOFF_MKW_COL}{row_idx}"]
                if _is_writable(cell):
                    cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _apply_driver_fills(
    ws: Worksheet,
    row_idx: int,
    drivers: List[str],
    klasse_label: str,
    col_map: Dict[Tuple[str, str], str],
    is_feststoff_sheet: bool,
) -> None:
    """Colour each driving-parameter cell with the row's class fill.

    Makes it visually obvious WHICH parameter drove the worst-case class.

    Args:
        ws: target worksheet.
        row_idx: sample row.
        drivers: list of EBV parameter names that drove the rank.
        klasse_label: row's class label, used to pick the fill colour.
        col_map: param→column letter mapping for this sheet.
        is_feststoff_sheet: True for Feststoff (handles paired MKW column).
    """
    color_key = _classification_color_key(klasse_label)
    hex_color = CLASS_FILL_HEX[color_key]
    paint = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    for driver in drivers:
        for (param, _matrix), col in col_map.items():
            if param == driver:
                cell = ws[f"{col}{row_idx}"]
                if _is_writable(cell):
                    cell.fill = paint
                break
        if is_feststoff_sheet and driver in {
            "Kohlenwasserstoffe (C10-C22)",
            "Kohlenwasserstoffe (C10-C40)",
        }:
            cell = ws[f"{FESTSTOFF_MKW_COL}{row_idx}"]
            if _is_writable(cell):
                cell.fill = paint


#: Per-parameter column mapping on the Zusammenfassung sheet (row 10 header).
#: Mirrors the company workbook layout. Only the EBV parameters we evaluate
#: are listed here; non-EBV columns (Tabelle 4 Sb/Mo/V, herbicides, BTEX,
#: LHKW, Cyanide, etc.) stay blank.
ZUSAMMEN_PARAM_TO_COL: Dict[Tuple[str, str], str] = {
    ("Mineralische Fremdbestandteile", "Feststoff"): "L",
    ("pH-Wert", "Eluat"): "M",
    ("Elektrische Leitfähigkeit", "Eluat"): "N",
    ("Sulfat", "Eluat"): "P",
    ("PAK15", "Eluat"): "S",
    ("PAK16", "Feststoff"): "T",
    ("Arsen", "Feststoff"): "V",
    ("Blei", "Feststoff"): "W",
    ("Cadmium", "Feststoff"): "X",
    ("Chrom, gesamt", "Feststoff"): "Y",
    ("Kupfer", "Feststoff"): "Z",
    ("Nickel", "Feststoff"): "AB",
    ("Zink", "Feststoff"): "AD",
    ("Arsen", "Eluat"): "AM",
    ("Blei", "Eluat"): "AN",
    ("Cadmium", "Eluat"): "AO",
    ("Chrom, gesamt", "Eluat"): "AP",
    ("Kupfer", "Eluat"): "AQ",
    ("Nickel", "Eluat"): "AR",
    ("Quecksilber", "Feststoff"): "AS",
    ("Quecksilber", "Eluat"): "AT",
    ("Thallium", "Feststoff"): "AU",
    ("Thallium", "Eluat"): "AV",
    ("Zink", "Eluat"): "AW",
    ("TOC", "Feststoff"): "AX",
    ("Benzo(a)pyren", "Feststoff"): "AZ",
    ("Naphthalin und Methylnaphthaline, gesamt", "Eluat"): "BA",
    ("PCB6 und PCB-118", "Feststoff"): "BB",
    ("PCB6 und PCB-118", "Eluat"): "BC",
    ("EOX", "Feststoff"): "BD",
}

ZUSAMMEN_FIRST_SAMPLE_ROW: int = 41
ZUSAMMEN_LAST_COL: str = "O"


def _categorise_above_bm0(
    df: pd.DataFrame,
    feststoff_status: Dict[str, str],
) -> Tuple[List[Tuple[str, str]], List[Tuple[str, str]], List[Tuple[str, str]], List[Tuple[str, str]]]:
    """Categorise parameters that exceed BM-0 into decisive vs ignored
    per matrix.

    A parameter is "above BM-0" iff its individual rank > 0. It is
    "decisive" iff it is allowed to drive the Gesamt class; otherwise
    "ignored" (orientation parameter, FN3 cross-reference where the
    Feststoff partner is BM-0, or KW C10-C40 with C10-C22 at BM-0).

    Args:
        df: evaluator output DataFrame (long format).
        feststoff_status: param -> klasse lookup for Feststoff rows.

    Returns:
        Tuple ``(fest_decisive, fest_ignored, elu_decisive, elu_ignored)``,
        each a list of ``(param_name, displayed_value)`` tuples.
    """
    try:
        from evaluator import ELUAT_FESTSTOFF_XREF as _XREF
    except Exception:
        _XREF = {"PAK15": "PAK16", "Naphthalin und Methylnaphthaline, gesamt": "PAK16"}

    ORIENTATION = {"pH-Wert", "Elektrische Leitfähigkeit", "TOC"}
    FESTSTOFF_PAIRED = {
        "Kohlenwasserstoffe (C10-C40)": "Kohlenwasserstoffe (C10-C22)",
    }

    fest_decisive: List[Tuple[str, str]] = []
    fest_ignored: List[Tuple[str, str]] = []
    elu_decisive: List[Tuple[str, str]] = []
    elu_ignored: List[Tuple[str, str]] = []

    for _, row in df.iterrows():
        param = str(row.get("Parameter", ""))
        einheit = str(row.get("Einheit", ""))
        klasse = str(row.get("Eingestufte Klasse", ""))
        rank = CLASS_RANKS.get(klasse, -1)
        if rank <= 0:
            continue  # at or below BM-0 → don't report
        messwert = str(row.get("Messwert", "")).strip()
        # Strip the operator from the display (it's redundant with "<…")
        display = messwert if messwert else klasse

        u = einheit.lower()
        is_feststoff = ("kg" in u) or ("vol" in u) or ("m%" in u) or (
            "%" in u and "kg" not in u and "/" not in u
        )

        if is_feststoff:
            if param in ORIENTATION:
                fest_ignored.append((param, display))
            elif (paired := FESTSTOFF_PAIRED.get(param)) and feststoff_status.get(paired) in ("BM-0", "No Value (< LOQ)"):
                fest_ignored.append((param, display))
            else:
                fest_decisive.append((param, display))
        else:
            if param in ORIENTATION:
                elu_ignored.append((param, display))
            elif param != "Sulfat":
                xref = _XREF.get(param, param)
                if feststoff_status.get(xref) in ("BM-0", "No Value (< LOQ)"):
                    elu_ignored.append((param, display))
                else:
                    elu_decisive.append((param, display))
            else:
                elu_decisive.append((param, display))

    return fest_decisive, fest_ignored, elu_decisive, elu_ignored


def _format_rel_schadstoffe(
    fest_decisive: List[Tuple[str, str]],
    fest_ignored: List[Tuple[str, str]],
    elu_decisive: List[Tuple[str, str]],
    elu_ignored: List[Tuple[str, str]],
) -> str:
    """Format the 'Relevante Schadstoffe > BM-0' summary cell.

    Always emits a ``Feststoff:`` and an ``Eluat:`` line. Decisive
    drivers are listed plain; ignored drivers (orientation, FN3, KW
    pairing) are wrapped in parentheses so the reviewer sees that the
    parameter was considered but excluded from the Gesamt.
    """
    def join_one(decisive: List[Tuple[str, str]], ignored: List[Tuple[str, str]]) -> str:
        parts: List[str] = [f"{p} {v}".strip() for p, v in decisive]
        parts += [f"({p} {v})".strip() for p, v in ignored]
        return "; ".join(parts) if parts else "-"

    return (
        "Feststoff: " + join_one(fest_decisive, fest_ignored)
        + "\nEluat: " + join_one(elu_decisive, elu_ignored)
    )


def _populate_zusammenfassung(
    ws: Worksheet,
    row_idx: int,
    sample_meta: SampleMeta,
    feststoff_label: str,
    eluat_label: str,
    gesamt_label: str,
    feststoff_drivers: List[str],
    eluat_drivers: List[str],
    messwerte: Dict[Tuple[str, str], str],
) -> None:
    """Write one sample row into the "Zusammenfassung Dekklaration" sheet.

    Matches the company deliverable summary layout (page 3 of
    ``A_4_3_1_Auswertung_Labor.pdf``):

        A: Probenbezeichnung
        B: Petrographische Beschreibung
        C: Stratigraphie
        D: Labor-Nummer
        I: Zuordnung Feststoff (colour-coded by class)
        K: Zuordnung Eluat (colour-coded)
        M: Zuordnung Gesamt (colour-coded)
        O: Relevante Schadstoffe > BM-0 — free-text summary of the
           parameters (with measured value + unit) that pushed the sample
           above BM-0. Format mirrors the company example:
           ``"Feststoff: PAK16 8,6 mg/kg; Eluat: PAK15 1,71 µg/L"``.

    Args:
        ws: target Zusammenfassung Dekklaration worksheet.
        row_idx: 1-indexed target row (>= 41 = first slot after the
            reference grid).
        sample_meta: per-sample metadata.
        feststoff_label / eluat_label / gesamt_label: class labels.
        feststoff_drivers / eluat_drivers: parameter names driving the
            respective ranks (output of :func:`_drivers_for_split`).
        messwerte: (param, matrix) → display string lookup.
    """
    _set_cell(ws, f"A{row_idx}", sample_meta.probenbezeichnung)
    _set_cell(ws, f"B{row_idx}", sample_meta.petrographische_beschreibung)
    _set_cell(ws, f"C{row_idx}", sample_meta.stratigraphie)
    _set_cell(ws, f"D{row_idx}", sample_meta.labor_nummer)

    # Three Zuordnung columns at I / K / M, each colour-coded per class.
    for coord, label in (
        (f"I{row_idx}", feststoff_label),
        (f"K{row_idx}", eluat_label),
        (f"M{row_idx}", gesamt_label),
    ):
        _set_cell(ws, coord, label)
        color_key = _classification_color_key(label)
        hex_color = CLASS_FILL_HEX[color_key]
        cell = ws[coord]
        if _is_writable(cell):
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, size=10)

    # Build the Feststoff status lookup the categoriser needs.
    feststoff_status: Dict[str, str] = {}
    for _, row in (messwerte_df if False else []):
        pass
    # Re-derive feststoff_status from messwerte keys + the original df is
    # not available here. We pass an empty dict — the categoriser then
    # falls back to "decisive" for every above-BM-0 param. That's the
    # safer default; the caller can pre-compute and pass status if it
    # wants the FN3 bracketing applied here too.
    # In the caller we supply the df via attached attribute on messwerte.
    df_attached = getattr(messwerte, "_df", None)
    if df_attached is not None:
        for _, row in df_attached.iterrows():
            p = str(row.get("Parameter", ""))
            ein = str(row.get("Einheit", "")).lower()
            kl = str(row.get("Eingestufte Klasse", ""))
            is_f = ("kg" in ein) or ("vol" in ein) or ("m%" in ein) or (
                "%" in ein and "kg" not in ein and "/" not in ein
            )
            if is_f:
                feststoff_status[p] = kl

    if df_attached is not None:
        fdec, fign, edec, eign = _categorise_above_bm0(df_attached, feststoff_status)
        summary_text = _format_rel_schadstoffe(fdec, fign, edec, eign)
    else:
        # Legacy fallback: use the drivers list
        feststoff_parts: List[str] = []
        eluat_parts: List[str] = []
        fest_set = set(feststoff_drivers)
        elu_set = set(eluat_drivers)
        for (param, matrix), val in messwerte.items():
            if not val:
                continue
            if param in fest_set and matrix == "Feststoff":
                feststoff_parts.append(f"{param} {val}")
            elif param in elu_set and matrix == "Eluat":
                eluat_parts.append(f"{param} {val}")
        summary_text = (
            "Feststoff: " + ("; ".join(feststoff_parts) if feststoff_parts else "-")
            + "\nEluat: " + ("; ".join(eluat_parts) if eluat_parts else "-")
        )

    _set_cell(ws, f"O{row_idx}", summary_text)
    o_cell = ws[f"O{row_idx}"]
    if _is_writable(o_cell):
        o_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        o_cell.font = Font(size=9)




def _generate_new_design_pdf(
    sheet_dict: Dict[str, pd.DataFrame],
    output_dir: str,
    base_name: str,
    project: ProjectMeta,
    sample_meta_map: Dict[str, SampleMeta],
    bodenart_internal: str,
    skeleton_path: str,
    keep_intermediate_xlsx: bool = False,
) -> Optional[str]:
    """Build the company-format PDF (Feststoff + Eluat) via skeleton + LibreOffice.

    Args:
        sheet_dict: per-sample evaluator output.
        output_dir: where to write the PDF.
        base_name: output basename (e.g. ``"All_Samples"``).
        project: project header meta.
        sample_meta_map: sheet_name -> SampleMeta.
        bodenart_internal: internal BM-0 soil subtype (for class labeling).
        skeleton_path: path to ``template_skeleton.xlsx``.
        keep_intermediate_xlsx: if True, retain the populated .xlsx alongside the PDF.

    Returns:
        Path to the produced PDF, or None on failure.
    """
    if not os.path.exists(skeleton_path):
        print(f"  -> ERROR: template_skeleton.xlsx not found at {skeleton_path}")
        return None

    # Load the skeleton with rich_text=True so footnote superscripts on
    # threshold-row cells (e.g. TOC "1⁷" — value "1" with footnote-7
    # superscript) survive into the output. Without rich_text=True,
    # openpyxl flattens "1⁷" to plain "17", which renders as the
    # nonsensical literal "17" the user flagged.
    try:
        wb = load_workbook(skeleton_path, rich_text=True)
    except TypeError:
        # Older openpyxl without rich_text kwarg — fall back to plain load.
        wb = load_workbook(skeleton_path)
    if "Feststoff" not in wb.sheetnames or "Eluat" not in wb.sheetnames:
        print("  -> ERROR: skeleton missing required sheets (Feststoff / Eluat)")
        return None

    feststoff_ws = wb["Feststoff"]
    eluat_ws = wb["Eluat"]
    # 3rd page: Zusammenfassung (optional — older skeletons may lack it).
    # Uses the company Mantelverordnung Zusammenfassung sheet AS-IS: one
    # Zuordnung column (H, merged H10:K12 with the "Zuordnung Feststoff"
    # title), parameter columns L..BL. Per-sample data lands at row 43+.
    zusammen_ws = (
        wb["Zusammenfassung Dekklaration"]
        if "Zusammenfassung Dekklaration" in wb.sheetnames
        else (wb["Zusammenfassung"] if "Zusammenfassung" in wb.sheetnames else None)
    )

    # Replace project header tokens on all visible sheets
    _replace_header_tokens(feststoff_ws, project)
    _replace_header_tokens(eluat_ws, project)
    if zusammen_ws is not None:
        _replace_header_tokens(zusammen_ws, project)

    # Write one row per sample
    last_row_used: int = FIRST_SAMPLE_ROW - 1
    for offset, (sheet_name, df) in enumerate(sheet_dict.items()):
        row_idx = FIRST_SAMPLE_ROW + offset
        last_row_used = row_idx
        sample_meta = sample_meta_map.get(sheet_name, SampleMeta(probenbezeichnung=str(sheet_name)))
        feststoff_label, eluat_label, gesamt_label = _classify_split(df, bodenart_internal)
        messwerte = _build_messwert_lookup(df)
        feststoff_drivers, eluat_drivers, _, _ = _drivers_for_split(df)
        # Attach df to messwerte so _populate_zusammenfassung can compute
        # the bracketed-ignored driver categorisation. Plain dict can't
        # take arbitrary attributes, so wrap in a thin subclass.
        if not hasattr(messwerte, "_df"):
            class _MDict(dict): pass
            wrapped = _MDict(messwerte)
            wrapped._df = df
            messwerte = wrapped

        # 1) Apply uniform formatting on the row first (clears inherited skeleton fills)
        _copy_style_from_reference(feststoff_ws, row_idx, FIRST_SAMPLE_ROW, "AD")
        _copy_style_from_reference(eluat_ws, row_idx, FIRST_SAMPLE_ROW, "AS")

        # 2) Write data
        _write_sample_row_v05(
            feststoff_ws,
            row_idx,
            sample_meta,
            feststoff_label,
            messwerte,
            FESTSTOFF_PARAM_TO_COL,
            is_feststoff_sheet=True,
        )
        _write_sample_row_v05(
            eluat_ws,
            row_idx,
            sample_meta,
            eluat_label,
            messwerte,
            ELUAT_PARAM_TO_COL,
            is_feststoff_sheet=False,
        )

        # 3) Color the H cell (Zuordnung) per actual classification
        _apply_classification_fill(feststoff_ws, row_idx, feststoff_label)
        _apply_classification_fill(eluat_ws, row_idx, eluat_label)

        # 4) Per-cell parameter colouring — each parameter's cell carries
        #    the fill of ITS OWN class. Reviewer sees, at a glance, every
        #    threshold breach (including orientation drivers like
        #    Leitfähigkeit that are ignored at the Gesamt stage). Mirrors
        #    the user's master Mantelverordnung workbook.
        _apply_per_cell_class_fills(
            feststoff_ws, row_idx, df,
            FESTSTOFF_PARAM_TO_COL, is_feststoff_sheet=True,
        )
        _apply_per_cell_class_fills(
            eluat_ws, row_idx, df,
            ELUAT_PARAM_TO_COL, is_feststoff_sheet=False,
        )

        # 5) Populate the Zusammenfassung sheet (3rd page) — Fix C.
        if zusammen_ws is not None:
            _populate_zusammenfassung(
                zusammen_ws,
                ZUSAMMEN_FIRST_SAMPLE_ROW + offset,
                sample_meta,
                feststoff_label,
                eluat_label,
                gesamt_label,
                feststoff_drivers,
                eluat_drivers,
                messwerte,
            )

    # Extend print area to include all newly added sample rows. The skeleton
    # inherits the company template's print area which only covers the
    # reference grid; we must widen it or LibreOffice truncates the PDF.
    feststoff_ws.print_area = f"A1:AD{last_row_used}"
    eluat_ws.print_area = f"A1:AS{last_row_used}"
    if zusammen_ws is not None:
        zusammen_last = ZUSAMMEN_FIRST_SAMPLE_ROW + len(sheet_dict) - 1
        zusammen_ws.print_area = f"A1:{ZUSAMMEN_LAST_COL}{zusammen_last}"
        # Force single-page width on A3 landscape — matches A_4_3_1_Auswertung_Labor.pdf 3-page layout.
        zusammen_ws.page_setup.orientation = "landscape"
        zusammen_ws.page_setup.paperSize = zusammen_ws.PAPERSIZE_A3
        zusammen_ws.page_setup.fitToWidth = 1
        zusammen_ws.page_setup.fitToHeight = 1
        if zusammen_ws.sheet_properties.pageSetUpPr is None:
            from openpyxl.worksheet.properties import PageSetupProperties
            zusammen_ws.sheet_properties.pageSetUpPr = PageSetupProperties()
        zusammen_ws.sheet_properties.pageSetUpPr.fitToPage = True
        # Same for Feststoff/Eluat to guarantee 1 page each
        for _ws in (feststoff_ws, eluat_ws):
            _ws.page_setup.orientation = "landscape"
            _ws.page_setup.paperSize = _ws.PAPERSIZE_A3
            _ws.page_setup.fitToWidth = 1
            _ws.page_setup.fitToHeight = 1
            if _ws.sheet_properties.pageSetUpPr is None:
                from openpyxl.worksheet.properties import PageSetupProperties
                _ws.sheet_properties.pageSetUpPr = PageSetupProperties()
            _ws.sheet_properties.pageSetUpPr.fitToPage = True

    intermediate_xlsx = os.path.join(output_dir, f"Evaluation_{base_name}.xlsx")
    wb.save(intermediate_xlsx)

    pdf_path = _convert_xlsx_to_pdf(intermediate_xlsx, output_dir)

    if not keep_intermediate_xlsx and pdf_path and os.path.exists(intermediate_xlsx):
        try:
            os.remove(intermediate_xlsx)
        except OSError:
            pass

    return pdf_path if os.path.exists(pdf_path) else None


def _convert_xlsx_to_pdf(xlsx_path: str, output_dir: str) -> Optional[str]:
    """Convert ``xlsx_path`` to a same-name PDF via headless LibreOffice."""
    try:
        completed = subprocess.run(
            [LIBREOFFICE_BIN, "--headless", "--convert-to", "pdf",
             "--outdir", output_dir, xlsx_path],
            capture_output=True,
            text=True,
            timeout=120,
            check=False,
        )
    except (subprocess.TimeoutExpired, OSError) as e:
        print(f"  -> ERROR: LibreOffice conversion failed: {e}")
        return None

    if completed.returncode != 0:
        print(f"  -> ERROR: LibreOffice rc={completed.returncode}: {completed.stderr}")
        return None

    expected = os.path.join(output_dir, os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf")
    return expected if os.path.exists(expected) else None


# ---------------------------------------------------------------------------
# Public entry point — called by step2_auswertung
# ---------------------------------------------------------------------------


def create_combined_report(
    sheet_dict: Dict[str, pd.DataFrame],
    output_dir: str,
    original_filename: str,
    bodenart: str,
    project_meta: Optional[ProjectMeta] = None,
    sample_meta_map: Optional[Dict[str, SampleMeta]] = None,
    skeleton_path: Optional[str] = None,
) -> None:
    """Generate v04 legacy reports + v05 company-format PDF in one call."""
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    base_name = os.path.splitext(original_filename)[0]
    _generate_old_design_reports(sheet_dict, output_dir, base_name, bodenart)
    if skeleton_path is None:
        here = os.path.dirname(__file__)
        candidates = (
            os.path.join(here, "templates", "ebv_template_skeleton.xlsx"),
            os.path.join(here, "template_skeleton.xlsx"),
        )
        skel = next((p for p in candidates if os.path.exists(p)), candidates[0])
    else:
        skel = skeleton_path
    proj = project_meta or ProjectMeta()
    smm: Dict[str, SampleMeta] = sample_meta_map or {
        name: SampleMeta(probenbezeichnung=str(name)) for name in sheet_dict
    }
    _generate_new_design_pdf(
        sheet_dict=sheet_dict,
        output_dir=output_dir,
        base_name=base_name,
        project=proj,
        sample_meta_map=smm,
        bodenart_internal=bodenart,
        skeleton_path=skel,
    )
    print(f"  -> REPORTS GENERATED IN: {output_dir}")
