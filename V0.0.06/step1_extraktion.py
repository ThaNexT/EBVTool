"""
EBV Tool v05 — Step 1 extraction (dual-flow).

Two parallel flows are supported:

* **EBV flow** — reads PDF lab reports from ``0_input/EBV/``, extracts
  EBV-relevant data via ``pdf_parser``, and produces a validation workbook
  in ``1_validation/EBV/<timestamp>_Validierung/`` with one sheet per sample
  plus a ``_Project`` sheet and (new in this build) a transposed ``UMWELT``
  echo sheet matching the company workbook layout.

* **Aggressivität flow** — reads filled Aggressivität input workbooks from
  ``0_input/Aggressivität/`` and echoes them to
  ``1_validation/Aggressivität/<timestamp>_Validierung/`` after sheet/column
  validation. *Implementation pending — placeholder in this build.*

CLI::

    python step1_extraktion.py [--flow {ebv,aggressivität,all}]

``--flow all`` (default) runs both branches; either branch with no input
files is silently skipped.

EBV per-sample sheet schema (unchanged from previous build):
    * ``Soil_Type``                   user selector (s / l / c / undef)
    * ``Probenbezeichnung``           auto-filled from filename
    * ``Petrographische_Beschreibung``  empty, user fills during validation
    * ``Stratigraphie``               empty, user fills during validation
    * ``Labor_Nummer``                auto-filled from filename when matchable
    * EBV measurement columns from ``pdf_parser`` output.

EBV project-level header sheet ``_Project`` (unchanged):
    Projektnummer (auto from filename prefix), Bauvorhaben, LOS, Bauwerk.

New UMWELT echo sheet (read-only style; for verification only):
    Transposed parameter×sample grid mirroring the company
    ``2604XX_Rohdaten & Aggressivität.xlsx`` UMWELT layout. Step 2 does
    not read this sheet — it is for human review only.
"""
from __future__ import annotations

import argparse
import glob
import os
import re
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
    _BERLIN_TZ = ZoneInfo("Europe/Berlin")
except Exception:
    _BERLIN_TZ = None


def _berlin_now() -> "datetime":
    """Return current datetime in Europe/Berlin so timestamps line up
    with the user's wall clock instead of the (UTC) container clock."""
    if _BERLIN_TZ is not None:
        return datetime.now(_BERLIN_TZ)
    return datetime.now()
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from pdf_parser import extract_all_data_from_pdf

#: Per-flow input/output directory layout. Top-level ``0_input``/``1_validation``
#: directories from earlier builds are no longer touched directly.
INPUT_DIR_EBV: str = os.path.join("0_input", "EBV")
INPUT_DIR_AGGR: str = os.path.join("0_input", "Aggressivität")
INPUT_DIR_PAK: str = os.path.join("0_input", "PAK")
VALIDATION_ROOT: str = "1_validation"
OUTPUT_DIR_EBV: str = VALIDATION_ROOT
OUTPUT_DIR_AGGR: str = VALIDATION_ROOT
OUTPUT_DIR_PAK: str = VALIDATION_ROOT

#: Shared timestamp set lazily on first ``_get_session_dir()`` call so all
#: flows within a single ``step1`` invocation land in the SAME folder.
_SESSION_TS_VAL: Optional[str] = None


def _get_session_dir_validation() -> str:
    """Return (and lazily create) the shared validation-session folder.

    All flows invoked within a single ``step1`` run share the timestamp,
    so a project's EBV/Aggressivität/PAK validation files end up in
    ``1_validation/<ts>_Validierung/`` together rather than three
    separate per-flow folders.

    Returns:
        Absolute or relative path to the timestamped session folder.
    """
    global _SESSION_TS_VAL
    if _SESSION_TS_VAL is None:
        _SESSION_TS_VAL = _berlin_now().strftime("%Y-%m-%d_%H-%M")
    session_dir = os.path.join(VALIDATION_ROOT, f"{_SESSION_TS_VAL}_Validierung")
    os.makedirs(session_dir, exist_ok=True)
    return session_dir

#: New metadata columns inserted after ``Soil_Type`` on every sample sheet.
METADATA_COLUMNS: Tuple[str, ...] = (
    "Probenbezeichnung",
    "Petrographische_Beschreibung",
    "Stratigraphie",
    "Labor_Nummer",
)


def _extract_metadata_from_filename(filename: str) -> Tuple[str, str, str]:
    """Parse filename to obtain (projektnummer, probenbezeichnung, labor_nummer).

    Filename pattern observed in real projects::

        e327524___BW_6524_5180_MP03_UST250114832_031.pdf
        ^^^^^^^               ^^^^ ^^^^^^^^^^^^^^^^^
        projektnummer         probe labor_nummer

    All three are best-effort extractions; the user can override them
    in the validation Excel before running Step 2.

    Args:
        filename: just the basename, with or without extension.

    Returns:
        ``(projektnummer, probenbezeichnung, labor_nummer)``. Missing
        values are returned as ``""``.
    """
    base = os.path.splitext(os.path.basename(filename))[0]

    # Projektnummer: leading 'e' + digits, optionally followed by '-' + suffix
    projekt_match = re.match(r"^(e\d+)", base, flags=re.IGNORECASE)
    projektnummer = ""
    if projekt_match:
        raw = projekt_match.group(1)
        # Insert "-" after the leading letter for the typical "e-327524" rendering
        projektnummer = f"{raw[0].lower()}-{raw[1:]}"

    # Probenbezeichnung: pattern MP## or P## or P-## etc.
    probe_match = re.search(r"(?:^|[_\s\-,])(M?P\s*-?\s*\d{2,3})(?=[_\s\-,]|$)", base)
    probenbezeichnung = ""
    if probe_match:
        probenbezeichnung = probe_match.group(1).replace(" ", "")

    # Labor_Nummer: pattern like UST250114832_031 -> "UST-25-0114832-03"
    labor_match = re.search(r"(?:^|[_\s\-,])(UST)(\d{2})(\d+)[_\-](\d{2,3})(?=[_\s\-.,]|$)", base)
    labor_nummer = ""
    if labor_match:
        prefix, year, project, sample = labor_match.groups()
        sample_clean = sample.lstrip("0").zfill(2) if sample.startswith("0") else sample
        # Drop trailing "1" suffix on sample if present (matches "01" convention)
        if len(sample) == 3 and sample[-1] in {"0", "1"}:
            sample_clean = sample[:2]
        labor_nummer = f"{prefix}-{year}-{project}-{sample_clean}"

    return projektnummer, probenbezeichnung, labor_nummer




def _extract_pdf_header_meta(pdf_path: str) -> Dict[str, str]:
    """Extract sample-identification fields from an SGS/AGROLAB-style PDF header.

    Looks for keys like ``Probenbezeichnung:``, ``Probe Nr.:``, ``Prüfbericht Nr.:``,
    ``Auftrag-Nr.:``, ``Projekt:`` on page 1. Authoritative source of
    metadata — preferred over filename regex.

    Args:
        pdf_path: filesystem path to the PDF.

    Returns:
        Dict with keys ``probenbezeichnung``, ``probennummer``, ``labor_nummer``,
        ``auftrag_nr``, ``projektnummer``, ``bauvorhaben``. Missing keys map to "".
    """
    import pdfplumber
    out: Dict[str, str] = {
        "probenbezeichnung": "",
        "probennummer": "",
        "labor_nummer": "",
        "auftrag_nr": "",
        "projektnummer": "",
        "bauvorhaben": "",
    }
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages[:2]:  # header is on page 1; page 2 may repeat
                text += (page.extract_text() or "") + "\n"
    except Exception:
        return out

    # SGS-style "Probenbezeichnung: MP05"
    m = re.search(r"Probenbezeichnung:\s*(\S.+?)(?:\n|$)", text)
    if m:
        out["probenbezeichnung"] = m.group(1).strip()
    # AGROLAB-style "Kunden-Probenbezeichnung 6324-5180-GWM3"
    if not out["probenbezeichnung"]:
        m = re.search(r"Kunden-?Probenbezeichnung\s+(\S.+?)(?:\s{2,}|\n|$)", text)
        if m:
            out["probenbezeichnung"] = m.group(1).strip()

    # Probe Nr.: UST-25-0114832-05  (per-sample lab number)
    m = re.search(r"Probe\s*Nr\.?:\s*(\S+)", text)
    if m:
        out["probennummer"] = m.group(1).strip()
    # AGROLAB analysennr
    if not out["probennummer"]:
        m = re.search(r"Analysennr\.?\s+(\S+)", text)
        if m:
            out["probennummer"] = m.group(1).strip()

    # Prüfbericht Nr.: UST-25-0114832/05-1  (lab nr.)
    m = re.search(r"Pr[uü]?fbericht\s*Nr\.?:\s*(\S+)", text)
    if m:
        out["labor_nummer"] = m.group(1).strip()

    # Auftrag-Nr.: UST-25-0114832
    m = re.search(r"Auftrag[\-\s]*Nr\.?:\s*(\S+)", text)
    if m:
        out["auftrag_nr"] = m.group(1).strip()
    # AGROLAB: "Auftrag 3774747"
    if not out["auftrag_nr"]:
        m = re.search(r"\bAuftrag\s+(\d+)\b", text)
        if m:
            out["auftrag_nr"] = m.group(1).strip()

    # Projekt: e-327524 / BW 6524 5180
    m = re.search(r"Projekt:?\s*(\S.+?)(?:\n|$)", text)
    if m:
        proj_line = m.group(1).strip()
        # Split "e-327524 / BW 6524 5180" → projektnummer + bauvorhaben
        parts = re.split(r"\s*/\s*", proj_line, maxsplit=1)
        out["projektnummer"] = parts[0].strip()
        if len(parts) > 1:
            out["bauvorhaben"] = parts[1].strip()
    return out




def _read_background_metadata(input_root: str = "0_input") -> Dict[str, object]:
    """Parse the optional ``background_data.txt`` file in the input root.

    File format (tab-separated key/value in ``-DATA-``, parallel arrays
    under ``-SAMPLE-``)::

        -DATA-
        Projektnummer   e-327524
        Bauvorhaben     Feldwegbrücke bei Distelhausen
        LOS             1
        Bauwerk         6324 536 0

        -SAMPLE-
        Name
        MP01-BW...
        MP02-BW...
        ...

        Art
        Auffüllung
        ...

        Stratigraphie
        qhy
        ...

        type (Sand/Lehm/Ton/Wasser)
        undef
        ...

        Petrographische Beschreibung
        Auffüllung
        ...

    Args:
        input_root: directory that may contain ``background_data.txt``.

    Returns:
        Dict with keys:
          * ``"data"`` — flat ``{key: value}`` map (Projektnummer etc.).
          * ``"samples"`` — list of per-sample dicts keyed by
            Probenbezeichnung, e.g.
            ``{"MP02-BW...": {"Art": "Auffüllung", "Stratigraphie": "qhy",
              "type": "undef", "petrographische": "Auffüllung"}, ...}``.
        Empty dicts when the file is missing or malformed.
    """
    out: Dict[str, object] = {"data": {}, "samples": {}}
    path = os.path.join(input_root, "background_data.txt")
    if not os.path.exists(path):
        return out
    try:
        with open(path, "r", encoding="utf-8") as f:
            lines = f.read().splitlines()
    except Exception:
        return out

    # Walk sections delimited by -SECTION- markers
    section: Optional[str] = None
    sample_blocks: Dict[str, List[str]] = {}
    current_block: Optional[str] = None
    current_values: List[str] = []
    BLOCK_FIELDS = {"Name", "Art", "Stratigraphie", "type", "Petrographische Beschreibung"}

    def flush_block() -> None:
        nonlocal current_block, current_values
        if current_block is not None:
            sample_blocks[current_block] = current_values
        current_block = None
        current_values = []

    for raw in lines:
        line = raw.rstrip()
        s = line.strip()
        if not s:
            # blank line ends a block but stays inside the section
            if section == "SAMPLE":
                flush_block()
            continue
        if s == "-DATA-":
            section = "DATA"; flush_block(); continue
        if s == "-SAMPLE-":
            section = "SAMPLE"; flush_block(); continue
        if section == "DATA":
            # key\tvalue OR key (whitespace) value
            parts = [p.strip() for p in re.split(r"\t+|\s{2,}", s, maxsplit=1) if p.strip()]
            if len(parts) == 2:
                out["data"][parts[0]] = parts[1]
        elif section == "SAMPLE":
            if current_block is None:
                # First non-blank line in a section names the block
                # (Name / Art / Stratigraphie / type ... / Petrographische ...).
                # Match liberally — the user may add explanatory text in
                # parentheses (e.g. "type (Sand/Lehm/Ton/Wasser)").
                head = s
                for f in BLOCK_FIELDS:
                    if head.startswith(f):
                        current_block = f
                        current_values = []
                        break
                else:
                    # Unknown header — skip silently
                    current_block = "__skip__"
                    current_values = []
            else:
                current_values.append(s)
    flush_block()

    # Pivot block lists into per-sample dicts, keyed by Name.
    names = sample_blocks.get("Name", [])
    for idx, name in enumerate(names):
        clean = name.strip()
        if not clean: continue
        rec: Dict[str, str] = {}
        for field in ("Art", "Stratigraphie", "type", "Petrographische Beschreibung"):
            vals = sample_blocks.get(field, [])
            rec[field] = vals[idx].strip() if idx < len(vals) else ""
        out["samples"][clean] = rec
    return out

def _write_project_sheet(wb: Workbook, projektnummer_default: str = "") -> None:
    """Insert the ``_Project`` sheet at position 0 with header placeholders."""
    proj_ws = wb.create_sheet("_Project", 0)
    proj_ws["A1"] = "Field"
    proj_ws["B1"] = "Value"
    proj_ws["A1"].font = Font(bold=True)
    proj_ws["B1"].font = Font(bold=True)
    proj_ws["A1"].fill = PatternFill("solid", fgColor="E0E0E0")
    proj_ws["B1"].fill = PatternFill("solid", fgColor="E0E0E0")

    # Pull project header values from background_data.txt if available.
    bg = _BG_META.get("data", {})  # type: ignore[union-attr]
    rows: Tuple[Tuple[str, str], ...] = (
        ("Projektnummer", bg.get("Projektnummer", "") or projektnummer_default),
        ("Bauvorhaben",   bg.get("Bauvorhaben", "")   or "No Value"),
        ("LOS",           bg.get("LOS", "")           or "No Value"),
        ("Bauwerk",       bg.get("Bauwerk", "")       or "No Value"),
    )
    for i, (key, val) in enumerate(rows, start=2):
        proj_ws[f"A{i}"] = key
        proj_ws[f"A{i}"].font = Font(bold=True)
        proj_ws[f"B{i}"] = val

    proj_ws.column_dimensions["A"].width = 20
    proj_ws.column_dimensions["B"].width = 60

    # Hint cell
    proj_ws["A7"] = (
        "Fill in the values above. Step 2 reads this sheet for the report header. "
        "Sample-level fields (Probenbezeichnung, Petrographische_Beschreibung, "
        "Stratigraphie, Labor_Nummer) are on each sample sheet."
    )
    proj_ws["A7"].font = Font(italic=True, color="666666")
    proj_ws.merge_cells("A7:B7")




#: Cached background-metadata read on module import. ``_BG_META["data"]``
#: holds project-level fields (Projektnummer, Bauvorhaben, LOS, Bauwerk);
#: ``_BG_META["samples"]`` maps Probenbezeichnung → per-sample dict.
_BG_META: Dict[str, object] = {"data": {}, "samples": {}}
def _ingest_ebv_dir(input_dir: str = INPUT_DIR_EBV) -> Tuple[List[Tuple[str, "pd.DataFrame"]], str]:
    """Ingest all EBV PDFs from ``input_dir`` and return parsed sample list.

    Args:
        input_dir: directory containing EBV lab-report PDFs.

    Returns:
        Tuple ``(samples, projektnummer)`` where ``samples`` is a list of
        ``(sample_name, parsed_dataframe)`` and ``projektnummer`` is the
        first non-empty projekt code found across all filenames.
    """
    samples: List[Tuple[str, pd.DataFrame]] = []
    projektnummer_first: str = ""
    if not os.path.exists(input_dir):
        return samples, projektnummer_first
    pdf_files = sorted(glob.glob(os.path.join(input_dir, "*.pdf")))
    for pdf_path in pdf_files:
        filename = os.path.basename(pdf_path)
        sheet_name = filename[:31]
        projektnummer, probenbezeichnung, labor_nummer = _extract_metadata_from_filename(filename)
        # Authoritative metadata: try to pull from the PDF header itself.
        pdf_meta = _extract_pdf_header_meta(pdf_path)
        probenbezeichnung = pdf_meta.get("probenbezeichnung") or probenbezeichnung
        # Prefer Probennummer (hyphen-format "UST-26-0002769-03") over
        # Prüfbericht Nr. (slash-format) — that's the per-sample lab ID
        # the user enters into their master Mantelverordnung workbook.
        labor_nummer = (
            pdf_meta.get("probennummer") or pdf_meta.get("labor_nummer") or labor_nummer
        )
        projektnummer = pdf_meta.get("projektnummer") or projektnummer
        if not projektnummer_first and projektnummer:
            projektnummer_first = projektnummer
        df = extract_all_data_from_pdf(pdf_path)
        if df.empty:
            continue
        # Overlay per-sample fields from background_data.txt if present.
        # EBV per-sample metadata fallback chain:
        #   1. background_data.txt entry for this probenbezeichnung
        #   2. EBV default: Petrographische = "Auffüllung", Stratigraphie = "qhy"
        bg_sample = _BG_META["samples"].get(probenbezeichnung, {})
        soil_type = (bg_sample.get("type", "") or "").strip() or "undef"
        if soil_type.lower() == "wasser":
            soil_type = "undef"
        strat = (bg_sample.get("Stratigraphie", "") or "").strip() or "qhy"
        petro = (bg_sample.get("Petrographische Beschreibung", "") or "").strip() or "Auffüllung"
        df.insert(0, "Soil_Type", soil_type)
        df.insert(1, "Probenbezeichnung", probenbezeichnung)
        df.insert(2, "Petrographische_Beschreibung", petro)
        df.insert(3, "Stratigraphie", strat)
        df.insert(4, "Labor_Nummer", labor_nummer)
        samples.append((sheet_name, df))
    return samples, projektnummer_first


def _parse_pak_multi_sample(pdf_path: str) -> List[Tuple[str, str, "pd.DataFrame"]]:
    """Parse a PAK lab PDF that may carry multiple samples per file.

    SGS PAK reports list every sample as its own data column on page 2,
    headed by ``Probe Nr.:`` and ``Bezeichnung:`` lines. Each parameter row
    then carries N tokens after the unit — one per sample::

        Probe Nr.:        UST-26-0002769-01    UST-26-0002769-02
        Bezeichnung:      MP01-BW63245360      Asp03-BW6324536
        Naphthalin  mg/kg <0,05                <0,05
        Phenanthren mg/kg 0,13                 0,088
        Summe PAK EPA mg/kg 0,85               0,43
        Phenol-Index mg/l <0,01                <0,01

    This helper splits that into one ``(probenbezeichnung, labor_nummer,
    DataFrame)`` tuple per sample. The DataFrame schema matches
    :func:`pdf_parser.extract_all_data_from_pdf` so the unified validation
    writer can consume it identically to the single-sample EBV path.

    Args:
        pdf_path: filesystem path to a multi-sample PAK PDF.

    Returns:
        List of tuples. Empty when the PDF doesn't follow the
        Probe-Nr./Bezeichnung-side-by-side layout (callers can fall back
        to the single-sample EBV ingest).
    """
    import pdfplumber
    from pdf_parser import map_parameter_name, determine_matrix, parse_value_and_unit

    samples: List[Tuple[str, str, "pd.DataFrame"]] = []
    probennummern: List[str] = []
    bezeichnungen: List[str] = []
    matrix_state = "Feststoff"  # flips to "Eluat" once the "Eluat" banner appears
    rows_per_sample: List[List[Dict[str, Any]]] = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
                lines = text.split("\n")
                i = 0
                while i < len(lines):
                    raw = lines[i].strip()
                    low = raw.lower()
                    # Probe Nr. header — next non-blank line carries the IDs
                    if low.startswith("probe nr."):
                        j = i + 1
                        while j < len(lines) and not lines[j].strip():
                            j += 1
                        if j < len(lines):
                            ids = lines[j].strip().split()
                            # Heuristic: lab IDs look like "UST-..." or contain digits
                            cand = [t for t in ids if any(c.isdigit() for c in t) and len(t) > 4]
                            if cand and not probennummern:
                                probennummern = cand
                                rows_per_sample = [[] for _ in probennummern]
                        i = j + 1
                        continue
                    if low.startswith("bezeichnung:"):
                        # Strip the "Bezeichnung:" prefix and split
                        tail = raw.split(":", 1)[1].strip()
                        # Expect N whitespace-separated names matching the N probennummern
                        names = tail.split()
                        if probennummern and len(names) >= len(probennummern):
                            bezeichnungen = names[: len(probennummern)]
                        i += 1
                        continue
                    if low == "eluat" or low.startswith("eluat "):
                        matrix_state = "Eluat"
                        i += 1
                        continue
                    if not probennummern:
                        i += 1
                        continue
                    # Try to parse a data row: <param> ... <unit> <val_1> ... <val_N>
                    tokens = raw.split()
                    if len(tokens) < 2 + len(probennummern):
                        i += 1
                        continue
                    # Last N tokens are the per-sample values (incl. operator/<BG)
                    n = len(probennummern)
                    values = tokens[-n:]
                    head = tokens[:-n]
                    # Find the unit token by scanning head from the right
                    unit_idx = -1
                    for k in range(len(head) - 1, -1, -1):
                        t = head[k].lower()
                        if any(u in t for u in ("mg/kg", "µg/l", "ug/l", "mg/l", "%", "vol", "ts", "tr")):
                            unit_idx = k
                            break
                    if unit_idx < 0:
                        i += 1
                        continue
                    raw_param = " ".join(head[:unit_idx]).strip()
                    unit = head[unit_idx]
                    if not raw_param:
                        i += 1
                        continue
                    mapped = map_parameter_name(raw_param)
                    full_row = raw
                    for k, val_token in enumerate(values):
                        # Run the existing per-sample unit/value parser so we
                        # get consistent operator + numeric extraction.
                        u, op, num = parse_value_and_unit([raw_param, unit, val_token])
                        matrix = determine_matrix(u) if u and u != "-" else matrix_state
                        rows_per_sample[k].append({
                            "Lab_Original_String": raw_param,
                            "Full_Row": full_row,
                            "EBV_Parameter": mapped if mapped else "",
                            "Matrix": matrix,
                            "Lab_Operator": op,
                            "Lab_Value": num,
                            "Lab_Unit": u or unit,
                        })
                    i += 1
    except Exception as e:  # noqa: BLE001
        import logging
        logging.error("PAK multi-sample parser failed on %s: %s", pdf_path, e)
        return []

    # Materialise one DataFrame per probe found in the PDF.
    for idx, probennummer in enumerate(probennummern):
        bezeichnung = bezeichnungen[idx] if idx < len(bezeichnungen) else f"Probe {probennummer}"
        df = pd.DataFrame(rows_per_sample[idx])
        samples.append((bezeichnung, probennummer, df))
    return samples


def _ingest_pak_dir(input_dir: str = INPUT_DIR_PAK) -> Tuple[List[Tuple[str, "pd.DataFrame"]], str]:
    """Ingest all PAK lab PDFs, splitting multi-sample reports into one
    entry per Probenbezeichnung found inside the PDF.

    SGS PAK reports for Straßenaufbruch can carry several samples (e.g.
    MP01 + Asp03) in one PDF as side-by-side columns. The multi-sample
    parser :func:`_parse_pak_multi_sample` finds the per-sample column
    headers and emits one DataFrame per probe. Single-sample PDFs are
    parsed with the EBV PDF parser as before.
    """
    samples: List[Tuple[str, "pd.DataFrame"]] = []
    projektnummer_first: str = ""
    if not os.path.exists(input_dir):
        return samples, projektnummer_first
    pdf_files = sorted(glob.glob(os.path.join(input_dir, "*.pdf")))
    for pdf_path in pdf_files:
        filename = os.path.basename(pdf_path)
        pdf_meta = _extract_pdf_header_meta(pdf_path)
        projektnummer = pdf_meta.get("projektnummer") or _extract_metadata_from_filename(filename)[0]
        if not projektnummer_first and projektnummer:
            projektnummer_first = projektnummer

        # Try the multi-sample parser first; fall back to the EBV single
        # parser if the PDF doesn't follow the Probe-Nr./Bezeichnung layout.
        multi = _parse_pak_multi_sample(pdf_path)
        if not multi:
            # Single-sample fallback (legacy behaviour)
            df = extract_all_data_from_pdf(pdf_path)
            if df.empty:
                continue
            bg_sample = _BG_META["samples"].get(pdf_meta.get("probenbezeichnung") or "", {})
            soil_type = bg_sample.get("type", "").strip() or "undef"
            if soil_type.lower() == "wasser":
                soil_type = "undef"
            df.insert(0, "Soil_Type", soil_type)
            df.insert(1, "Probenbezeichnung", pdf_meta.get("probenbezeichnung") or filename)
            df.insert(2, "Petrographische_Beschreibung", bg_sample.get("Petrographische Beschreibung", ""))
            df.insert(3, "Stratigraphie", bg_sample.get("Stratigraphie", ""))
            df.insert(4, "Labor_Nummer", pdf_meta.get("probennummer") or pdf_meta.get("labor_nummer") or "")
            samples.append((filename[:31], df))
            continue

        # Multi-sample: one entry per probe found in the PDF.
        # Metadata fallback chain (per-field):
        #   1. background_data.txt entry for this probe's Bezeichnung
        #   2. any other probe in THIS SAME PDF that has the field set
        #      (e.g. Asp03 inherits Petrographische from MP01 since they
        #       came from the same Straßenaufbruch sampling)
        #   3. PAK default: Petrographische = "Auffüllung", Stratigraphie = "qhy"
        per_sample_bg = [
            _BG_META["samples"].get(bezeichnung, {}) for (bezeichnung, _pn, _df) in multi
        ]
        def _sibling_or_default(field: str, default: str) -> str:
            for bg in per_sample_bg:
                v = (bg.get(field, "") or "").strip()
                if v:
                    return v
            return default
        sibling_petro = _sibling_or_default("Petrographische Beschreibung", "Auffüllung")
        sibling_strat = _sibling_or_default("Stratigraphie", "qhy")
        sibling_soil  = _sibling_or_default("type", "undef")
        if sibling_soil.lower() == "wasser":
            sibling_soil = "undef"

        for idx, (bezeichnung, probennummer, df) in enumerate(multi):
            if df.empty:
                continue
            bg = per_sample_bg[idx]
            soil_type = (bg.get("type", "") or "").strip() or sibling_soil or "undef"
            if soil_type.lower() == "wasser":
                soil_type = "undef"
            petro = (bg.get("Petrographische Beschreibung", "") or "").strip() or sibling_petro
            strat = (bg.get("Stratigraphie", "") or "").strip() or sibling_strat
            df.insert(0, "Soil_Type", soil_type)
            df.insert(1, "Probenbezeichnung", bezeichnung)
            df.insert(2, "Petrographische_Beschreibung", petro)
            df.insert(3, "Stratigraphie", strat)
            df.insert(4, "Labor_Nummer", probennummer)
            sheet_name = bezeichnung[:31]
            samples.append((sheet_name, df))
    return samples, projektnummer_first


def _ingest_aggr_dir(input_dir: str = INPUT_DIR_AGGR) -> Tuple[List[Tuple[str, "pd.DataFrame"]], str]:
    """Ingest all Aggressivität PDFs via pdf_parser_aggressivität.

    Returns:
        ``(samples, projektnummer)`` with the same shape as the EBV helper.
    """
    from pdf_parser_aggressivität import (
        extract_all_data_from_pdf as extract_aggr_data,
        extract_probenbezeichnung as extract_aggr_probe,
    )
    samples: List[Tuple[str, pd.DataFrame]] = []
    projektnummer_first: str = ""
    if not os.path.exists(input_dir):
        return samples, projektnummer_first
    pdf_files = sorted(glob.glob(os.path.join(input_dir, "*.pdf")))
    for pdf_path in pdf_files:
        filename = os.path.basename(pdf_path)
        sheet_name = filename[:31]
        projektnummer, _, _ = _extract_metadata_from_filename(filename)
        # Authoritative metadata: pull from the PDF header text.
        pdf_meta = _extract_pdf_header_meta(pdf_path)
        probe = pdf_meta.get("probenbezeichnung") or extract_aggr_probe(pdf_path) or filename
        probennummer = pdf_meta.get("probennummer") or pdf_meta.get("labor_nummer") or ""
        projektnummer = pdf_meta.get("projektnummer") or projektnummer
        if not projektnummer_first and projektnummer:
            projektnummer_first = projektnummer
        df = extract_aggr_data(pdf_path)
        if df.empty:
            continue
        # Aggressivität default fallback: water samples typically have no
        # petrographische / stratigraphie. Use a placeholder "-" so the
        # column isn't blank in the validation workbook, and pull from
        # background_data.txt if defined for this probe.
        bg = _BG_META["samples"].get(probe, {})
        petro = (bg.get("Petrographische Beschreibung", "") or "").strip() or "-"
        strat = (bg.get("Stratigraphie", "") or "").strip() or "-"
        df.insert(0, "Probenbezeichnung", probe)
        df.insert(1, "Analysennr", probennummer)
        df.insert(2, "Petrographische_Beschreibung", petro)
        df.insert(3, "Stratigraphie", strat)
        df.insert(4, "Tiefe", "")
        df.insert(5, "Formation", "")
        df.insert(6, "Sample_Type", "Wasser")
        samples.append((sheet_name, df))
    return samples, projektnummer_first


def _write_unified_validation(
    session_dir: str,
    pak_samples: List[Tuple[str, "pd.DataFrame"]],
    ebv_samples: List[Tuple[str, "pd.DataFrame"]],
    aggr_samples: List[Tuple[str, "pd.DataFrame"]],
    projektnummer: str,
) -> Tuple[str, str]:
    """Write ONE ``Validation.xlsx`` containing UMWELT (master layout) +
    per-sample tabs with [PAK]/[EBV]/[Aggr] prefixes for step2 dispatch.

    Args:
        session_dir: target directory.
        pak_samples / ebv_samples / aggr_samples: ingested sample lists.
        projektnummer: auto-extracted project code for the _Project sheet.

    Returns:
        Tuple ``(xlsx_path, html_path)``.
    """
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell

    master_path = os.path.join("templates", "umwelt_master.xlsx")
    if not os.path.exists(master_path):
        raise FileNotFoundError(f"Master UMWELT template not found: {master_path}")
    wb = load_workbook(master_path)
    keep = {"UMWELT"}
    for s in list(wb.sheetnames):
        if s not in keep:
            del wb[s]

    umw = wb["UMWELT"]

    # Clear the example sample columns C..I (master's demo data) but
    # preserve column-A labels + row formatting.
    EXAMPLE_COL_START = 3   # col C
    EXAMPLE_COL_END = 9     # col I
    for r in range(1, 132):
        for c in range(EXAMPLE_COL_START, EXAMPLE_COL_END + 1):
            cell = umw.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    # Insert _Project sheet at position 0
    _write_project_sheet(wb, projektnummer_default=projektnummer)
    wb.move_sheet("_Project", offset=-(len(wb.sheetnames) - 1))

    # The master UMWELT splits parameters into three flow-specific sections:
    #   PAK section:   rows 5..23  (grey fill, PAK16 individuals + Phenolindex)
    #   EBV section:   rows 24..65 (default fill, EBV Feststoff + Eluat PAK + PCB + EOX)
    #   Aggr section:  rows 66..131 (water/Eluat parameters)
    # Same parameter names (e.g. "Naphthalin") appear in MULTIPLE sections, so
    # matching is restricted per flow.
    from umwelt_template import is_feststoff_unit, is_eluat_unit
    SECTION_ROWS: Dict[str, range] = {
        "PAK":  range(5, 24),     # grey FFF2F2F2 — PAK16 Feststoff + Phenolindex
        "EBV":  range(24, 110),   # default — EBV Feststoff + Eluat (metals, PAK Eluat, PCB)
        "Aggr": range(110, 132),  # blue FFDEEAF6 — Aggressivität-Wasser
    }

    def _norm_label(s: str) -> str:
        """Normalise a label for matching: lowercase + hyphen/comma → space."""
        s = (s or "").strip().lower()
        s = re.sub(r"\s*\[[^\]]*\]\s*", " ", s)  # drop "[25°C]"
        s = s.replace("-", " ").replace("_", " ").replace(",", " ").replace(".", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def _fold_umlauts(s: str) -> str:
        """Drop German umlauts entirely so OCR-stripped lab text matches.

        PDF OCR commonly drops the umlaut character: "Leitfähigkeit" becomes
        "Leitfhigkeit" in the parsed text. Folding ä→"" (not ä→a) restores
        equality with the template's umlauted form once both go through this
        normalisation.
        """
        return (s.replace("ä", "").replace("ö", "").replace("ü", "")
                 .replace("ß", "ss"))

    def _norm_label_tight(s: str) -> str:
        """Even more aggressive: also drop all whitespace.

        Lets "phenol index" match "phenolindex" and "summe pak 16" match
        "summepak16".
        """
        return _norm_label(s).replace(" ", "")

    #: Synonym pairs to bridge minor master/parser label drift.
    #: Both sides are normalised lower-case label fragments.
    # IMPORTANT: entries must use the NORMALISED form (commas/periods/
    # hyphens replaced by spaces) because the matcher compares against
    # ``_norm_label(ln)`` and ``_norm_label(lkey)`` — both pass through
    # the same normaliser, so raw forms with commas would never match.
    LABEL_SYNONYMS: List[Tuple[str, str]] = [
        # Säurekapazität — long lab form vs. short master abbreviation.
        ("säurekap bis ph 4 3", "säurekapazität bis ph 4 3"),
        ("säurekap bis ph 4 3", "ks 4 3"),
        ("säurekap bis ph 4 3", "ks4 3"),
        # Kalk(l) Kohlensäure variants
        ("kalk kohlensäure", "kalkl kohlensäure"),
        ("kalk kohlensäure", "kalklösende kohlensäure"),
        # Leitfähigkeit variants
        ("leitfähigkeit", "leitfähigkeit bei 20 °c"),
        ("leitfähigkeit", "leitfähigkeit bei 25 °c"),
        ("leitfähigkeit", "elektrische leitfähigkeit"),
        # PAK / Phenol synonyms
        ("summe pak 16", "summe pak epa"),
        ("summe pak 16", "summe pak nach epa"),
        ("summe pak 16", "summe pak (16)"),
        ("phenolindex", "phenol index"),
        # Dibenz(o)(a,h)anthracen — labs use both "Dibenz" and "Dibenzo"
        # spellings (IUPAC "dibenzo[a,h]anthracene" vs. common German
        # "Dibenz(a,h)anthracen"). Bridge them paren-stripped.
        ("dibenz anthracen", "dibenzo anthracen"),
        ("dibenz(a h)anthracen", "dibenzo(a h)anthracen"),
        # Benzo(g,h,i)perylen — template Eluat row uses commas, Feststoff
        # row drops them ("Benzo(ghi)perylen" vs "Benzo(g,h,i)perylen").
        ("benzo(ghi)perylen", "benzo(g h i)perylen"),
        # Neutralsalze vs. template typo "Neutralssalze" (double s).
        ("neutralsalze", "neutralssalze"),
        # Gesamthärte mmol/l — lab form "Gesamthärte (Summe Erdalkalien)"
        # bridges to the short master label "Gesamthärte" (the unit
        # disambiguates r127 mg/l vs r128 mmol/l).
        ("gesamthärte", "gesamthärte summe erdalkalien"),
        # Betonaggressivität — lab string contains "(Angriffsgrad DIN..."
        # which doesn't survive paren-strip cleanly; bridge to master row.
        ("betonaggressivität labor", "betonaggressivität angriffsgrad din 4030"),
        ("betonaggressivität labor", "betonaggressivität"),
    ]

    def _build_row_index(flow: str) -> List[Tuple[str, str, int]]:
        """List of (normalised label, normalised unit, row_idx) for this flow.

        The unit lets the matcher distinguish Feststoff (mg/kg) from Eluat
        (mg/l) rows that share a label (e.g. "Arsen" / "Blei" appear in BOTH
        matrices in the EBV section).
        """
        out: List[Tuple[str, str, int]] = []
        for r in SECTION_ROWS[flow]:
            label = umw.cell(row=r, column=1).value
            unit = umw.cell(row=r, column=2).value
            if not label or not isinstance(label, str):
                continue
            out.append((_norm_label(label), str(unit or "").lower().strip(), r))
        return out

    section_idx: Dict[str, List[Tuple[str, str, int]]] = {
        flow: _build_row_index(flow) for flow in SECTION_ROWS
    }

    def _unit_matrix(unit: str) -> str:
        """Return ``"Feststoff"`` / ``"Eluat"`` / ``""`` from a unit string."""
        if is_feststoff_unit(unit):
            return "Feststoff"
        if is_eluat_unit(unit):
            return "Eluat"
        return ""

    def _rank_target_rows(label: str, lab_unit: str,
                          primary: List[Tuple[str, str, int]],
                          fallback: List[Tuple[str, str, int]],
                          strict: bool = False) -> List[Tuple[float, int]]:
        """Score every candidate row and return ``[(score, row), ...]``
        sorted by score descending, filtered above the cutoff.

        Lets the caller pick the next-best row when a higher-scoring row
        is already claimed by another lab candidate. Returns from the
        FIRST non-empty index view (primary preferred over fallback) —
        ranking across both would let weak fallback hits dilute strong
        primary ones.

        Args:
            label: raw lab-report label text.
            lab_unit: unit string from the lab row (drives matrix gating).
            primary: indexed rows of the lab's own section (PAK/EBV/Aggr).
            fallback: indexed rows of a fall-through section (Aggr → EBV).
            strict: when True, only counts the exact / tight / paren-eq
                passes (used by Phase A pre-locking).

        Returns:
            List of ``(score, row_idx)`` tuples in descending score order.
            Empty list if no row exceeds the cutoff.
        """
        ln = _norm_label(label)
        if not ln:
            return []
        ln_tight = _norm_label_tight(label)
        lab_matrix = _unit_matrix(lab_unit)

        def _strip_parens(s: str) -> str:
            return re.sub(r"\s*\([^)]*\)\s*", " ", s or "").strip()

        ln_np = _strip_parens(ln)
        ln_np_tokens = set(_fold_umlauts(t) for t in ln_np.split() if len(t) > 2)
        ln_tokens = set(_fold_umlauts(t) for t in ln.split() if len(t) > 2)

        def _matrix_compat(ukey: str) -> bool:
            if not lab_matrix:
                return True
            tpl_matrix = _unit_matrix(ukey)
            return (not tpl_matrix) or (tpl_matrix == lab_matrix)

        def _score_one(lkey: str, ukey: str) -> float:
            if not lkey:
                return 0.0
            lkey_tight = lkey.replace(" ", "")
            lkey_np = _strip_parens(lkey)
            lkey_np_tokens = set(_fold_umlauts(t) for t in lkey_np.split() if len(t) > 2)

            if lkey == ln:
                return 1000.0
            if lkey_tight == ln_tight:
                return 950.0
            if lkey_np and lkey_np == ln_np:
                return 900.0
            if strict:
                return 0.0

            def ratio(a: int, b: int) -> float:
                if a == 0 or b == 0:
                    return 0.0
                return min(a, b) / max(a, b)

            best = 0.0
            # Direction A: lab fully INSIDE template label (template more specific).
            if ln and ln in lkey:
                rA = ratio(len(ln), len(lkey))
                # Penalty for low-coverage matches: a very short lab label
                # buried in a long template label (e.g. lab "Naphthalin" inside
                # template "Summe Naphthaline nach EBV") usually means the
                # template row is too specific — not a real match.
                if rA >= 0.6:
                    best = max(best, 800.0 + 100.0 * rA)
                else:
                    best = max(best, 450.0 + 100.0 * rA)
            # Direction B: template fully INSIDE lab label (template less specific).
            # Heavily penalised when coverage is low: a short template label
            # ("Naphthalin", 10 chars) absorbing a long lab string ("Summe
            # Naphthaline (EBV)", 22 chars) is the exact failure mode we want
            # to avoid — that long lab string belongs on the dedicated
            # "Summe Naphthaline nach EBV" row.
            if lkey and lkey in ln:
                rB = ratio(len(lkey), len(ln))
                if rB >= 0.7:
                    best = max(best, 700.0 + 100.0 * rB)
                else:
                    best = max(best, 300.0 + 100.0 * rB)
            # Paren-stripped substring (handles "(Eluat)" / "(EBV)" suffixes).
            if ln_np and lkey_np and ln_np in lkey_np:
                best = max(best, 500.0 + 100.0 * ratio(len(ln_np), len(lkey_np)))
            if ln_np and lkey_np and lkey_np in ln_np:
                best = max(best, 400.0 + 100.0 * ratio(len(lkey_np), len(ln_np)))
            # Token-overlap (umlaut-folded). Symmetric: both directions count.
            if ln_np_tokens and lkey_np_tokens:
                if len(lkey_np_tokens) >= 2 and lkey_np_tokens.issubset(ln_np_tokens):
                    best = max(best, 200.0 + 80.0 * (len(lkey_np_tokens) / max(1, len(ln_np_tokens))))
                if len(ln_np_tokens) >= 2 and ln_np_tokens.issubset(lkey_np_tokens):
                    best = max(best, 200.0 + 80.0 * (len(ln_np_tokens) / max(1, len(lkey_np_tokens))))
            # Synonym bridge
            for a, b in LABEL_SYNONYMS:
                a_np = _strip_parens(a)
                b_np = _strip_parens(b)
                if ((a_np == ln_np or a_np in ln_np or ln_np in a_np)
                    and (b_np in lkey_np or lkey_np in b_np)):
                    best = max(best, 300.0)
                if ((b_np == ln_np or b_np in ln_np or ln_np in b_np)
                    and (a_np in lkey_np or lkey_np in a_np)):
                    best = max(best, 300.0)
            # Unit-equality tiebreaker: when two rows score identically on
            # label similarity but have different units, prefer the row
            # whose unit matches the lab unit verbatim (Gesamthärte mg/l
            # vs Gesamthärte mmol/l disambiguation).
            if best > 0 and lab_unit and ukey:
                ul = str(lab_unit).strip().lower().replace(" ", "")
                uk = str(ukey).strip().lower().replace(" ", "")
                if ul and uk and ul == uk:
                    best += 50.0
            return best

        CUTOFF = 700.0 if strict else 180.0
        for idx in (primary, fallback):
            if not idx:
                continue
            mx_idx = [(l, u, r) for (l, u, r) in idx if _matrix_compat(u)]
            views_to_try = (mx_idx,) if lab_matrix else (mx_idx, idx)
            for view in views_to_try:
                if not view:
                    continue
                ranked: List[Tuple[float, int]] = []
                for lkey, ukey, r in view:
                    s = _score_one(lkey, ukey)
                    if s > CUTOFF:
                        ranked.append((s, r))
                if ranked:
                    ranked.sort(key=lambda x: -x[0])
                    return ranked
        return []

    def _find_target_row(label: str, lab_unit: str,
                         primary: List[Tuple[str, str, int]],
                         fallback: List[Tuple[str, str, int]],
                         strict: bool = False,
                         skip_rows: Optional[set] = None) -> Optional[int]:
        """Return the highest-ranked target row not in ``skip_rows``.

        Thin wrapper over :func:`_rank_target_rows` that walks the ranked
        list and returns the first row not yet claimed by another lab
        candidate. Lets Phase B fall through to the next-best row when its
        top pick was already locked in Phase A.

        Args:
            label: raw lab-report label text.
            lab_unit: unit string from the lab row.
            primary: indexed rows of the lab's own section.
            fallback: indexed rows of a fall-through section.
            strict: same semantics as :func:`_rank_target_rows`.
            skip_rows: optional set of row indices to ignore (already
                claimed). If exhausted, returns None.

        Returns:
            The 1-based row index of the best unclaimed match, or None.
        """
        ranked = _rank_target_rows(label, lab_unit, primary, fallback, strict)
        if not ranked:
            return None
        if not skip_rows:
            return ranked[0][1]
        for _score, r in ranked:
            if r not in skip_rows:
                return r
        return None

    # Assign per-sample columns: PAK first (cols C..), then EBV, then Aggr.
    sample_blocks: List[Tuple[str, List[Tuple[str, pd.DataFrame]]]] = [
        ("PAK", pak_samples),
        ("EBV", ebv_samples),
        ("Aggr", aggr_samples),
    ]
    material_label = {
        "PAK": "Straßenaufbruch",
        "EBV": "Auffüllung",
        "Aggr": "Grundwasser",
    }

    col = EXAMPLE_COL_START  # C
    for flow, samples in sample_blocks:
        for name, df in samples:
            # Header rows 1-4
            umw.cell(row=1, column=col, value=material_label[flow])
            probe = ""
            if "Probenbezeichnung" in df.columns and len(df) > 0:
                v = df["Probenbezeichnung"].iloc[0]
                probe = str(v) if pd.notna(v) else ""
            labor_num = ""
            if "Labor_Nummer" in df.columns and len(df) > 0:
                v = df["Labor_Nummer"].iloc[0]
                labor_num = str(v) if pd.notna(v) else ""
            elif "Analysennr" in df.columns and len(df) > 0:
                v = df["Analysennr"].iloc[0]
                labor_num = str(v) if pd.notna(v) else ""
            umw.cell(row=2, column=col, value=labor_num)
            umw.cell(row=4, column=col, value=probe or name)

            # Map each parsed row to a canonical template row WITHIN this
            # flow's section (PAK/EBV/Aggr). For Aggressivität samples, also
            # try the EBV section as a fall-through (water labs sometimes
            # include params shared with EBV like Sulfat).
            primary_section = flow  # "PAK" | "EBV" | "Aggr"
            primary_idx = section_idx[primary_section]
            fallback_idx = section_idx["EBV"] if primary_section == "Aggr" else {}

            # Noise prefixes / substrings to drop before matching — temperature
            # auxiliaries, SGS footer text, address lines etc. would otherwise
            # leak into the canonical rows via substring match (e.g.
            # "Bei-Temperatur für pH-Wert" would hit the pH-Wert row).
            _NOISE_PREFIXES = (
                "bei", "sgs ", "gbm ", "metechnik", "datum:", "prfbericht",
                "auftrag-nr", "ihr auftrag", "projekt:", "eingangsdatum",
                "untersuchungs", "prfzeitraum:", "der prfbericht",
                "ohne unterschrift", "probenbezeichnung:", "probe nr",
                "polycyclische", "hinweis", "sofern nicht", "ausschlielich",
                "verffentl", "sonstigen fll", "dieses dokument", "www.sgs",
                "zum gerichts", "ucp 600", "zeitpunkt", "wiedergeben",
                "aber nicht", "oder des ueren", "regel fr die", "analytischen",
                "bestimmungsgrenze pausch", "76275 ettlingen", "herr david",
                "nobelstrae", "standort", "(f) -",
            )
            _NOISE_SUBSTR = ("temperatur",)

            def _fmt_de(s: str) -> str:
                """Format value string with German decimal comma.

                Below-detection markers like ``n.n.`` / ``n.b.`` / ``n.d.``
                are passed through verbatim — only DECIMAL dots are
                replaced with commas (digit before AND after the dot).
                """
                if not s:
                    return s
                # Replace only "<digit>.<digit>" → "<digit>,<digit>"
                return re.sub(r"(\d)\.(\d)", r"\1,\2", s)

            _HEADER_LABELS = {"eluat", "feststoff", "königswasseraufschluss",
                              "knigswasseraufschluss", "aufschlussfaktor",
                              "probenvorbereitung", "siebung < 2 mm",
                              "summarische parameter", "anionen", "kationen",
                              "berechnete werte", "sensorische prüfungen",
                              "physikalisch-chemische parameter"}

            # Pre-build a candidate list (label, unit, op, val, has_real,
            # display) once per sample — cheaper than recomputing for each
            # phase and lets us reason about ordering.
            candidates: List[Tuple[str, str, str, object, bool, str]] = []
            for _, lab_row in df.iterrows():
                label = str(lab_row.get("Lab_Original_String", "")).strip()
                unit = str(lab_row.get("Lab_Unit", "")).strip()
                if not label:
                    continue
                label_lower = label.lower()
                if any(label_lower.startswith(p) for p in _NOISE_PREFIXES):
                    continue
                if any(s in label_lower for s in _NOISE_SUBSTR):
                    continue
                if label_lower.strip() in _HEADER_LABELS:
                    continue
                op = str(lab_row.get("Lab_Operator", "")).strip()
                val = lab_row.get("Lab_Value", None)
                has_real = (val is not None) and not (isinstance(val, float) and pd.isna(val))
                # Lab_Display_Override carries textual results that the
                # numeric parse path discards (Färbung "farblos", Trübung
                # "klar mit Bodensatz", Lab DIN 4030 verdict "nicht
                # angreifend"). When present, treat it as content even
                # though Lab_Value is None.
                override_raw = lab_row.get("Lab_Display_Override", "") if "Lab_Display_Override" in lab_row.index else ""
                override = str(override_raw).strip() if override_raw is not None else ""
                if not has_real and op.strip().lower() not in {"< bg", "<bg"} and not override:
                    continue
                op_lower = op.lower().replace(" ", "")
                if override:
                    # Textual result wins over the < BG default.
                    display = override
                    has_real = True  # so downstream "skip empty" gates pass
                elif "bg" in op_lower or op_lower in {"n.n.", "n.b.", "n.d.", "nn", "nb", "nd"}:
                    display = "n.n."
                elif has_real:
                    display = f"{op}{val}" if op else str(val)
                else:
                    display = "n.n."
                # German decimal comma only for numeric content; leave
                # textual overrides (e.g. "klar mit Bodensatz") untouched.
                if not override:
                    display = _fmt_de(display.replace(" ", ""))
                candidates.append((label, unit, op, val, has_real, display))

            # First-write-wins: track which rows have been claimed in THIS
            # sample column so a later substring/fuzzy match cannot
            # overwrite an earlier exact match. Two-phase resolution:
            #   Phase A — strict pass (exact + tight + paren-stripped eq only)
            #             locks high-confidence rows.
            #   Phase B — full scored matching fills the rest, skipping
            #             any row already claimed in Phase A or by an
            #             earlier Phase-B candidate.
            claimed_rows: set[int] = set()

            def _try_write(target_row: int, has_real: bool, display: str) -> bool:
                """Write to target_row if free; return True if a write occurred."""
                if target_row in claimed_rows:
                    return False
                tpl_unit = umw.cell(target_row, 2).value
                if not tpl_unit and not has_real:
                    return False
                cell = umw.cell(row=target_row, column=col)
                if isinstance(cell, MergedCell):
                    return False
                cell.value = display
                claimed_rows.add(target_row)
                return True

            # Phase A: strict matches only (exact / tight / paren-stripped eq).
            # Pass ``claimed_rows`` so a strict candidate whose top template
            # row was already claimed by an earlier strict match falls
            # through to the next-best strict row — needed e.g. for
            # Gesamthärte mg/l (r127) vs Gesamthärte mmol/l (r128), which
            # share a paren-stripped label and differ only in unit.
            phase_b_queue: List[Tuple[str, str, str, object, bool, str]] = []
            for label, unit, op, val, has_real, display in candidates:
                target_row = _find_target_row(
                    label, unit, primary_idx, fallback_idx,
                    strict=True, skip_rows=claimed_rows,
                )
                if target_row is not None:
                    if not _try_write(target_row, has_real, display):
                        # All strict candidates claimed — defer to Phase B
                        # so a looser score can still find a home for this
                        # row (rather than dropping it silently).
                        phase_b_queue.append((label, unit, op, val, has_real, display))
                else:
                    phase_b_queue.append((label, unit, op, val, has_real, display))

            # Phase B: full scored matching for everything Phase A didn't lock.
            # Pass `claimed_rows` as ``skip_rows`` so a candidate whose top
            # match is already claimed falls through to the next-best row
            # instead of being silently dropped (e.g. "Summe Naphthaline
            # (EBV)" whose top score might be the already-locked
            # ``Naphthalin`` row must still land on
            # ``Summe Naphthaline nach EBV``).
            for label, unit, op, val, has_real, display in phase_b_queue:
                target_row = _find_target_row(
                    label, unit, primary_idx, fallback_idx,
                    strict=False, skip_rows=claimed_rows,
                )
                if target_row is not None:
                    _try_write(target_row, has_real, display)

            col += 1

    # Section-aware border + style extension.
    #
    # Master template's visible table boundaries come from:
    #   * thin top border on every parameter row (cols 1..end)
    #   * thin LEFT border on the first column of each flow section (E for
    #     EBV-start, I for Aggr-start in the master's 6-EBV+1-Grundwasser
    #     example)
    # Real projects vary in sample counts per flow, so we recompute the
    # section starts from how many samples each flow actually has, then:
    #   1. clear any stray "section-start" L-borders on cols inside a section
    #   2. apply the L-border on the actual section-start columns
    #   3. clone the top-border + font/fill styling from a known-good cell
    #      across every populated sample column so the table grid extends
    #      uniformly to the rightmost sample.
    from copy import copy as _copycell
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter

    pak_n = len(pak_samples)
    ebv_n = len(ebv_samples)
    aggr_n = len(aggr_samples)
    first_col = EXAMPLE_COL_START               # = 3 (col C)
    pak_first = first_col
    ebv_first = first_col + pak_n
    aggr_first = first_col + pak_n + ebv_n
    last_col = first_col + pak_n + ebv_n + aggr_n - 1
    section_starts = {pak_first, ebv_first, aggr_first}

    if last_col >= first_col:
        # Style reference cell for each row — col E (5) in master has the
        # canonical per-row formatting we want to propagate.
        ref_col = 5
        thin = Side(border_style="thin", color="808080")
        for r in range(1, 132):
            ref_cell = umw.cell(row=r, column=ref_col)
            if isinstance(ref_cell, MergedCell):
                continue
            for c in range(first_col, last_col + 1):
                tgt = umw.cell(row=r, column=c)
                if isinstance(tgt, MergedCell):
                    continue
                # Propagate styling from the reference cell
                if ref_cell.has_style:
                    tgt.font = _copycell(ref_cell.font)
                    tgt.alignment = _copycell(ref_cell.alignment)
                    tgt.fill = _copycell(ref_cell.fill)
                    tgt.number_format = ref_cell.number_format
                    tgt.protection = _copycell(ref_cell.protection)
                # Border: top from ref_cell, left only on section starts,
                # right only on the very last sample column.
                left_side = thin if c in section_starts else None
                right_side = thin if c == last_col else None
                top_side = _copycell(ref_cell.border.top) if (ref_cell.border and ref_cell.border.top) else thin
                bottom_side = _copycell(ref_cell.border.bottom) if (ref_cell.border and ref_cell.border.bottom) else None
                tgt.border = Border(
                    left=left_side,
                    right=right_side,
                    top=top_side,
                    bottom=bottom_side,
                )

        # Column widths: match the master's col-E width across all sample cols.
        h_width = umw.column_dimensions["E"].width or 13
        for c in range(first_col, last_col + 1):
            umw.column_dimensions[get_column_letter(c)].width = h_width

    # Per-sample tabs with flow prefix.
    # Sheet names use the Probenbezeichnung (e.g. "MP03") rather than the
    # raw filename — avoids 31-char-truncation collisions that previously
    # caused Excel to flag the workbook as "needs recovery".
    used_titles: set[str] = set()
    for flow_tag, samples in sample_blocks:
        for name, df in samples:
            probe = ""
            if "Probenbezeichnung" in df.columns and len(df) > 0:
                v = df["Probenbezeichnung"].iloc[0]
                probe = str(v).strip() if pd.notna(v) else ""
            base = probe or name
            # Clean characters Excel forbids: \ / ? * : [ ]
            base = re.sub(r"[\\/?\*:\[\]]", "_", base)
            tab_name = f"{flow_tag}_{base}"[:31]
            # Disambiguate if multiple samples normalise to the same probe
            i = 2
            unique = tab_name
            while unique in used_titles:
                suffix = f"_{i}"
                unique = (tab_name[: 31 - len(suffix)]) + suffix
                i += 1
            used_titles.add(unique)
            ws = wb.create_sheet(title=unique)
            # Strip ASCII control chars (\x00-\x08, \x0b-\x0c, \x0e-\x1f) from
            # any string cell — PDF text extraction occasionally yields these
            # and openpyxl rejects them with IllegalCharacterError. Keep \t,
            # \n, \r since they are valid in cells.
            _ctrl_re = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
            for r in dataframe_to_rows(df, index=False, header=True):
                cleaned = [
                    _ctrl_re.sub("", v) if isinstance(v, str) else v
                    for v in r
                ]
                ws.append(cleaned)

    xlsx_path = os.path.join(session_dir, "Validation.xlsx")
    wb.save(xlsx_path)

    # Augment each Aggr_<sample> sheet with the DIN 4030 + DIN 50929
    # calculations so the user can manually verify the classification.
    try:
        from aggr_validation_writer import augment_validation as _aug
        n = _aug(xlsx_path)
        if n > 0:
            print(f"  Aggr-Berechnungen ergänzt in {n} Validierungssheet(s).")
    except Exception as _exc:
        print(f"  WARN: Aggr-Berechnung konnte nicht angehängt werden: {_exc}")

    # HTML preview — concatenate per-sample tables for human review
    html_path = os.path.join(session_dir, "Validation.html")
    html_content = """<!DOCTYPE html><html><head><meta charset="utf-8"/><style>
    body { font-family: Helvetica, sans-serif; font-size: 13px; line-height: 1.4; margin: 20px; }
    table { border-collapse: collapse; width: 100%; margin-bottom: 30px; }
    th, td { border: 1px solid #a0a0a0; padding: 6px; text-align: left; }
    th { background-color: #e0e0e0; font-weight: bold; border-bottom: 2px solid #555; }
    h2 { border-bottom: 2px solid #000; padding-bottom: 5px; margin-top: 40px;}
    </style></head><body><h1>Validation Data Extraction</h1>"""
    for flow_tag, samples in sample_blocks:
        for name, df in samples:
            html_content += f"<h2>[{flow_tag}] {name}</h2>"  # bracket fine in HTML
            html_content += df.to_html(index=False)
    html_content += "</body></html>"
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    return xlsx_path, html_path


def _process_ebv(input_dir: str = INPUT_DIR_EBV, output_dir: str = OUTPUT_DIR_EBV, flow_name: str = "EBV") -> bool:
    """Run the EBV branch of Step 1.

    Parses every ``*.pdf`` in ``input_dir`` via :func:`extract_all_data_from_pdf`,
    builds a validation workbook with one sheet per sample plus a ``_Project``
    sheet and (best-effort) a transposed ``UMWELT`` echo sheet, and writes it
    to a timestamped subfolder of ``output_dir``.

    Args:
        input_dir: directory containing EBV lab-report PDFs.
        output_dir: parent directory for the timestamped validation folder.

    Returns:
        True if at least one PDF produced usable data; False if the branch
        had nothing to do (missing folder or no PDFs).

    Raises:
        OSError: if ``output_dir`` cannot be created.
    """
    print("STEP 1 [EBV]: Starting data extraction from PDF reports...")

    if not os.path.exists(input_dir):
        print(
            f"Directory '{input_dir}' not found. Creating it now. "
            "Please add your PDFs and restart."
        )
        os.makedirs(input_dir, exist_ok=True)
        return False

    pdf_files = sorted(glob.glob(os.path.join(input_dir, "*.pdf")))

    if not pdf_files:
        print(f"No PDF files found in '{input_dir}'.")
        return False

    timestamp = _berlin_now().strftime("%Y-%m-%d_%H-%M")
    session_dir = os.path.join(output_dir, f"{timestamp}_Validierung")
    os.makedirs(session_dir, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    files_processed = 0
    all_dfs: Dict[str, pd.DataFrame] = {}
    sheet_meta: Dict[str, Tuple[str, str, str]] = {}
    projektnummer_first: str = ""

    for pdf_path in pdf_files:
        filename = os.path.basename(pdf_path)
        sheet_name = filename[:31]
        print(f"  - Processing: {filename}")

        projektnummer, probenbezeichnung, labor_nummer = _extract_metadata_from_filename(filename)
        if not projektnummer_first and projektnummer:
            projektnummer_first = projektnummer

        df = extract_all_data_from_pdf(pdf_path)

        if not df.empty:
            df.insert(0, "Soil_Type", "undef")
            df.insert(1, "Probenbezeichnung", probenbezeichnung)
            df.insert(2, "Petrographische_Beschreibung", "")
            df.insert(3, "Stratigraphie", "")
            df.insert(4, "Labor_Nummer", labor_nummer)

            columns_order = [
                "Soil_Type",
                "Probenbezeichnung",
                "Petrographische_Beschreibung",
                "Stratigraphie",
                "Labor_Nummer",
                "EBV_Parameter",
                "Matrix",
                "Lab_Unit",
                "Lab_Original_String",
                "Lab_Operator",
                "Lab_Value",
            ]
            for col in columns_order:
                if col not in df.columns:
                    df[col] = ""

            remaining_cols = [c for c in df.columns if c not in columns_order]
            df = df[columns_order + remaining_cols]
            all_dfs[sheet_name] = df
            sheet_meta[sheet_name] = (probenbezeichnung, labor_nummer, "")

            ws = wb.create_sheet(title=sheet_name)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            files_processed += 1

    if files_processed == 0:
        print("\nNo relevant EBV data could be extracted from the provided PDFs.")
        return False

    # Insert _Project sheet at position 0 with auto-extracted Projektnummer
    _write_project_sheet(wb, projektnummer_default=projektnummer_first)

    # Insert transposed UMWELT echo sheet at position 1 (after _Project)
    try:
        _write_umwelt_echo_sheet(wb, all_dfs, sheet_meta)
    except Exception as e:  # noqa: BLE001
        # UMWELT echo is best-effort — never block the run on a styling issue
        print(f"  -> WARNING: UMWELT echo sheet could not be generated: {e}")

    out_excel = os.path.join(session_dir, f"{flow_name}.xlsx")
    wb.save(out_excel)

    # HTML preview (unchanged from previous build)
    out_html = os.path.join(session_dir, f"{flow_name}.html")
    html_content = """<!DOCTYPE html><html><head><meta charset="utf-8"/><style>
    body { font-family: Helvetica, sans-serif; font-size: 13px; line-height: 1.4; margin: 20px; }
    table { border-collapse: collapse; width: 100%; margin-bottom: 30px; }
    th, td { border: 1px solid #a0a0a0; padding: 6px; text-align: left; }
    th { background-color: #e0e0e0; font-weight: bold; border-bottom: 2px solid #555; }
    h2 { border-bottom: 2px solid #000; padding-bottom: 5px; margin-top: 40px;}
    </style></head><body><h1>Validation Data Extraction</h1>"""
    for name, d_frame in all_dfs.items():
        html_content += f"<h2>Sample: {name}</h2>"
        html_content += d_frame.to_html(index=False)
    html_content += "</body></html>"
    with open(out_html, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"\n[{flow_name}] Extraction completed - processed {files_processed} file(s).")
    print(f"-> {out_excel}")
    print(f"-> {out_html}")
    print(
        "\nIMPORTANT: Open the validation Excel and fill in the '_Project' sheet "
        "(Bauvorhaben, LOS, Bauwerk) plus per-sample 'Petrographische_Beschreibung' "
        "and 'Stratigraphie' before running Step 2. "
        "The 'UMWELT' sheet is a read-only verification echo; do not edit it."
    )
    return True


def _write_umwelt_echo_sheet(
    wb: Workbook,
    sample_dfs: Dict[str, pd.DataFrame],
    sheet_meta: Dict[str, Tuple[str, str, str]],
) -> None:
    """Insert the canonical ``UMWELT`` echo sheet into the validation workbook.

    Layout mirrors the company workbook's UMWELT sheet (rows 1..64 of
    ``2604XX_Rohdaten & Aggressivität.xlsx``). The row order is the fixed
    canonical list from :mod:`umwelt_template` — every project's UMWELT
    looks identical regardless of which lab parameters were reported.
    Parsed PDF rows are mapped to canonical rows via
    :func:`umwelt_template.match_lab_row`; lab rows that don't match any
    canonical entry are appended at the bottom under a
    "Weitere Parameter (lab-spezifisch)" section so nothing is dropped
    silently.

    Cell layout:

        Col A: section heading or parameter display name
        Col B: unit (canonical)
        Col C: lab raw label (only for unmapped append rows — blank
               for canonical rows)
        Col D..: per-sample columns. Row 1 carries the material category
               (placeholder), row 2 the Probennummer/Labor_Nummer, row 4
               the Probenbezeichnung. Sample data starts at row 5.

    Args:
        wb: target validation workbook.
        sample_dfs: mapping ``sheet_name -> parsed DataFrame`` (one per sample).
        sheet_meta: mapping ``sheet_name -> (probenbezeichnung, labor_nummer, material)``.
    """
    from umwelt_template import UMWELT_TEMPLATE, match_lab_row

    sheet_title = "UMWELT"
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    insert_idx = 1 if "_Project" in wb.sheetnames else 0
    ws = wb.create_sheet(sheet_title, insert_idx)

    header_font = Font(bold=True)
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="305496")
    header_fill = PatternFill("solid", fgColor="E0E0E0")

    # Header column A meta-rows (row 1..4)
    ws.cell(row=1, column=1, value="Material").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=2, column=1, value="Probennummer").font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=3, column=1, value="Eingang").font = header_font
    ws.cell(row=3, column=1).fill = header_fill
    ws.cell(row=4, column=1, value="Probenbezeichnung").font = header_font
    ws.cell(row=4, column=1).fill = header_fill

    # Canonical template rows starting at row 5
    canonical_row_index: Dict[int, int] = {}  # template_idx -> excel_row
    excel_row = 5
    for tpl_idx, row in enumerate(UMWELT_TEMPLATE):
        if row.kind == "blank":
            excel_row += 1
            continue
        if row.kind == "section":
            ws.cell(row=excel_row, column=1, value=row.label).font = section_font
            ws.cell(row=excel_row, column=1).fill = section_fill
            ws.merge_cells(start_row=excel_row, start_column=1,
                           end_row=excel_row, end_column=max(3, 3 + len(sample_dfs)))
            excel_row += 1
            continue
        # param row
        ws.cell(row=excel_row, column=1, value=row.label).font = header_font
        ws.cell(row=excel_row, column=2, value=row.unit).font = header_font
        canonical_row_index[tpl_idx] = excel_row
        excel_row += 1

    # Per-sample columns start at column D (col 4).
    # Column C is reserved for raw lab labels of unmapped rows.
    sample_first_col = 4

    # Collect unmapped rows across all samples (de-dup by (label, unit))
    unmapped_keys: List[Tuple[str, str]] = []
    unmapped_seen: set[Tuple[str, str]] = set()

    # Per-sample-column writes — fill canonical rows + collect unmapped
    for sample_idx, (sheet_name, df) in enumerate(sample_dfs.items()):
        col = sample_first_col + sample_idx
        probe, labor, material = sheet_meta.get(sheet_name, (sheet_name, "", ""))
        ws.cell(row=1, column=col, value=material or "")
        ws.cell(row=2, column=col, value=labor or "")
        ws.cell(row=4, column=col, value=probe or sheet_name).font = header_font

        # Track unmapped per sample for later append
        sample_unmapped_values: Dict[Tuple[str, str], str] = {}

        if "Lab_Original_String" not in df.columns:
            continue
        for _, lab_row in df.iterrows():
            label = str(lab_row.get("Lab_Original_String", "")).strip()
            unit = str(lab_row.get("Lab_Unit", "")).strip()
            ebv = str(lab_row.get("EBV_Parameter", "")).strip()
            if not label:
                continue
            # Pre-filter: drop rows that aren't real chemistry — temperature
            # auxiliaries, lab metadata, footer boilerplate. These leak into
            # the canonical schema via substring match (e.g. "Bei-Temperatur
            # für pH-Wert" matches "ph-wert").
            label_lower = label.lower()
            _NOISE_PREFIXES = ("bei", "sgs ", "gbm ", "metechnik", "datum:", "prfbericht",
                               "auftrag-nr", "ihr auftrag", "projekt:", "eingangsdatum",
                               "untersuchungs", "prfzeitraum:", "der prfbericht",
                               "ohne unterschrift", "probenbezeichnung:", "probe nr",
                               "parameter", "polycyclische", "hinweis", "sofern nicht",
                               "ausschlielich", "verffentl", "sonstigen fll",
                               "dieses dokument", "www.sgs", "zum gerichts",
                               "ucp 600", "zeitpunkt", "wiedergeben", "aber nicht",
                               "oder des ueren", "regel fr die", "analytischen",
                               "bestimmungsgrenze pausch", "76275 ettlingen",
                               "herr david", "nobelstrae", "standort",
                               "(f) -")
            _NOISE_SUBSTR = ("temperatur",)
            if any(label_lower.startswith(p) for p in _NOISE_PREFIXES):
                continue
            if any(s in label_lower for s in _NOISE_SUBSTR):
                continue
            op = str(lab_row.get("Lab_Operator", "")).strip()
            val = lab_row.get("Lab_Value", None)
            has_real_value = (val is not None) and not (isinstance(val, float) and pd.isna(val))
            if not has_real_value:
                # Drop unless the operator is a meaningful below-detection marker
                if op.strip().lower() not in {"< bg", "<bg"}:
                    continue
                display = "< BG"
            else:
                display = f"{op}{val}".replace(" ", "") if op else str(val)
            if not display:
                continue

            tpl_idx = match_lab_row(label, unit, ebv)
            if tpl_idx >= 0 and tpl_idx in canonical_row_index:
                target_row = canonical_row_index[tpl_idx]
                ws.cell(row=target_row, column=col, value=display)
            elif has_real_value:
                # Only append unmapped rows that carry a real measurement.
                # Pure "< BG" parser-noise rows (footer text, addresses, etc.)
                # are dropped silently — they're not chemistry.
                key = (label, unit)
                sample_unmapped_values[key] = display
                if key not in unmapped_seen:
                    unmapped_seen.add(key)
                    unmapped_keys.append(key)

        # Stash for later write — defer writing until all keys collected
        df.attrs["__umwelt_unmapped"] = sample_unmapped_values  # type: ignore

    # Append "Weitere Parameter" section + unmapped rows below canonical block
    if unmapped_keys:
        excel_row += 1
        ws.cell(row=excel_row, column=1, value="Weitere Parameter (lab-spezifisch, kein Standard-UMWELT-Eintrag)").font = section_font
        ws.cell(row=excel_row, column=1).fill = section_fill
        ws.merge_cells(start_row=excel_row, start_column=1,
                       end_row=excel_row, end_column=max(3, 3 + len(sample_dfs)))
        excel_row += 1
        unmapped_excel_rows: Dict[Tuple[str, str], int] = {}
        for label, unit in unmapped_keys:
            ws.cell(row=excel_row, column=1, value=label).font = header_font
            ws.cell(row=excel_row, column=2, value=unit).font = header_font
            ws.cell(row=excel_row, column=3, value="(unmapped)").font = Font(italic=True, color="888888", size=9)
            unmapped_excel_rows[(label, unit)] = excel_row
            excel_row += 1

        for sample_idx, (sheet_name, df) in enumerate(sample_dfs.items()):
            col = sample_first_col + sample_idx
            unmapped_values = df.attrs.get("__umwelt_unmapped", {})
            for key, display in unmapped_values.items():
                r = unmapped_excel_rows.get(key)
                if r is not None:
                    ws.cell(row=r, column=col, value=display)

    # Column widths
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    for sample_idx in range(len(sample_dfs)):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(sample_first_col + sample_idx)
        ws.column_dimensions[col_letter].width = 18


def _process_aggressivität(
    input_dir: str = INPUT_DIR_AGGR,
    output_dir: str = OUTPUT_DIR_AGGR,
) -> bool:
    """Run the Aggressivität branch of Step 1.

    Parses every ``*.pdf`` in ``input_dir`` via
    :func:`pdf_parser_aggressivität.extract_all_data_from_pdf`, builds a
    validation workbook with one sheet per sample plus ``_Project`` and a
    transposed ``UMWELT_Aggr`` echo sheet, and writes it to a timestamped
    subfolder of ``output_dir``.

    Output schema per sample sheet (column order, all required by Step 2):

        Probenbezeichnung, Analysennr, Aggr_Parameter, Lab_Unit,
        Lab_Operator, Lab_Value, Lab_Verdict_Text, Full_Row

    Args:
        input_dir: directory containing Aggressivität lab-report PDFs.
        output_dir: parent directory for the timestamped validation folder.

    Returns:
        True if at least one PDF produced usable data; False if the folder
        was missing or no PDFs were present.
    """
    # Local import keeps the EBV branch decoupled if pdf_parser_aggressivität
    # ever fails to import (e.g. config_aggressivität typo).
    from pdf_parser_aggressivität import (
        extract_all_data_from_pdf as extract_aggr_data,
        extract_probenbezeichnung as extract_aggr_probe,
    )
    from config_aggressivität import WATER_AGGR_PARAMETERS, LAB_DIN4030_KEY

    print("STEP 1 [Aggressivität]: Starting data extraction from PDF reports...")

    if not os.path.exists(input_dir):
        print(
            f"Directory '{input_dir}' not found. Creating it now. "
            "Please add your Aggressivität PDFs and restart."
        )
        os.makedirs(input_dir, exist_ok=True)
        return False

    pdf_files = sorted(glob.glob(os.path.join(input_dir, "*.pdf")))
    if not pdf_files:
        print(f"No PDF files found in '{input_dir}'. Skipping branch.")
        return False

    session_dir = _get_session_dir_validation()

    wb = Workbook()
    wb.remove(wb.active)

    files_processed = 0
    all_dfs: Dict[str, pd.DataFrame] = {}
    sample_meta: Dict[str, Tuple[str, str, str]] = {}  # sheet -> (probe, analysennr, material)
    projektnummer_first: str = ""

    for pdf_path in pdf_files:
        filename = os.path.basename(pdf_path)
        sheet_name = filename[:31]
        print(f"  - Processing: {filename}")

        projektnummer, _, _ = _extract_metadata_from_filename(filename)
        if not projektnummer_first and projektnummer:
            projektnummer_first = projektnummer

        probe = extract_aggr_probe(pdf_path) or filename
        df = extract_aggr_data(pdf_path)
        if df.empty:
            print(f"  -> WARNING: no Aggressivität parameters recognised in {filename}")
            continue

        # Front-fill the sample's per-row metadata
        df.insert(0, "Probenbezeichnung", probe)
        df.insert(1, "Analysennr", "")  # placeholder; user fills if needed
        df.insert(2, "Petrographische_Beschreibung", "")
        df.insert(3, "Stratigraphie", "")
        df.insert(4, "Tiefe", "")
        df.insert(5, "Formation", "")
        df.insert(6, "Sample_Type", "Wasser")  # only water flow currently supported

        all_dfs[sheet_name] = df
        sample_meta[sheet_name] = (probe, "", "Grundwasser")

        ws = wb.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        files_processed += 1

    if files_processed == 0:
        print("\nNo Aggressivität data could be extracted from the provided PDFs.")
        return False

    # _Project sheet (same convention as EBV)
    _write_project_sheet(wb, projektnummer_default=projektnummer_first)

    # Transposed UMWELT_Aggr echo (params as rows, samples as columns)
    try:
        _write_umwelt_aggr_echo(wb, all_dfs, sample_meta, WATER_AGGR_PARAMETERS, LAB_DIN4030_KEY)
    except Exception as e:  # noqa: BLE001
        print(f"  -> WARNING: UMWELT_Aggr echo sheet could not be generated: {e}")

    out_excel = os.path.join(session_dir, "Aggressivität.xlsx")
    wb.save(out_excel)

    print(f"\n[Aggressivität] Extraction completed - processed {files_processed} file(s).")
    print(f"-> {out_excel}")
    print(
        "\nIMPORTANT: Open the validation Excel and fill in the '_Project' sheet "
        "(Bauvorhaben, LOS, Bauwerk) plus per-sample Tiefe / Formation / Analysennr "
        "before running Step 2. The 'UMWELT_Aggr' sheet is a read-only verification "
        "echo; do not edit it. The lab's own DIN 4030 verdict is captured in the "
        "row with Aggr_Parameter='Lab_DIN4030_assessment' for cross-check."
    )
    return True


def _write_umwelt_aggr_echo(
    wb: "Workbook",
    sample_dfs: Dict[str, pd.DataFrame],
    sheet_meta: Dict[str, Tuple[str, str, str]],
    parameter_catalog: List[Tuple[str, str, str]],
    lab_din4030_key: str,
) -> None:
    """Insert a transposed UMWELT_Aggr echo sheet (parameters x samples).

    Mirrors the EBV ``UMWELT`` sheet convention but uses the Aggressivität
    parameter catalogue. Row 1 = material category, row 2 = Analysennr,
    row 4 = Probenbezeichnung, rows 5+ = one canonical parameter per row.
    A final row labelled ``"Lab DIN 4030 Verdict"`` carries the lab's own
    DIN 4030 verdict per sample for cross-check.

    Args:
        wb: target validation workbook.
        sample_dfs: mapping ``sheet_name -> parsed DataFrame``.
        sheet_meta: mapping ``sheet_name -> (probe, analysennr, material)``.
        parameter_catalog: WATER_AGGR_PARAMETERS list of (id, display, unit).
        lab_din4030_key: canonical ID of the lab-verdict capture row.
    """
    sheet_title = "UMWELT_Aggr"
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    insert_idx = 1 if "_Project" in wb.sheetnames else 0
    ws = wb.create_sheet(sheet_title, insert_idx)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E0E0E0")

    ws.cell(row=1, column=1, value="Material").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=2, column=1, value="Analysennr").font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=3, column=1, value="Eingang").font = header_font
    ws.cell(row=3, column=1).fill = header_fill
    ws.cell(row=4, column=1, value="Probenbezeichnung").font = header_font
    ws.cell(row=4, column=1).fill = header_fill

    # Parameter labels in column A starting row 5
    for i, (canonical, display, unit) in enumerate(parameter_catalog, start=5):
        ws.cell(row=i, column=1, value=display).font = header_font
        ws.cell(row=i, column=2, value=unit).font = header_font

    verdict_row = 5 + len(parameter_catalog)
    ws.cell(row=verdict_row, column=1, value="Lab DIN 4030 Verdict").font = header_font

    for col_offset, (sheet_name, df) in enumerate(sample_dfs.items(), start=3):
        probe, analysennr, material = sheet_meta.get(sheet_name, (sheet_name, "", ""))
        ws.cell(row=1, column=col_offset, value=material or "")
        ws.cell(row=2, column=col_offset, value=analysennr or "")
        ws.cell(row=4, column=col_offset, value=probe or sheet_name).font = header_font

        per_param: Dict[str, str] = {}
        verdict: str = ""
        for _, row in df.iterrows():
            cid = str(row.get("Aggr_Parameter", "")).strip()
            if cid == lab_din4030_key:
                verdict = str(row.get("Lab_Verdict_Text", "") or "")
                continue
            if not cid:
                continue
            op = str(row.get("Lab_Operator", "")).strip()
            val = row.get("Lab_Value", None)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                display = op if op else ""
            else:
                display = f"{op}{val}".replace(" ", "") if op else str(val)
            per_param[cid] = display

        for i, (canonical, _disp, _unit) in enumerate(parameter_catalog, start=5):
            v = per_param.get(canonical)
            if v is not None:
                ws.cell(row=i, column=col_offset, value=v)
        ws.cell(row=verdict_row, column=col_offset, value=verdict)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    for col_offset in range(3, 3 + len(sample_dfs)):
        ws.column_dimensions[ws.cell(row=1, column=col_offset).column_letter].width = 22


def _process_pak(
    input_dir: str = INPUT_DIR_PAK,
    output_dir: str = OUTPUT_DIR_PAK,
) -> bool:
    """Run the PAK / RuA-StB branch of Step 1.

    Reuses the EBV ingest pipeline (same lab format, same parser) and writes
    to the PAK validation folder. RuA-StB classification will be applied by
    Step 2 once the threshold table is in place.

    Args:
        input_dir: directory containing PAK lab-report PDFs (Strassenaufbruch).
        output_dir: parent directory for the timestamped validation folder.

    Returns:
        True if at least one PDF produced usable data; False if the folder
        was missing or no PDFs were present.
    """
    print("STEP 1 [PAK]: Starting data extraction from PAK lab reports...")
    if not os.path.exists(input_dir):
        print(
            f"Directory '{input_dir}' not found. Creating it now. "
            "Please add your PAK / Strassenaufbruch PDFs and restart."
        )
        os.makedirs(input_dir, exist_ok=True)
        return False
    pdf_files = sorted(glob.glob(os.path.join(input_dir, "*.pdf")))
    if not pdf_files:
        print(f"No PDF files found in '{input_dir}'. Skipping branch.")
        return False
    return _process_ebv(input_dir=input_dir, output_dir=output_dir, flow_name="PAK")


def main() -> None:
    """Entry point for Step 1 — ingests all selected flows and writes ONE
    consolidated Validation.xlsx into the shared session folder.

    The CLI argument ``--flow`` selects which branch(es) to ingest:

      * ``ebv``           — only EBV PDFs
      * ``aggressivität`` — only Aggressivität PDFs
      * ``pak``           — only PAK / Strassenaufbruch PDFs
      * ``all``           — every branch that has at least one PDF
                            (empty input folders silently skipped)

    All flows write to the SAME timestamped folder under ``1_validation/``,
    so that a single project's EBV/Aggressivität/PAK data lives together
    rather than in three sibling subfolders.
    """
    parser = argparse.ArgumentParser(
        description=(
            "EBV Tool - Step 1 (extraction). "
            "Three-flow ingest into one Validation.xlsx."
        )
    )
    parser.add_argument(
        "--flow",
        choices=("ebv", "aggressivität", "pak", "all"),
        default="all",
        help="Which flows to ingest. Default: all (empty input folders skipped).",
    )
    args = parser.parse_args()
    flow = args.flow

    pak_samples: List[Tuple[str, "pd.DataFrame"]] = []
    ebv_samples: List[Tuple[str, "pd.DataFrame"]] = []
    aggr_samples: List[Tuple[str, "pd.DataFrame"]] = []
    projektnummer: str = ""

    # Read background metadata once so all branches see the same fields.
    global _BG_META
    _BG_META = _read_background_metadata()
    if _BG_META["data"]:
        print("STEP 1: Loaded background_data.txt:")
        for k, v in _BG_META["data"].items():
            print(f"  {k}: {v}")

    if flow in ("pak", "all"):
        print("STEP 1 [PAK]: Ingesting PAK PDFs...")
        pak_samples, p_pak = _ingest_pak_dir()
        if pak_samples:
            print(f"  PAK: {len(pak_samples)} sample(s) parsed.")
            projektnummer = projektnummer or p_pak
        else:
            print("  PAK: no PDFs found.")

    if flow in ("ebv", "all"):
        print("STEP 1 [EBV]: Ingesting EBV PDFs...")
        ebv_samples, p_ebv = _ingest_ebv_dir()
        if ebv_samples:
            print(f"  EBV: {len(ebv_samples)} sample(s) parsed.")
            projektnummer = projektnummer or p_ebv
        else:
            print("  EBV: no PDFs found.")

    if flow in ("aggressivität", "all"):
        print("STEP 1 [Aggressivität]: Ingesting Aggressivität PDFs...")
        aggr_samples, p_aggr = _ingest_aggr_dir()
        if aggr_samples:
            print(f"  Aggressivität: {len(aggr_samples)} sample(s) parsed.")
            projektnummer = projektnummer or p_aggr
        else:
            print("  Aggressivität: no PDFs found.")

    total = len(pak_samples) + len(ebv_samples) + len(aggr_samples)
    if total == 0:
        print("\nNo samples across any flow — nothing to write.")
        return

    session_dir = _get_session_dir_validation()
    xlsx_path, html_path = _write_unified_validation(
        session_dir=session_dir,
        pak_samples=pak_samples,
        ebv_samples=ebv_samples,
        aggr_samples=aggr_samples,
        projektnummer=projektnummer,
    )
    print(f"\n[STEP 1] {total} sample(s) written to consolidated Validation.")
    print(f"-> {xlsx_path}")
    print(f"-> {html_path}")
    print(
        "\nIMPORTANT: Open Validation.xlsx, fill in the '_Project' sheet "
        "(Bauvorhaben / LOS / Bauwerk) plus per-sample metadata before running Step 2."
    )


if __name__ == "__main__":
    main()
