"""
EBV Tool — RuVA-StB 01 PAK reporter (Berlin Fassung 2018).

Generates a self-contained Excel + PDF report for the PAK / Straßenaufbruch
flow. There is no company "Mantelverordnung-style" skeleton for RuVA-StB 01
in the user's source materials, so this module builds the workbook from
scratch using the same colour-coding conventions as the EBV reporter:

    A:                green (C6EFCE)
    B:                yellow (FFEB9C)
    C:                red    (FFC7CE)
    Gefährlicher Abfall:   dark red on white text (9C0006)

Layout (single sheet ``RuVA``):

    Row 1-2    Project header (Projektnummer, Bauvorhaben)
    Row 4      Standards reference
    Row 6-11   Tabelle 1 reference grid (Klasse | PAK16 | Phenol | Verwertung)
    Row 13     Hazardous-waste trigger summary
    Row 16     Per-sample header
    Row 17+    One row per sample

Strict type hints + comprehensive docstrings per project convention.
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config_pak import (
    PAK_CLASS_FILL_HEX,
    RUVA_HAZARDOUS_TRIGGERS,
)
from evaluator_pak import PakResult


@dataclass(frozen=True)
class PakProjectMeta:
    """Project-level header data for the PAK output workbook.

    Attributes:
        projektnummer: e.g. ``"e-327524"``.
        bauvorhaben: free-form construction-project name.
        los: optional LOS identifier.
        bauwerk: optional Bauwerk identifier.
    """

    projektnummer: str = ""
    bauvorhaben: str = ""
    los: str = ""
    bauwerk: str = ""


@dataclass(frozen=True)
class PakSampleMeta:
    """Per-sample metadata for the PAK output workbook.

    Attributes:
        probenbezeichnung: sample identifier (e.g. ``"MP01"``).
        petrographische_beschreibung: optional petrographic description.
        stratigraphie: optional stratigraphic context.
        labor_nummer: optional lab analysis ID.
        tiefe: optional sample depth.
    """

    probenbezeichnung: str = ""
    petrographische_beschreibung: str = ""
    stratigraphie: str = ""
    labor_nummer: str = ""
    tiefe: str = ""


# ---------------------------------------------------------------------------
# Visual constants
# ---------------------------------------------------------------------------

_THIN: Side = Side(border_style="thin", color="808080")
_BORDER: Border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL: PatternFill = PatternFill("solid", fgColor="E0E0E0")
_HEADER_FONT: Font = Font(bold=True, size=10)
_TITLE_FONT: Font = Font(bold=True, size=14, color="000000")
_NOTE_FONT: Font = Font(italic=True, color="666666", size=9)

LIBREOFFICE_BIN: str = "soffice"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _set(ws: Worksheet, coord: str, value: Any, *, font: Optional[Font] = None,
         fill: Optional[PatternFill] = None, alignment: Optional[Alignment] = None,
         border: Optional[Border] = None) -> None:
    """Write a cell and apply optional styling in one call.

    Args:
        ws: target worksheet.
        coord: A1-style coordinate (e.g. ``"B17"``).
        value: scalar to write.
        font: optional Font override.
        fill: optional PatternFill override.
        alignment: optional Alignment override.
        border: optional Border override.
    """
    cell = ws[coord]
    cell.value = value
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def _fmt_value(v: Optional[float]) -> str:
    """Format a numeric value for display, with German comma decimals.

    Args:
        v: numeric value, or None.

    Returns:
        Display string (e.g. ``"7,2"``, ``"<0,01"``) or ``"–"`` for None.
    """
    if v is None:
        return "–"
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).replace(".", ",")


def _convert_xlsx_to_pdf(xlsx_path: str, output_dir: str) -> Optional[str]:
    """Convert ``.xlsx`` → ``.pdf`` via LibreOffice headless.

    Args:
        xlsx_path: input workbook path.
        output_dir: directory for the produced PDF.

    Returns:
        PDF path on success; None on LibreOffice unavailability or failure.
    """
    binary = shutil.which("soffice") or shutil.which("libreoffice")
    if binary is None:
        print(
            "  -> ERROR: LibreOffice (soffice) not found on PATH. "
            "Install LibreOffice to enable PAK PDF export."
        )
        return None
    try:
        completed = subprocess.run(
            [binary, "--headless", "--convert-to", "pdf", "--outdir", output_dir, xlsx_path],
            capture_output=True, text=True, timeout=120, check=False,
        )
    except (subprocess.TimeoutExpired, OSError) as e:
        print(f"  -> ERROR: LibreOffice conversion failed: {e}")
        return None
    if completed.returncode != 0:
        print(f"  -> ERROR: LibreOffice rc={completed.returncode}: {completed.stderr}")
        return None
    expected = os.path.join(
        output_dir, os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf"
    )
    return expected if os.path.exists(expected) else None


# ---------------------------------------------------------------------------
# Worksheet builders
# ---------------------------------------------------------------------------


def _write_header(ws: Worksheet, project: PakProjectMeta) -> None:
    """Write the project header block (rows 1-4).

    Args:
        ws: target worksheet.
        project: project metadata.
    """
    _set(ws, "A1", project.projektnummer or "—", font=_TITLE_FONT)
    _set(ws, "A2", project.bauvorhaben or "")
    _set(ws, "A3", project.los or "")
    _set(ws, "A4",
         "Bewertung nach RuVA-StB 01 (Ausgabe 2001, Fassung 2005) – "
         "Berlin Fassung 2018 (Amtsblatt Berlin Nr. 07/2018 S. 900)",
         font=_NOTE_FONT)


def _write_tabelle_1_reference(ws: Worksheet, start_row: int) -> int:
    """Write the RuVA-StB 01 Tabelle 1 reference grid.

    Args:
        ws: target worksheet.
        start_row: first row of the table header.

    Returns:
        Row index immediately AFTER the last reference row written.
    """
    r = start_row
    _set(ws, f"A{r}", "Tabelle 1 – Verwertungsklassen (RuVA-StB 01, Berlin Fassung 2018)",
         font=Font(bold=True, size=11))
    r += 1

    headers: List[Tuple[str, str]] = [
        ("A", "Klasse"),
        ("B", "PAK nach EPA, Feststoff [mg/kg TS]"),
        ("C", "Phenolindex, Eluat [mg/l]"),
        ("D", "Verwertungsverfahren nach Abschnitt"),
    ]
    for col, label in headers:
        _set(ws, f"{col}{r}", label, font=_HEADER_FONT, fill=_HEADER_FILL,
             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
             border=_BORDER)
    r += 1

    rows: List[Tuple[str, str, str, str, str]] = [
        ("A", "≤ 25",          "≤ 0,1",          "4.1 (oder ausnahmsweise 4.2 / 4.3)", "A"),
        ("B", "> 25 und ≤ 100", "≤ 0,1",          "kein (Entsorgung)",                   "B"),
        ("C", "> 25 und ≤ 100", "> 0,1 und ≤ 50", "kein (Entsorgung)",                   "C"),
    ]
    for klasse, pak, phen, verf, fill_key in rows:
        fill_hex = PAK_CLASS_FILL_HEX.get(fill_key, "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_hex)
        for col, val in zip(("A", "B", "C", "D"), (klasse, pak, phen, verf)):
            _set(ws, f"{col}{r}", val,
                 fill=fill if col == "A" else None,
                 font=Font(bold=True) if col == "A" else None,
                 alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                 border=_BORDER)
        r += 1

    return r


def _write_hazardous_block(ws: Worksheet, start_row: int) -> int:
    """Write the hazardous-waste trigger summary (Berlin amendment point 2).

    Args:
        ws: target worksheet.
        start_row: row to begin writing.

    Returns:
        Row index immediately after the block.
    """
    r = start_row + 1
    _set(ws, f"A{r}", "Gefährlicher Abfall (Abfallschlüssel 170301*) bei Überschreitung eines der folgenden Werte:",
         font=Font(bold=True, color="9C0006"))
    r += 1
    triggers_lines = [
        f"PAK nach EPA (Feststoff) > {int(RUVA_HAZARDOUS_TRIGGERS['PAK16'][0])} mg/kg TS",
        f"Benzo(a)pyren (Feststoff) > {int(RUVA_HAZARDOUS_TRIGGERS['Benzo(a)pyren'][0])} mg/kg TS",
        f"Phenolindex (Eluat) > {int(RUVA_HAZARDOUS_TRIGGERS['Phenolindex'][0])} mg/l",
    ]
    for line in triggers_lines:
        _set(ws, f"A{r}", "  • " + line, font=Font(color="9C0006"))
        r += 1
    return r + 1


def _write_samples_table(
    ws: Worksheet,
    start_row: int,
    samples: List[Tuple[PakSampleMeta, PakResult]],
) -> None:
    """Write the per-sample evaluation table.

    Args:
        ws: target worksheet.
        start_row: row of the table header.
        samples: list of ``(metadata, result)`` tuples.
    """
    r = start_row
    headers: List[Tuple[str, str]] = [
        ("A", "Probenbezeichnung"),
        ("B", "Petrographische\nBeschreibung"),
        ("C", "Stratigraphie"),
        ("D", "Labor-Nummer"),
        ("E", "PAK16\n[mg/kg TS]"),
        ("F", "Benzo(a)pyren\n[mg/kg TS]"),
        ("G", "Phenolindex\n[mg/l]"),
        ("H", "Verwertungs-\nklasse"),
        ("I", "Verwertungsverfahren\nnach Abschnitt"),
        ("J", "Maßgebliche\nParameter"),
        ("K", "Anmerkungen"),
    ]
    for col, label in headers:
        _set(ws, f"{col}{r}", label, font=_HEADER_FONT, fill=_HEADER_FILL,
             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
             border=_BORDER)
    ws.row_dimensions[r].height = 36
    r += 1

    for meta, result in samples:
        cls = result.klasse
        fill_hex = PAK_CLASS_FILL_HEX.get(cls, "FFFFFF")
        cls_fill = PatternFill("solid", fgColor=fill_hex)
        # Dark-red class needs white text
        cls_font = Font(bold=True, color="FFFFFF") if cls == "Gefährlicher Abfall" else Font(bold=True)

        drivers_text = ", ".join(result.driving_parameters) if result.driving_parameters else "—"
        notes_text = "; ".join(result.notes) if result.notes else ""
        if result.triggers:
            notes_text = (notes_text + "; " if notes_text else "") + f"Trigger: {', '.join(result.triggers)}"

        cells: List[Tuple[str, Any, Optional[Font], Optional[PatternFill]]] = [
            ("A", meta.probenbezeichnung,                     None, None),
            ("B", meta.petrographische_beschreibung,          None, None),
            ("C", meta.stratigraphie,                         None, None),
            ("D", meta.labor_nummer,                          None, None),
            ("E", _fmt_value(result.pak16_value),             None, None),
            ("F", _fmt_value(result.bap_value),               None, None),
            ("G", _fmt_value(result.phenol_value),            None, None),
            ("H", cls,                                        cls_font, cls_fill),
            ("I", result.verwertungsverfahren,                None, None),
            ("J", drivers_text,                               None, None),
            ("K", notes_text,                                 _NOTE_FONT if notes_text else None, None),
        ]
        for col, val, font, fill in cells:
            _set(ws, f"{col}{r}", val,
                 font=font,
                 fill=fill,
                 alignment=Alignment(horizontal="center" if col in {"E","F","G","H"} else "left",
                                     vertical="center", wrap_text=True),
                 border=_BORDER)
        r += 1


def _set_column_widths(ws: Worksheet) -> None:
    """Set sensible column widths for the PAK output."""
    widths: Dict[str, int] = {
        "A": 22, "B": 26, "C": 22, "D": 22,
        "E": 14, "F": 14, "G": 14, "H": 16,
        "I": 32, "J": 18, "K": 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def create_pak_report(
    samples: List[Tuple[PakSampleMeta, PakResult]],
    output_dir: str,
    project: PakProjectMeta,
) -> Tuple[str, Optional[str]]:
    """Generate the RuVA-StB 01 PAK report (.xlsx + .pdf) for one project.

    Args:
        samples: list of ``(metadata, result)`` tuples in display order.
        output_dir: target directory; created if needed.
        project: project header metadata.

    Returns:
        Tuple ``(xlsx_path, pdf_path)``. ``pdf_path`` is None on LibreOffice
        failure.
    """
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "RuVA"

    _write_header(ws, project)
    next_row = _write_tabelle_1_reference(ws, start_row=6)
    next_row = _write_hazardous_block(ws, start_row=next_row)
    _write_samples_table(ws, start_row=next_row + 1, samples=samples)

    _set_column_widths(ws)
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    safe_proj = re.sub(r"[^A-Za-z0-9_\-]", "_", (project.projektnummer or "").strip() or "Project")
    xlsx_path = os.path.join(output_dir, f"RuVA_PAK_{safe_proj}.xlsx")
    wb.save(xlsx_path)
    pdf_path = _convert_xlsx_to_pdf(xlsx_path, output_dir)
    return xlsx_path, pdf_path
