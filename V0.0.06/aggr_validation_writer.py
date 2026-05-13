"""Augment the Aggr_<sample> sheets in Validation.xlsx with the DIN 4030-1
and DIN 50929-3 calculations.

Rationale: Step 1's parser writes only the *parsed* parameter rows; the
user cannot verify the eventual XA-class / W0-W1-WD-WL classification by
hand without seeing the threshold buckets each parameter fell into, the
intermediate N/M digits, and the formula sums. This module appends a
"BERECHNUNGEN" block below the parsed rows of every ``Aggr_*`` sheet so
the validation file becomes self-checkable — matching the layout of the
company reference workbook ``2604XX_Rohdaten & Aggressivität.xlsx``
(sheets ``Beton_Wasser`` and ``Korrosion_Wasser``).

The block is read-only with respect to the evaluator: the same call into
``evaluator_aggressivität`` that ``step2_auswertung`` makes at report
time is replicated here. No mutation of any other sheet occurs.
"""
from __future__ import annotations

from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from config_aggressivität import (
    DIN4030_DIRECTION,
    DIN4030_THRESHOLDS,
    N1_WASSERART,
    N2_LAGE,
    N3_SALT_LOAD,
    N4_KS43,
    N5_CALCIUM,
    N6_PH,
    mgL_to_molM3,
)
from evaluator_aggressivität import (
    Din4030Result,
    Din50929WaterResult,
    evaluate_beton_wasser,
    evaluate_korrosion_wasser,
)

# ---------------------------------------------------------------------------
# Display helpers
# ---------------------------------------------------------------------------

_BORDER_THIN: Border = Border(
    left=Side(style="thin", color="FF808080"),
    right=Side(style="thin", color="FF808080"),
    top=Side(style="thin", color="FF808080"),
    bottom=Side(style="thin", color="FF808080"),
)
_HEADER_FONT: Font = Font(bold=True, size=10)
_BODY_FONT: Font = Font(size=10)
_BODY_BOLD: Font = Font(bold=True, size=10)
_CENTER: Alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT: Alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

#: XA-class fill palette (matches the reporter's DIN4030_FILL_HEX).
_XA_FILL: Dict[str, str] = {
    "XA0": "C6EFCE",
    "XA1": "FFEB9C",
    "XA2": "FFC7CE",
    "XA3": "F4B084",
    "Milieu unstimmig": "DA9694",
}

#: W bucket fill palette (matches reporter's W_BUCKET_FILL_HEX for class_W*).
_W_FILL: Dict[str, str] = {
    "sehr gering": "C6EFCE",
    "gering": "FFEB9C",
    "mittel": "FFC7CE",
    "hoch": "F4B084",
}


def _de(value: object) -> str:
    """Render ``value`` with German decimal comma; pass-through for non-floats."""
    if value is None:
        return ""
    if isinstance(value, float):
        # Trim trailing zeros for nice display.
        s = ("%g" % value)
        return s.replace(".", ",")
    if isinstance(value, int):
        return str(value)
    return str(value)


def _interval_text(
    table: List[Tuple[float, float, bool, bool, float, float]],
    n_or_m_value: Optional[float],
    pick_index: int,
) -> str:
    """Find the table row that produced ``n_or_m_value`` and render its range.

    Args:
        table: N/M interval table (lo, hi, lo_inc, hi_inc, N, M).
        n_or_m_value: the rating digit that came out of the lookup; we
            search for the first row whose N value (``pick_index=4``) or
            M value (``pick_index=5``) matches.
        pick_index: 4 to match by N value, 5 to match by M value.

    Returns:
        Compact human range string like ``"5 bis 25"`` or ``"> 6"``. Empty
        string when ``n_or_m_value`` is None or no row matches (engine
        skipped the lookup).
    """
    if n_or_m_value is None:
        return ""
    for row in table:
        lo, hi, lo_inc, hi_inc, n_val, m_val = row
        if (pick_index == 4 and n_val == n_or_m_value) or (
            pick_index == 5 and m_val == n_or_m_value
        ):
            return _fmt_interval(lo, hi, lo_inc, hi_inc)
    return ""


def _fmt_interval(lo: float, hi: float, lo_inc: bool, hi_inc: bool) -> str:
    """Render an interval ``(lo, hi, lo_inc, hi_inc)`` as ``"a bis b"``-style text."""
    if lo == float("-inf"):
        op = "≤" if hi_inc else "<"
        return f"{op} {_de(hi)}"
    if hi == float("inf"):
        op = "≥" if lo_inc else ">"
        return f"{op} {_de(lo)}"
    lo_op = "≥" if lo_inc else ">"
    hi_op = "≤" if hi_inc else "<"
    return f"{lo_op} {_de(lo)} und {hi_op} {_de(hi)}"


# ---------------------------------------------------------------------------
# Measurement extraction
# ---------------------------------------------------------------------------

#: Column indices in step1's Aggr sheet layout (1-based).
_COL_AGGR_PARAM: int = 10
_COL_LAB_UNIT: int = 11
_COL_LAB_OP: int = 12
_COL_LAB_VAL: int = 13


def _extract_measurements(ws: Worksheet) -> Tuple[Dict[str, Optional[float]], Dict[str, str]]:
    """Build the ``canonical_id -> value`` dict the evaluator expects.

    Args:
        ws: an ``Aggr_<sample>`` worksheet produced by step1.

    Returns:
        Tuple ``(measurements, display)``:
          * ``measurements``: ``canonical_id -> numeric_value_or_None``.
            ``<`` operator values are treated as half of the limit (same
            convention as the evaluator's "<BG" handling).
          * ``display``: ``canonical_id -> original_display_string`` used
            for the "Wert" column in the calculation rows (preserves
            ``"<1"`` etc. for the user).
    """
    meas: Dict[str, Optional[float]] = {}
    disp: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        param = ws.cell(r, _COL_AGGR_PARAM).value
        if not param:
            continue
        cid = str(param).strip()
        raw_val = ws.cell(r, _COL_LAB_VAL).value
        op = ws.cell(r, _COL_LAB_OP).value
        op_str = "" if op is None else str(op).strip()
        if raw_val is None:
            meas[cid] = None
            disp[cid] = ""
            continue
        try:
            v = float(raw_val)
        except (TypeError, ValueError):
            meas[cid] = None
            disp[cid] = str(raw_val)
            continue
        # "<" operator → engine treats as zero contribution; preserve display.
        if "<" in op_str:
            meas[cid] = 0.0
            disp[cid] = f"<{_de(v)}"
        else:
            meas[cid] = v
            disp[cid] = _de(v)
    return meas, disp


# ---------------------------------------------------------------------------
# Block writers
# ---------------------------------------------------------------------------


def _write_cell(
    ws: Worksheet,
    coord: str,
    value: object,
    *,
    font: Font = _BODY_FONT,
    alignment: Alignment = _LEFT,
    border: Optional[Border] = _BORDER_THIN,
    fill: Optional[PatternFill] = None,
) -> None:
    """Write a styled value into a single cell of the worksheet."""
    cell = ws[coord]
    cell.value = value
    cell.font = font
    cell.alignment = alignment
    if border is not None:
        cell.border = border
    if fill is not None:
        cell.fill = fill


def _write_din4030_block(
    ws: Worksheet,
    start_row: int,
    meas: Dict[str, Optional[float]],
    disp: Dict[str, str],
    result: Din4030Result,
) -> int:
    """Write the DIN 4030-1 per-parameter table + Gesamteinstufung.

    Returns the next free row after the block.
    """
    r = start_row
    _write_cell(ws, f"A{r}", "BERECHNUNG DIN 4030-1 — Beton-Wasser",
                font=Font(bold=True, size=11), alignment=_LEFT, border=None)
    r += 1
    headers = ("Parameter", "Wert", "XA1 schwach", "XA2 mäßig",
               "XA3 stark", "Milieu unstimmig", "Eingestuft")
    for i, h in enumerate(headers, start=1):
        col = chr(ord("A") + i - 1)
        _write_cell(ws, f"{col}{r}", h, font=_HEADER_FONT, alignment=_CENTER)
    r += 1

    # Parameter display labels (canonical_id -> "label [unit]")
    label_map: Dict[str, str] = {
        "pH": "pH-Wert [-]",
        "Mg": "Mg²⁺ [mg/l]",
        "NH4": "NH₄⁺ [mg/l]",
        "SO4": "SO₄²⁻ [mg/l]",
        "CO2_angr": "CO₂ angreifend [mg/l]",
    }
    for cid in DIN4030_DIRECTION:
        thresholds = DIN4030_THRESHOLDS.get(cid, {})
        cells = [
            label_map.get(cid, cid),
            disp.get(cid, ""),
            _fmt_interval(*thresholds["XA1"]) if "XA1" in thresholds else "—",
            _fmt_interval(*thresholds["XA2"]) if "XA2" in thresholds else "—",
            _fmt_interval(*thresholds["XA3"]) if "XA3" in thresholds else "—",
            _fmt_interval(*thresholds["Milieu unstimmig"]) if "Milieu unstimmig" in thresholds else "—",
            result.per_parameter_class.get(cid, "—"),
        ]
        for i, val in enumerate(cells, start=1):
            col = chr(ord("A") + i - 1)
            align = _CENTER if i > 1 else _LEFT
            fill = None
            if i == 7 and val in _XA_FILL:
                fill = PatternFill("solid", fgColor=_XA_FILL[val])
            _write_cell(ws, f"{col}{r}", val, font=_BODY_FONT, alignment=align, fill=fill)
        r += 1

    # Optional Sulfid note
    if "S2" in meas and meas["S2"] is not None:
        _write_cell(ws, f"A{r}", "S²⁻ [mg/l]", alignment=_LEFT)
        _write_cell(ws, f"B{r}", disp.get("S2", ""), alignment=_CENTER)
        _write_cell(ws, f"C{r}", "keine Grenzwerte in DIN 4030-1 (Information)",
                    font=_BODY_FONT, alignment=_LEFT)
        for col in ("D", "E", "F", "G"):
            _write_cell(ws, f"{col}{r}", "", alignment=_CENTER)
        r += 1

    r += 1  # spacer
    overall_text = f"Gesamteinstufung DIN 4030-1: {result.overall_class}"
    _write_cell(ws, f"A{r}", overall_text, font=_BODY_BOLD, alignment=_LEFT, border=None,
                fill=PatternFill("solid", fgColor=_XA_FILL.get(result.overall_class, "FFFFFF")))
    r += 2
    return r


def _write_din50929_block(
    ws: Worksheet,
    start_row: int,
    meas: Dict[str, Optional[float]],
    result: Din50929WaterResult,
    wasserart: str = "stehend",
) -> int:
    """Write the DIN 50929-3 Nx/Mx breakdown + W formulas.

    Returns the next free row after the block.
    """
    r = start_row
    _write_cell(ws, f"A{r}", "BERECHNUNG DIN 50929-3 — Korrosion-Wasser",
                font=Font(bold=True, size=11), alignment=_LEFT, border=None)
    r += 1
    headers = ("Nr.", "Merkmal", "Eingangswert", "Bereich (Tabelle)",
               "N (unleg. Stahl)", "M (verz. Stahl)")
    for i, h in enumerate(headers, start=1):
        col = chr(ord("A") + i - 1)
        _write_cell(ws, f"{col}{r}", h, font=_HEADER_FONT, alignment=_CENTER)
    r += 1

    n = result.n_values
    m = result.m_values

    # Row 1 — Wasserart
    n1, m1 = N1_WASSERART.get(wasserart, (0.0, 0.0))
    rows: List[Tuple[object, ...]] = [
        ("1", "Wasserart", wasserart, "kategorial", n1, m1),
        ("2", "Lage des Objektes", "nicht angegeben", "—",
         n.get("N2", 0.0), m.get("M2", 0.0)),
    ]

    # Row 3 — c(Cl-)+2·c(SO4²-) in mol/m³
    cl = meas.get("Cl") or 0.0
    so4 = meas.get("SO4") or 0.0
    salt_load = (mgL_to_molM3(cl, "Cl") if cl else 0.0) + 2.0 * (
        mgL_to_molM3(so4, "SO4") if so4 else 0.0
    )
    rows.append((
        "3", "c(Cl⁻) + 2·c(SO₄²⁻)",
        f"{_de(salt_load)} mol/m³ (Cl={_de(cl)}, SO₄={_de(so4)} mg/l)",
        _interval_text(N3_SALT_LOAD, n.get("N3"), 4),
        n.get("N3"), m.get("M3"),
    ))

    # Row 4 — KS 4,3 (mmol/l → mol/m³ is identity)
    ks43 = meas.get("KS43") or 0.0
    rows.append((
        "4", "Säurekapazität KS₄,₃",
        f"{_de(ks43)} mmol/l (= {_de(ks43)} mol/m³)",
        _interval_text(N4_KS43, n.get("N4"), 4),
        n.get("N4"), m.get("M4"),
    ))

    # Row 5 — Ca²⁺ (mg/l → mol/m³)
    ca = meas.get("Ca") or 0.0
    ca_mol = mgL_to_molM3(ca, "Ca") if ca else 0.0
    rows.append((
        "5", "Calcium c(Ca²⁺)",
        f"{_de(ca_mol)} mol/m³ ({_de(ca)} mg/l)",
        _interval_text(N5_CALCIUM, n.get("N5"), 4),
        n.get("N5"), m.get("M5"),
    ))

    # Row 6 — pH
    ph = meas.get("pH") or 0.0
    rows.append((
        "6", "pH-Wert", _de(ph),
        _interval_text(N6_PH, n.get("N6"), 4),
        n.get("N6"), m.get("M6"),
    ))

    # Row 7 — Object/water potential (only if measured)
    if "N7" in n:
        rows.append(("7", "Objekt/Wasser-Potential Uh", "—",
                     "—", n.get("N7"), "—"))

    for vals in rows:
        for i, v in enumerate(vals, start=1):
            col = chr(ord("A") + i - 1)
            align = _CENTER if i in (1, 5, 6) else _LEFT
            _write_cell(ws, f"{col}{r}", v, font=_BODY_FONT, alignment=align)
        r += 1

    r += 1  # spacer

    # Formula lines + bucket classes
    def _wfill(label: str) -> Optional[PatternFill]:
        return PatternFill("solid", fgColor=_W_FILL[label]) if label in _W_FILL else None

    formulas: List[Tuple[str, str, str]] = [
        ("W0 = N1 + N3 + N4 + N5 + N6 + N3/N4",
         f"= {result.W0}", f"Mulden/Loch: {result.class_W0}; Fläche: {result.flaechen_class_W0}; "
         f"Abtrag {_de(result.rate_W0_mm_per_a)} mm/a, max Eindring {_de(result.rate_W0_max_mm_per_a)} mm/a"),
        ("W1 = W0 − N1 + N2·N3",
         f"= {result.W1}", f"Mulden/Loch: {result.class_W1}; Fläche: {result.flaechen_class_W1}"),
        ("WD = M1 + M3 + M4 + M5 + M6",
         f"= {result.WD}", f"Deckschicht-Güte: {result.class_WD}"),
        ("WL = WD + M2",
         f"= {result.WL}", f"Deckschicht-Güte: {result.class_WL}"),
    ]
    for formula, total, classes in formulas:
        _write_cell(ws, f"A{r}", formula, font=_BODY_FONT, alignment=_LEFT)
        _write_cell(ws, f"E{r}", total, font=_BODY_BOLD, alignment=_CENTER)
        _write_cell(ws, f"F{r}", classes, font=_BODY_FONT, alignment=_LEFT)
        r += 1

    return r + 1


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def augment_validation(xlsx_path: str) -> int:
    """Append DIN 4030 + DIN 50929 calculations to every ``Aggr_*`` sheet.

    Args:
        xlsx_path: path to ``Validation.xlsx`` (mutated in place).

    Returns:
        Number of Aggr sheets that received a calculations block.
    """
    wb = load_workbook(xlsx_path)
    augmented = 0
    for sn in [s for s in wb.sheetnames if s.startswith("Aggr_")]:
        ws = wb[sn]
        meas, disp = _extract_measurements(ws)
        try:
            d4030 = evaluate_beton_wasser(meas)
            d50929 = evaluate_korrosion_wasser(meas, wasserart="stehend")
        except Exception as exc:
            # Evaluator failure shouldn't break step1; log to the sheet itself.
            r = ws.max_row + 2
            _write_cell(ws, f"A{r}", f"Berechnung übersprungen — Fehler im Evaluator: {exc}",
                        font=_BODY_FONT, alignment=_LEFT, border=None)
            continue

        start = ws.max_row + 2
        next_row = _write_din4030_block(ws, start, meas, disp, d4030)
        _write_din50929_block(ws, next_row, meas, d50929)
        augmented += 1

    wb.save(xlsx_path)
    return augmented
