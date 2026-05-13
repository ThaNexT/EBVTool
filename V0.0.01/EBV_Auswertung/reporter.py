import os
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from config import EBV_VERSION

FUSSNOTEN_TEXTE = [
    "1: Die Materialwerte gelten für Bodenmaterial und Baggergut mit bis zu 10 Volumenprozent (BM und BG) oder bis zu 50 Volumenprozent (BM-F und BG-F) mineralischer Fremdbestandteile im Sinne von § 2 Nummer 8 der Bundes-Bodenschutz- und Altlastenverordnung mit nur vernachlässigbaren Anteilen an Störstoffen im Sinne von § 2 Nummer 9 der Bundes-Bodenschutz- und Altlastenverordnung. Bodenmaterial der Klasse BM-0 und Baggergut der Klasse BG-0 erfüllen die wertebezogenen Anforderungen an das Auf- oder Einbringen gemäß § 7 Absatz 3 der Bundes-Bodenschutz- und Altlastenverordnung. Bodenmaterial der Klasse BM-0 und Baggergut der Klasse BG-0 Sand erfüllen die wertebezogenen Anforderungen an das Auf- oder Einbringen gemäß § 8 Absatz 2 der Bundes-Bodenschutz- und Altlastenverordnung; Bodenmaterial der Klasse BM-0* und Baggergut der Klasse BG-0* erfüllen die wertebezogenen Anforderungen an das Auf- oder Einbringen gemäß § 8 Absatz 3 Nummer 1 der Bundes-Bodenschutz- und Altlastenverordnung.",
    "2: Bodenarten-Hauptgruppen gemäß Bodenkundlicher Kartieranleitung, 5. Auflage, Hannover 2005 (KA5); stark schluffige Sande, lehmig-schluffige Sande und stark lehmige Sande sowie Materialien, die nicht bodenartspezifisch zugeordnet werden können, sind entsprechend der Bodenart Lehm, Schluff zu bewerten.",
    "3: Die Eluatwerte in Spalte 6 sind mit Ausnahme des Eluatwertes für Sulfat nur maßgeblich, wenn für den betreffenden Stoff der jeweilige Feststoffwert nach Spalte 3 bis 5 überschritten wird. Der Eluatwert für PAK15 und Napthalin und Methylnaphtaline, gesamt, ist maßgeblich, wenn der Feststoffwert für PAK16 nach Spalte 3 bis 5 überschritten wird. Die in Klammern genannten Werte gelten jeweils bei einem TOC-Gehalt von ≥ 0,5 %.",
    "4: Stoffspezifischer Orientierungswert; bei Abweichungen ist die Ursache zu prüfen.",
    "5: Bei Überschreitung des Wertes ist die Ursache zu prüfen. Handelt es sich um naturbedingt erhöhte Sulfatkonzentrationen, ist eine Verwertung innerhalb der betroffenen Gebiete möglich. Außerhalb dieser Gebiete ist über die Verwertungseignung im Einzelfall und in Abstimmung mit der zuständigen Behörde zu entscheiden.",
    "6: Der Wert 1 mg/kg gilt für Sand und Lehm/Schluff. Für Ton gilt der Wert 1,5 mg/kg.",
    "7: Bodenmaterialspezifischer Orientierungswert. Bei heterogenen Bodenverhältnissen mineralischer Böden kann der TOC-Gehalt der Masse des anfallenden Materials als maßgeblich bei Verwertung im Umfeld des anfallenden Materials und Verwendung unter gleichen Bedingungen herangezogen werden. Beim Einbau sind Volumenbeständigkeit und Setzungsprozesse sowie die Vorgaben von § 6 Absatz 11 Satz 2 und 3 der Bundes-Bodenschutz- und Altlastenverordnung zu berücksichtigen. Beim Einbau sind Volumenbeständigkeit und Setzungsprozesse zu berücksichtigen.",
    "8: Die angegebenen Werte gelten für Kohlenwasserstoffverbindungen mit einer Kettenlänge von C10 bis C22. Der Gesamtgehalt bestimmt nach der DIN EN 14039, „Charakterisierung von Abfällen – Bestimmung des Gehalts an Kohlenwasserstoffen von C10 bis C40 mittels Gaschromatographie“, Ausgabe Januar 2005 darf insgesamt den in Klammern genannten Wert nicht überschreiten.",
    "9: PAK15: PAK16 ohne Naphthalin und Methylnaphthalin.",
    "10: PAK16: stellvertretend für die Gruppe der polyzyklischen aromatischen Kohlenwasserstoffe (PAK) werden nach der Liste der US-amerikanischen Umweltbehörde, Environmental Protection Agency (EPA), 16 ausgewählte PAK untersucht: Acenaphthen, Acenaphthylen, Anthracen, Benzo[a]anthracen, Benzo[a]pyren, Benzo[b]fluoranthen, Benzo[g,h,i]perylen, Benzo[k]fluoranthen, Chrysen, Dibenzo[a,h]anthracen, Fluoranthen, Fluoren, Indeno[1,2,3- cd]pyren, Naphthalin, Phenanthren und Pyren.",
    "11: Bei Überschreitung der Werte sind die Materialien auf fallspezifische Belastungen zu untersuchen.",
    "12: Bei Quecksilber und Thallium ist für die Klassifizierung (BM-F0* bis BM-F3) der Gesamtgehalt maßgeblich. Der Eluatwert für BM-0* ist einzuhalten."
]

def create_raw_data_log(raw_df, output_dir, original_filename):
    """Erstellt eine einfache Textdatei zur Prüfung der extrahierten Rohdaten."""
    if not os.path.exists(output_dir): os.makedirs(output_dir)
    base_name = os.path.splitext(original_filename)[0]
    txt_path = os.path.join(output_dir, f"Rohdaten_Check_{base_name}.txt")
    
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(f"--- ROHDATEN CHECK FÜR: {original_filename} ---\n")
        f.write(f"Gelesene Parameter-Zeilen gesamt: {len(raw_df)}\n")
        f.write("-" * 60 + "\n")
        
        if raw_df.empty:
            f.write("WARNUNG: Keine Daten aus dem PDF extrahiert.\n")
            return
        
        for idx, row in raw_df.iterrows():
            labor_str = str(row.get('Labor_Original_String', '')).ljust(35)
            ebv_name = str(row.get('EBV_Parameter', '')).ljust(35)
            val = str(row.get('Wert', ''))
            op = str(row.get('Operator', ''))
            einheit = str(row.get('Einheit', ''))
            f.write(f"Labor: {labor_str} | EBV: {ebv_name} | Ausgelesen: {op} {val} {einheit}\n")
            
    import logging
    logging.info(f"Rohdaten-Checkdatei gespeichert: {txt_path}")

def create_excel_report(df, output_dir, original_filename, bodenart):
    if not os.path.exists(output_dir): os.makedirs(output_dir)
        
    base_name = os.path.splitext(original_filename)[0]
    output_filename = f"Klassifizierung_{base_name}.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='EBV_Klassifizierung', startrow=3)
        
    wb = load_workbook(output_path)
    ws = wb['EBV_Klassifizierung']
    
    ws.cell(row=1, column=1).value = f"HINWEIS: Referenz-Bodenart für BM-0 wurde auf '{bodenart}' gesetzt. Auswertungsdatum: {datetime.date.today().isoformat()}"
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws.cell(row=2, column=1).value = (
        "RECHTLICHER HINWEIS: Diese Auswertung wurde maschinell erstellt und ersetzt NICHT die gutachterliche Prüfung. "
        f"Grundlage: {EBV_VERSION['gesetz']} ({EBV_VERSION['fundstelle']})."
    )
    ws.cell(row=2, column=1).font = Font(bold=True, color="9C0006")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    font_green = Font(color="006100")
    fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    font_yellow = Font(color="9C5700")
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    font_red = Font(color="9C0006")
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    font_gray = Font(color="7A7A7A")
    
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                try: max_length = max(max_length, len(str(cell_value)))
                except: pass
        ws.column_dimensions[column_letter].width = min(max_length + 3, 50)

    for row in range(5, ws.max_row + 1):
        klasse_cell = ws.cell(row=row, column=4) 
        val = str(klasse_cell.value)
        
        fill_to_use, font_to_use = None, None
        if "BM-0" in val: fill_to_use, font_to_use = fill_green, font_green
        elif "BM-F" in val: fill_to_use, font_to_use = fill_yellow, font_yellow
        elif "> BM-F3" in val: fill_to_use, font_to_use = fill_red, font_red
        elif "Nicht in EBV" in val or "Kein" in val: fill_to_use, font_to_use = fill_gray, font_gray
            
        if fill_to_use:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill_to_use
                ws.cell(row=row, column=col).font = font_to_use
                
    start_row = ws.max_row + 3
    ws.cell(row=start_row, column=1).value = "Regelbezüge & Fußnoten (Anlage 1 Tabelle 3 EBV):"
    ws.cell(row=start_row, column=1).font = Font(bold=True)
    
    for i, text in enumerate(FUSSNOTEN_TEXTE):
        cell = ws.cell(row=start_row + 1 + i, column=1)
        cell.value = text
        ws.merge_cells(start_row=start_row + 1 + i, start_column=1, end_row=start_row + 1 + i, end_column=6)
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[start_row + 1 + i].height = 30

    wb.save(output_path)
    import logging
    logging.info(f"Excel-Bericht gespeichert: {output_path}")