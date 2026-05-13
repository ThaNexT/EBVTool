import os
import html
import datetime
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from config import EBV_VERSION

from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

FUSSNOTEN_TEXTE = [
    "1: Die Materialwerte gelten für Bodenmaterial und Baggergut mit bis zu 10 Volumenprozent (BM und BG) oder bis zu 50 Volumenprozent (BM-F und BG-F) mineralischer Fremdbestandteile im Sinne von § 2 Nummer 8 der Bundes-Bodenschutz- und Altlastenverordnung mit nur vernachlässigbaren Anteilen an Störstoffen im Sinne von § 2 Nummer 9 der Bundes-Bodenschutz- und Altlastenverordnung. Bodenmaterial der Klasse BM-0 und Baggergut der Klasse BG-0 erfüllen die wertebezogenen Anforderungen an das Auf- oder Einbringen gemäß § 7 Absatz 3 der Bundes-Bodenschutz- und Altlastenverordnung. Bodenmaterial der Klasse BM-0 und Baggergut der Klasse BG-0 Sand erfüllen die wertebezogenen Anforderungen an das Auf- oder Einbringen gemäß § 8 Absatz 2 der Bundes-Bodenschutz- und Altlastenverordnung; Bodenmaterial der Klasse BM-0* und Baggergut der Klasse BG-0* erfüllen die wertebezogenen Anforderungen an das Auf- oder Einbringen gemäß § 8 Absatz 3 Nummer 1 der Bundes-Bodenschutz- und Altlastenverordnung.",
    "2: Bodenarten-Hauptgruppen gemäß Bodenkundlicher Kartieranleitung, 5. Auflage, Hannover 2005 (KA5); stark schluffige Sande, lehmig-schluffige Sande und stark lehmige Sande sowie Materialien, die nicht bodenartspezifisch zugeordnet werden können, sind entsprechend der Bodenart Lehm, Schluff zu bewerten.",
    "3: Die Eluatwerte in Spalte 6 sind mit Ausnahme des Eluatwertes für Sulfat nur maßgeblich, wenn für den betreffenden Stoff der jeweilige Feststoffwert nach Spalte 3 bis 5 überschritten wird. Der Eluatwert für PAK15 und Napthalin und Methylnaphtaline, gesamt, ist maßgeblich, wenn der Feststoffwert für PAK16 nach Spalte 3 bis 5 überschritten wird. Die in Klammern genannten Werte gelten jeweils bei einem TOC-Gehalt von ≥ 0,5 %.",
    "4: Stoffspezifischer Orientierungswert; bei Abweichungen ist die Ursache zu prüfen.",
    "5: Bei Überschreitung des Wertes ist die Ursache zu prüfen. Handelt es sich um naturbedingt erhöhte Sulfatkonzentrationen, ist eine Verwertung innerhalb der betroffenen Gebiete möglich. Außerhalb dieser Gebiete ist über die Verwertungseignung im Einzelfall und in Abstimmung mit der zuständigen Behörde zu entscheiden.",
    "6: Der Wert 1 mg/kg gilt für Sand und Lehm/Schluff. Für Ton gilt der Wert 1,5 mg/kg.",
    "7: Bodenmaterialspezifischer Orientierungswert. Bei heterogenen Bodenverhältnissen mineralischer Böden kann der TOC-Gehalt der Masse des anfallenden Materials als maßgeblich bei Verwertung im Umfeld des anfallenden Materials und Verwendung unter gleichen Bedingungen herangezogen werden. Beim Einbau sind Volumenbeständigkeit und Setzungsprozesse sowie die Vorgaben von § 6 Absatz 11 Satz 2 und 3 der Bundes-Bodenschutz- und Altlastenverordnung zu berücksichtigen. Beim Einbau sind Volumenbeständigkeit und Setzungsprozesse zu berücksichtigen.",
    "8: Die angegebenen Werte gelten für Kohlenwasserstoffverbindungen mit einer Kettenlänge von C10 bis C22. Der Gesamtgehalt bestimmt nach der DIN EN 14039, „Charakterisierung von Abfällen – Bestimmung des Gehalts an Kohlenwasserstoffen von C10 bis C40 mittels Gaschromatographie“, Ausgabe Januar 2005 darf insgesamt den in Klammern genannten Wert nicht überschreiten.",
    "9: PAK15: PAK16 ohne Naphthalin und Methylnaphthalin.",
    "10: PAK16: stellvertretend für die Gruppe der polyzyklischen aromatischen Kohlenwasserstoffe (PAK) werden nach der Liste der US-amerikanischen Umweltbehörde, Environmental Protection Agency (EPA), 16 ausgewählte PAK untersucht: Acenaphthen, Acenaphthylen, Anthracen, Benzo[a]anthracen, Benzo[a]pyren, Benzo[b]fluoranthen, Benzo[g,h,i]perylen, Benzo[k]fluoranthen, Chrysen, Dibenzo[a,h]anthracen, Fluoranthen, Fluoren, Indeno[1,2,3- cd]pyren, Naphthalin, Phenanthren und Pyren.",
    "11: Bei Überschreitung der Werte sind die Materialien auf fallspezifische Belastungen zu untersuchen.",
    "12: Bei Quecksilber und Thallium ist für die Klassifizierung (BM-F0* bis BM-F3) der Gesamtgehalt maßgeblich. Der Eluatwert für BM-0* ist einzuhalten."
]

CLASS_RANKS = {"BM-0": 0, "BM-0 (Eluat n. maßgeblich)": 0, "BM-0*": 1, "BM-F0*": 2, "BM-F1": 3, "BM-F2": 4, "BM-F3": 5, "> BM-F3 (Deponie!)": 6, "Kein Messwert": -1, "Kein Messwert (< BG)": -1}

def clean_float_string(val):
    if pd.isna(val) or val is None:
        return ""
    val_str = str(val)
    
    def round_match(match):
        num = float(match.group(0))
        return f"{num:.4f}".rstrip('0').rstrip('.')

    cleaned = re.sub(r'\d+\.\d{5,}', round_match, val_str)
    return html.escape(cleaned)

def create_excel_report(df, output_dir, original_filename, bodenart):
    if not os.path.exists(output_dir): os.makedirs(output_dir)
        
    base_name = os.path.splitext(original_filename)[0]
    output_filename = f"Auswertung_{base_name}.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    # 1. EXCEL BERICHT
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='EBV_Klassifizierung', startrow=3)
        
    wb = load_workbook(output_path)
    ws = wb['EBV_Klassifizierung']
    
    ws.cell(row=1, column=1).value = f"HINWEIS: Referenz-Bodenart für BM-0: '{bodenart}'. Datum: {datetime.date.today().isoformat()}"
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws.cell(row=2, column=1).value = f"RECHTLICHER HINWEIS: Automatisierte Vorprüfung. Ersetzt keine gutachterliche Freigabe. Grundlage: {EBV_VERSION['gesetz']}."
    ws.cell(row=2, column=1).font = Font(bold=True, color="9C0006")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    font_green, font_yellow, font_red, font_gray = Font(color="006100"), Font(color="9C5700"), Font(color="9C0006"), Font(color="7A7A7A")
    
    max_rank = -1
    worst_class = "BM-0"

    for row in range(5, ws.max_row + 1):
        val = str(ws.cell(row=row, column=4).value)
        rank = CLASS_RANKS.get(val, -1)
        if rank > max_rank:
            max_rank = rank
            worst_class = val
            
        fill_to_use, font_to_use = None, None
        if "BM-0" in val: fill_to_use, font_to_use = fill_green, font_green
        elif "BM-F" in val: fill_to_use, font_to_use = fill_yellow, font_yellow
        elif "> BM-F3" in val: fill_to_use, font_to_use = fill_red, font_red
        elif "Nicht in EBV" in val or "Kein" in val: fill_to_use, font_to_use = fill_gray, font_gray
            
        if fill_to_use:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill_to_use
                ws.cell(row=row, column=col).font = font_to_use

    wb.save(output_path)

    # 2. HTML BERICHT
    html_path = os.path.join(output_dir, f"Auswertung_{base_name}.html")
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <meta charset="utf-8"/>
    <style>
        body {{ font-family: Helvetica, Arial, sans-serif; font-size: 12px; color: #000; line-height: 1.3; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 15px; margin-bottom: 15px; }}
        th, td {{ border: 1px solid #a0a0a0; padding: 6px; text-align: left; vertical-align: middle; }}
        th {{ background-color: #e0e0e0; font-weight: bold; border-bottom: 2px solid #555; }}
        .header-red {{ color: #9C0006; font-weight: bold; margin-top: 5px; margin-bottom: 10px; }}
        .header-bold {{ font-weight: bold; font-size: 14px; }}
        .row-green {{ background-color: #C6EFCE; color: #006100; }}
        .row-yellow {{ background-color: #FFEB9C; color: #9C5700; }}
        .row-red {{ background-color: #FFC7CE; color: #9C0006; }}
        .row-gray {{ background-color: #F2F2F2; color: #7A7A7A; }}
        .summary {{ font-size: 15px; font-weight: bold; margin-top: 15px; margin-bottom: 15px; }}
        .footnotes {{ font-size: 10px; line-height: 1.4; color: #333; }}
    </style>
    </head>
    <body>
        <div class="header-bold">HINWEIS: Referenz-Bodenart für BM-0: '{bodenart}'. Datum: {datetime.date.today().isoformat()}</div>
        <div class="header-red">RECHTLICHER HINWEIS: Automatisierte Vorprüfung. Ersetzt keine gutachterliche Freigabe. Grundlage: {EBV_VERSION['gesetz']}.</div>
        <table>
            <thead>
                <tr>
                    <th width="30%">Parameter</th>
                    <th width="12%">Einheit</th>
                    <th width="15%">Messwert</th>
                    <th width="20%">Eingestufte Klasse</th>
                    <th width="15%">Maßgeblicher GW</th>
                    <th width="8%">Fußnote</th>
                </tr>
            </thead>
            <tbody>
    """
    for idx, row in df.iterrows():
        val = str(row.get("Eingestufte Klasse", ""))
        cls = ""
        if "BM-0" in val: cls = "row-green"
        elif "BM-F" in val: cls = "row-yellow"
        elif "> BM-F3" in val: cls = "row-red"
        elif "Nicht in EBV" in val or "Kein" in val: cls = "row-gray"

        html_content += f"<tr class='{cls}'>"
        html_content += f"<td>{clean_float_string(row.get('Parameter',''))}</td>"
        html_content += f"<td>{clean_float_string(row.get('Einheit',''))}</td>"
        html_content += f"<td>{clean_float_string(row.get('Messwert',''))}</td>"
        html_content += f"<td>{clean_float_string(val)}</td>"
        html_content += f"<td>{clean_float_string(row.get('Maßgeblicher GW',''))}</td>"
        html_content += f"<td>{clean_float_string(row.get('Fußnote',''))}</td>"
        html_content += "</tr>"

    summary_color = "row-green" if max_rank <= 1 else ("row-yellow" if max_rank <= 5 else "row-red")
    html_content += f"""
            </tbody>
        </table>
        <div class="summary">
            GESAMTEINSTUFUNG DER PROBE (Worst-Case): <span class="{summary_color}">{worst_class}</span>
        </div>
        <div class="footnotes">
            <b>Regelbezüge & Fußnoten (Anlage 1 Tabelle 3 EBV):</b><br/>
            {" | ".join(FUSSNOTEN_TEXTE)}
        </div>
    </body>
    </html>
    """
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)


    # 3. PDF DIREKT ZEICHNEN (Optimiert für exakt 1 Seite)
    pdf_path = os.path.join(output_dir, f"Auswertung_{base_name}.pdf")
    try:
        # Ränder oben/unten auf 20 reduziert
        doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A3), rightMargin=30, leftMargin=30, topMargin=20, bottomMargin=20)
        elements = []
        
        styles = getSampleStyleSheet()
        style_normal = styles["Normal"]
        style_normal.fontSize = 9   # Schrift von 10 auf 9
        style_normal.leading = 11   # Zeilenhöhe von 12 auf 11
        
        style_bold = styles["Normal"].clone("Bold")
        style_bold.fontName = "Helvetica-Bold"
        style_bold.fontSize = 9     # Schrift von 10 auf 9
        
        style_title = styles["Heading3"].clone("Title")
        style_title.textColor = colors.HexColor("#9C0006")
        style_title.fontSize = 11
        style_title.leading = 13
        
        elements.append(Paragraph(f"<b>HINWEIS: Referenz-Bodenart für BM-0: '{bodenart}'. Datum: {datetime.date.today().isoformat()}</b>", style_bold))
        elements.append(Paragraph(f"RECHTLICHER HINWEIS: Automatisierte Vorprüfung. Ersetzt keine gutachterliche Freigabe. Grundlage: {EBV_VERSION['gesetz']}.", style_title))
        elements.append(Spacer(1, 5)) # Spacer verkleinert
        
        data = [[
            Paragraph("<b>Parameter</b>", style_bold),
            Paragraph("<b>Einheit</b>", style_bold),
            Paragraph("<b>Messwert</b>", style_bold),
            Paragraph("<b>Eingestufte Klasse</b>", style_bold),
            Paragraph("<b>Maßgeblicher GW</b>", style_bold),
            Paragraph("<b>Fußnote</b>", style_bold)
        ]]
        
        bg_colors = []
        c_green = colors.HexColor("#C6EFCE")
        c_yellow = colors.HexColor("#FFEB9C")
        c_red = colors.HexColor("#FFC7CE")
        c_gray = colors.HexColor("#F2F2F2")
        
        for idx, row in df.iterrows():
            val = str(row.get("Eingestufte Klasse", ""))
            row_color = colors.white
            if "BM-0" in val: row_color = c_green
            elif "BM-F" in val: row_color = c_yellow
            elif "> BM-F3" in val: row_color = c_red
            elif "Nicht in EBV" in val or "Kein" in val: row_color = c_gray
            bg_colors.append(row_color)

            data.append([
                Paragraph(clean_float_string(row.get('Parameter','')), style_normal),
                Paragraph(clean_float_string(row.get('Einheit','')), style_normal),
                Paragraph(clean_float_string(row.get('Messwert','')), style_normal),
                Paragraph(clean_float_string(val), style_normal),
                Paragraph(clean_float_string(row.get('Maßgeblicher GW','')), style_normal),
                Paragraph(clean_float_string(row.get('Fußnote','')), style_normal)
            ])
            
        t = Table(data, colWidths=[350, 120, 140, 250, 150, 100], repeatRows=1)
        
        t_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#e0e0e0")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1), # Padding minimiert auf 1
            ('TOPPADDING', (0, 0), (-1, -1), 1),    # Padding minimiert auf 1
        ]
        
        for i, color in enumerate(bg_colors):
            t_style.append(('BACKGROUND', (0, i+1), (-1, i+1), color))
            
        t.setStyle(TableStyle(t_style))
        elements.append(t)
        elements.append(Spacer(1, 5)) # Spacer verkleinert
        
        sum_color = c_green if max_rank <= 1 else (c_yellow if max_rank <= 5 else c_red)
        sum_data = [[Paragraph("<b>GESAMTEINSTUFUNG DER PROBE (Worst-Case):</b>", style_bold), Paragraph(f"<b>{worst_class}</b>", style_bold)]]
        sum_t = Table(sum_data, colWidths=[300, 150])
        sum_t.setStyle(TableStyle([('BACKGROUND', (1,0), (1,0), sum_color), ('BOX', (1,0), (1,0), 1, colors.grey), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        elements.append(sum_t)
        
        elements.append(Spacer(1, 5)) # Spacer verkleinert
        
        elements.append(Paragraph("<b>Regelbezüge & Fußnoten (Anlage 1 Tabelle 3 EBV):</b>", style_bold))
        style_fn = styles["Normal"].clone("Footnote")
        style_fn.fontSize = 7       # Fußnoten-Schrift auf 7 verkleinert
        style_fn.leading = 8.5      # Zeilenhöhe auf 8.5 reduziert
        elements.append(Paragraph(" | ".join(FUSSNOTEN_TEXTE), style_fn))
        
        doc.build(elements)
        print(f"  -> Bericht generiert: {output_path} (Excel), HTML & A3-PDF (Einzelseite)")
        
    except Exception as e:
        print(f"  -> FEHLER beim PDF-Export: {e}")