"""
EBV Tool — Aggressivität reporter (water-only, results-only build).

Produces a single-sheet A3-landscape report combining DIN 4030-1
(Beton-Wasser) and DIN 50929-3 (Korrosion-Wasser) results per sample.
The output is built from scratch — earlier builds cloned the full
company workbook and ran into multi-page truncation. This module keeps
only what the engineer needs to read: input values, derived rating
digits, W0/W1/WD/WL, class labels, and lab-verdict cross-check.

Layout (sheet ``Aggressivität``):

    Row 1-4    Project header
    Row 6      Section title "DIN 4030-1 — Beton-Wasser (Concrete attack)"
    Row 7-11   Threshold reference grid (XA1/XA2/XA3/Milieu per parameter)
    Row 13     Beton-Wasser sample-row header
    Row 14+    One row per sample (DIN 4030)
    spacer
    Section title "DIN 50929-3 — Korrosion-Wasser (Steel corrosion)"
    Korrosion-Wasser sample-row header
    One row per sample (DIN 50929)

A3 landscape with fit-to-page-width=1. Class cells are colour-coded.
Strict type hints + comprehensive docstrings per project convention.
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from evaluator_aggressivität import (
    Din4030Result,
    Din50929WaterResult,
    evaluate_beton_wasser,
    evaluate_korrosion_wasser,
)


# ---------------------------------------------------------------------------
# Public data contracts (unchanged API for step2_auswertung)
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class AggrProjectMeta:
    """Project-level header data for the Aggressivität output workbook.

    Attributes:
        projektnummer: e.g. ``"e-327524"``.
        bauvorhaben: free-form construction-project name.
        los: optional LOS identifier.
        bauwerk: optional Bauwerk identifier.
    """

    projektnummer: str = ""
    bauvorhaben: str = ""
    los: str = ""
    bauwerk: str = ""


@dataclass(frozen=True)
class AggrSampleMeta:
    """Per-sample metadata for the Aggressivität output workbook.

    Attributes:
        probenbezeichnung: customer's sample identifier.
        tiefe: optional depth string.
        formation: optional geological-formation string.
        wasserart: one of N1_WASSERART keys (default "stehend").
        objektlage: optional one of N2_LAGE keys.
        u_potential: optional object/water potential in V (Cu/CuSO4 reference).
        lab_din4030_verdict: text captured from the lab PDF for cross-check.
    """

    probenbezeichnung: str = ""
    tiefe: str = ""
    formation: str = ""
    wasserart: str = "stehend"
    objektlage: Optional[str] = None
    u_potential: Optional[float] = None
    lab_din4030_verdict: str = ""


# ---------------------------------------------------------------------------
# Visual constants
# ---------------------------------------------------------------------------

_THIN: Side = Side(border_style="thin", color="808080")
_BORDER: Border = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL: PatternFill = PatternFill("solid", fgColor="E0E0E0")
_SECTION_FILL: PatternFill = PatternFill("solid", fgColor="305496")
_HEADER_FONT: Font = Font(bold=True, size=10)
_SECTION_FONT: Font = Font(bold=True, size=12, color="FFFFFF")
_TITLE_FONT: Font = Font(bold=True, size=14, color="000000")
_NOTE_FONT: Font = Font(italic=True, color="666666", size=9)
_BODY_FONT: Font = Font(size=10)
_BODY_BOLD: Font = Font(size=10, bold=True)
_CENTER: Alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT: Alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

#: DIN 4030-1 exposure-class fill colours.
DIN4030_FILL_HEX: Dict[str, str] = {
    "XA0":              "C6EFCE",  # green
    "XA1":              "DDEBF7",  # blue
    "XA2":              "FFEB9C",  # yellow
    "XA3":              "FFC7CE",  # red
    "Milieu unstimmig": "9C0006",  # dark red (white text)
}

#: DIN 50929-3 W0/W1 bucket fills (Mulden/Loch + Flächen).
W_BUCKET_FILL_HEX: Dict[str, str] = {
    "sehr gering": "C6EFCE",
    "gering":      "DDEBF7",
    "mittel":      "FFEB9C",
    "hoch":        "FFC7CE",
}

#: DIN 50929-3 Deckschicht-Güte bucket fills.
WD_BUCKET_FILL_HEX: Dict[str, str] = {
    "sehr gut":           "C6EFCE",
    "gut":                "DDEBF7",
    "befriedigend":       "FFEB9C",
    "nicht ausreichend":  "FFC7CE",
}

LIBREOFFICE_BIN: str = "soffice"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _set(
    ws: Worksheet,
    coord: str,
    value: Any,
    *,
    font: Optional[Font] = None,
    fill: Optional[PatternFill] = None,
    alignment: Optional[Alignment] = None,
    border: Optional[Border] = None,
) -> None:
    """Write a single cell value and apply optional styling.

    Args:
        ws: target worksheet.
        coord: A1-style coordinate.
        value: scalar to write.
        font: optional Font override.
        fill: optional PatternFill override.
        alignment: optional Alignment override.
        border: optional Border override.
    """
    cell = ws[coord]
    cell.value = value
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border


def _fmt_value(v: Optional[float], op: str = "") -> str:
    """Format a numeric value with optional operator prefix.

    Args:
        v: numeric value, or None.
        op: operator string like ``"<"``, ``">"``, ``"< BG"``.

    Returns:
        Display string with German comma decimals, ``"–"`` for None
        without a meaningful operator, or ``"< BG"`` when below detection.
    """
    if v is None and ("<" in op or "BG" in op):
        return "< BG"
    if v is None:
        return "–"
    if isinstance(v, float) and v == int(v):
        s = str(int(v))
    else:
        s = f"{v:g}"
    s = s.replace(".", ",")
    if op in ("<", ">"):
        return f"{op}{s}"
    return s


def _convert_xlsx_to_pdf(xlsx_path: str, output_dir: str) -> Optional[str]:
    """Convert ``.xlsx`` → ``.pdf`` via LibreOffice headless.

    Args:
        xlsx_path: input workbook path.
        output_dir: directory for the produced PDF.

    Returns:
        PDF path on success; None on LibreOffice unavailability or failure.
    """
    binary = shutil.which("soffice") or shutil.which("libreoffice")
    if binary is None:
        print(
            "  -> ERROR: LibreOffice (soffice) not found on PATH. "
            "Install LibreOffice to enable Aggressivität PDF export."
        )
        return None
    try:
        completed = subprocess.run(
            [binary, "--headless", "--convert-to", "pdf", "--outdir", output_dir, xlsx_path],
            capture_output=True, text=True, timeout=120, check=False,
        )
    except (subprocess.TimeoutExpired, OSError) as e:
        print(f"  -> ERROR: LibreOffice conversion failed: {e}")
        return None
    if completed.returncode != 0:
        print(f"  -> ERROR: LibreOffice rc={completed.returncode}: {completed.stderr}")
        return None
    expected = os.path.join(
        output_dir, os.path.splitext(os.path.basename(xlsx_path))[0] + ".pdf"
    )
    return expected if os.path.exists(expected) else None


def _apply_page_setup(ws: Worksheet, last_col: str, last_row: int) -> None:
    """Set print area + A3 landscape + fit-to-width=1 for clean PDF output.

    Args:
        ws: target worksheet.
        last_col: rightmost column letter to include in print area.
        last_row: last row to include.
    """
    ws.print_area = f"A1:{last_col}{last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    if ws.sheet_properties.pageSetUpPr is None:
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.sheet_properties.pageSetUpPr = PageSetupProperties()
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4


# ---------------------------------------------------------------------------
# Section builders
# ---------------------------------------------------------------------------


def _write_project_header(ws: Worksheet, project: AggrProjectMeta) -> int:
    """Write the project header block.

    Args:
        ws: target worksheet.
        project: project metadata.

    Returns:
        Next free row index.
    """
    _set(ws, "A1", project.projektnummer or "—", font=_TITLE_FONT)
    _set(ws, "A2", project.bauvorhaben or "")
    _set(ws, "A3", " / ".join(p for p in (project.los, project.bauwerk) if p))
    _set(ws, "A4",
         "Bewertung der Beton-Aggressivität nach DIN 4030-1:2024-07 und "
         "Stahl-Korrosion nach DIN 50929-3:2024-05.",
         font=_NOTE_FONT)
    return 6


def _write_din4030_section(
    ws: Worksheet,
    start_row: int,
    samples: List[Tuple[AggrSampleMeta, Dict[str, Optional[float]], Dict[str, str]]],
    results: List[Din4030Result],
) -> int:
    """Write the DIN 4030-1 Beton-Wasser section.

    Args:
        ws: target worksheet.
        start_row: first row to use.
        samples: list of ``(meta, measurements, operators)``.
        results: DIN 4030-1 results aligned to ``samples`` order.

    Returns:
        Next free row index after the section.
    """
    r = start_row
    _set(ws, f"A{r}", "DIN 4030-1 — Beton-Wasser (Beurteilung der Betonaggressivität)",
         font=_SECTION_FONT, fill=_SECTION_FILL, alignment=_LEFT)
    ws.merge_cells(f"A{r}:K{r}")
    r += 1

    # Threshold reference (compact)
    thresh_headers = ["Parameter", "Einheit", "XA1 schwach", "XA2 mäßig", "XA3 stark", "Milieu unstimmig"]
    for c, h in enumerate(thresh_headers, start=1):
        _set(ws, f"{get_column_letter(c)}{r}", h, font=_HEADER_FONT, fill=_HEADER_FILL,
             alignment=_CENTER, border=_BORDER)
    r += 1

    thresh_rows: List[Tuple[str, str, str, str, str, str]] = [
        ("pH-Wert",                "-",     "≤ 6,5 und ≥ 5,5",   "< 5,5 und ≥ 4,5",   "< 4,5 und ≥ 4,0",        "< 4,0"),
        ("Magnesium (Mg²⁺)",       "mg/l",  "≥ 300 und ≤ 1000",  "> 1000 und ≤ 3000", "> 3000 bis Sättigung",   "—"),
        ("Ammonium (NH₄⁺)",        "mg/l",  "≥ 15 und ≤ 30",     "> 30 und ≤ 60",     "> 60 und ≤ 100",         "> 100"),
        ("Sulfat (SO₄²⁻)",         "mg/l",  "≥ 200 und ≤ 600",   "> 600 und ≤ 3000",  "> 3000 und ≤ 6000",      "> 6000"),
        ("CO₂ (angreifend)",       "mg/l",  "≥ 15 und ≤ 40",     "> 40 und ≤ 100",    "> 100 bis Sättigung",    "—"),
    ]
    for row_data in thresh_rows:
        for c, v in enumerate(row_data, start=1):
            _set(ws, f"{get_column_letter(c)}{r}", v,
                 font=_BODY_FONT, alignment=_CENTER if c > 1 else _LEFT, border=_BORDER)
        r += 1
    r += 1

    # Sample table header
    sample_cols = [
        ("A", "Probenbezeichnung"),
        ("B", "Tiefe / Formation"),
        ("C", "pH-Wert"),
        ("D", "Mg²⁺ [mg/l]"),
        ("E", "NH₄⁺ [mg/l]"),
        ("F", "SO₄²⁻ [mg/l]"),
        ("G", "CO₂ angr. [mg/l]"),
        ("H", "S²⁻ [mg/l]"),
        ("I", "Einstufung"),
        ("J", "Lab-Aussage (DIN 4030)"),
        ("K", "Anmerkungen"),
    ]
    for col, label in sample_cols:
        _set(ws, f"{col}{r}", label, font=_HEADER_FONT, fill=_HEADER_FILL,
             alignment=_CENTER, border=_BORDER)
    ws.row_dimensions[r].height = 30
    r += 1

    for (meta, m, ops), res in zip(samples, results):
        cls = res.overall_class
        fill_hex = DIN4030_FILL_HEX.get(cls, "FFFFFF")
        cls_fill = PatternFill("solid", fgColor=fill_hex)
        cls_font = Font(bold=True, color="FFFFFF") if cls == "Milieu unstimmig" else Font(bold=True)
        tiefe_form = " / ".join(s for s in (meta.tiefe, meta.formation) if s) or "—"
        lab_text = meta.lab_din4030_verdict or "—"
        notes_text = "; ".join(res.notes) if res.notes else ""

        _set(ws, f"A{r}", meta.probenbezeichnung, font=_BODY_FONT, alignment=_LEFT, border=_BORDER)
        _set(ws, f"B{r}", tiefe_form, font=_BODY_FONT, alignment=_LEFT, border=_BORDER)
        _set(ws, f"C{r}", _fmt_value(m.get("pH"), ops.get("pH", "")), font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"D{r}", _fmt_value(m.get("Mg"), ops.get("Mg", "")), font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"E{r}", _fmt_value(m.get("NH4"), ops.get("NH4", "")), font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"F{r}", _fmt_value(m.get("SO4"), ops.get("SO4", "")), font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"G{r}", _fmt_value(m.get("CO2_angr"), ops.get("CO2_angr", "")), font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"H{r}", _fmt_value(m.get("S2"), ops.get("S2", "")), font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"I{r}", cls, font=cls_font, fill=cls_fill, alignment=_CENTER, border=_BORDER)
        _set(ws, f"J{r}", lab_text, font=_NOTE_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"K{r}", notes_text, font=_NOTE_FONT, alignment=_LEFT, border=_BORDER)
        r += 1

    return r + 1


def _write_din50929_section(
    ws: Worksheet,
    start_row: int,
    samples: List[Tuple[AggrSampleMeta, Dict[str, Optional[float]], Dict[str, str]]],
    results: List[Din50929WaterResult],
) -> int:
    """Write the DIN 50929-3 Korrosion-Wasser section.

    Args:
        ws: target worksheet.
        start_row: first row to use.
        samples: list of ``(meta, measurements, operators)``.
        results: DIN 50929-3 results aligned to ``samples`` order.

    Returns:
        Next free row index after the section.
    """
    r = start_row
    _set(ws, f"A{r}", "DIN 50929-3 — Korrosion-Wasser (Bewertung freie Korrosion an metallischen Werkstoffen)",
         font=_SECTION_FONT, fill=_SECTION_FILL, alignment=_LEFT)
    ws.merge_cells(f"A{r}:V{r}")
    r += 1

    # Compact lookup-table reference
    _set(ws, f"A{r}", "Bewertungsziffern Nx (unlegierter Stahl) und Mx (feuerverzinkter Stahl) nach DIN 50929-3:2024-05. "
                      "W0 = N1+N3+N4+N5+N6+(N3/N4); W1 = W0−N1+N2·N3; WD = M1+M3+M4+M5+M6; WL = WD+M2.",
         font=_NOTE_FONT, alignment=_LEFT)
    ws.merge_cells(f"A{r}:V{r}")
    r += 2

    # Sample-table header
    sample_cols = [
        ("A", "Probenbezeichnung"),
        ("B", "Wasserart"),
        ("C", "Lage"),
        ("D", "pH"),
        ("E", "Ca²⁺ [mg/l]"),
        ("F", "Cl⁻ [mg/l]"),
        ("G", "SO₄²⁻ [mg/l]"),
        ("H", "KS₄,₃ [mmol/l]"),
        ("I", "W0"),
        ("J", "W1"),
        ("K", "Mulden W0"),
        ("L", "Mulden W1"),
        ("M", "Abtragsrate W0 [mm/a]"),
        ("N", "Abtragsrate W1 [mm/a]"),
        ("O", "WD"),
        ("P", "WL"),
        ("Q", "Deck-Güte WD"),
        ("R", "Deck-Güte WL"),
        ("S", "Anmerkungen"),
    ]
    for col, label in sample_cols:
        _set(ws, f"{col}{r}", label, font=_HEADER_FONT, fill=_HEADER_FILL,
             alignment=_CENTER, border=_BORDER)
    ws.row_dimensions[r].height = 36
    r += 1

    for (meta, m, ops), res in zip(samples, results):
        wbf = lambda key: PatternFill("solid", fgColor=W_BUCKET_FILL_HEX.get(key, "FFFFFF"))
        wdf = lambda key: PatternFill("solid", fgColor=WD_BUCKET_FILL_HEX.get(key, "FFFFFF"))
        notes_text = "; ".join(res.notes) if res.notes else ""

        _set(ws, f"A{r}", meta.probenbezeichnung, font=_BODY_FONT, alignment=_LEFT, border=_BORDER)
        _set(ws, f"B{r}", meta.wasserart, font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"C{r}", meta.objektlage or "—", font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"D{r}", _fmt_value(m.get("pH"), ops.get("pH", "")), font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"E{r}", _fmt_value(m.get("Ca"), ops.get("Ca", "")), font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"F{r}", _fmt_value(m.get("Cl"), ops.get("Cl", "")), font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"G{r}", _fmt_value(m.get("SO4"), ops.get("SO4", "")), font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"H{r}", _fmt_value(m.get("KS43"), ops.get("KS43", "")), font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"I{r}", res.W0, font=_BODY_BOLD, alignment=_CENTER, border=_BORDER)
        _set(ws, f"J{r}", res.W1, font=_BODY_BOLD, alignment=_CENTER, border=_BORDER)
        _set(ws, f"K{r}", res.class_W0, font=_BODY_FONT, fill=wbf(res.class_W0), alignment=_CENTER, border=_BORDER)
        _set(ws, f"L{r}", res.class_W1, font=_BODY_FONT, fill=wbf(res.class_W1), alignment=_CENTER, border=_BORDER)
        _set(ws, f"M{r}", res.rate_W0_mm_per_a, font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"N{r}", res.rate_W1_mm_per_a, font=_BODY_FONT, alignment=_CENTER, border=_BORDER)
        _set(ws, f"O{r}", res.WD, font=_BODY_BOLD, alignment=_CENTER, border=_BORDER)
        _set(ws, f"P{r}", res.WL, font=_BODY_BOLD, alignment=_CENTER, border=_BORDER)
        _set(ws, f"Q{r}", res.class_WD, font=_BODY_FONT, fill=wdf(res.class_WD), alignment=_CENTER, border=_BORDER)
        _set(ws, f"R{r}", res.class_WL, font=_BODY_FONT, fill=wdf(res.class_WL), alignment=_CENTER, border=_BORDER)
        _set(ws, f"S{r}", notes_text, font=_NOTE_FONT, alignment=_LEFT, border=_BORDER)
        r += 1

    return r + 1


def _set_column_widths(ws: Worksheet) -> None:
    """Set sensible column widths for the Aggressivität output.

    DIN 4030 section uses cols A-K; DIN 50929 section uses cols A-S. Both
    sections share column A (Probenbezeichnung); we size for the wider of
    the two.
    """
    widths: Dict[str, float] = {
        "A": 24, "B": 18, "C": 11, "D": 12, "E": 12, "F": 12, "G": 14, "H": 13,
        "I": 12, "J": 19, "K": 14, "L": 14, "M": 17, "N": 17, "O": 8, "P": 8,
        "Q": 16, "R": 16, "S": 22,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Public entry point (signature preserved for step2_auswertung compatibility)
# ---------------------------------------------------------------------------


def create_aggressivität_report(
    samples: List[Tuple[AggrSampleMeta, Dict[str, Optional[float]], Dict[str, str]]],
    output_dir: str,
    project: AggrProjectMeta,
    source_workbook: str = "",
) -> Tuple[Optional[str], Optional[str], List[Tuple[Din4030Result, Din50929WaterResult]]]:
    """Generate the Aggressivität results report (.xlsx + .pdf).

    Builds a single A3-landscape worksheet from scratch with both DIN 4030-1
    (Beton-Wasser) and DIN 50929-3 (Korrosion-Wasser) per-sample rows.

    Args:
        samples: list of ``(sample_meta, measurements, operators)`` tuples.
        output_dir: target directory.
        project: project header metadata.
        source_workbook: unused in this build (kept for API compatibility
            with earlier reporter that cloned the company workbook).

    Returns:
        Tuple ``(xlsx_path, pdf_path, per_sample_results)``. ``pdf_path``
        is None on LibreOffice failure.
    """
    del source_workbook  # explicitly unused

    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "Aggressivität"

    # Evaluate per-sample first so the writers can stream the results.
    per_sample_results: List[Tuple[Din4030Result, Din50929WaterResult]] = []
    din4030_results: List[Din4030Result] = []
    din50929_results: List[Din50929WaterResult] = []
    for meta, m, _ops in samples:
        r1 = evaluate_beton_wasser(m)
        r2 = evaluate_korrosion_wasser(
            m,
            wasserart=meta.wasserart,
            objektlage=meta.objektlage,
            u_potential=meta.u_potential,
        )
        per_sample_results.append((r1, r2))
        din4030_results.append(r1)
        din50929_results.append(r2)

    next_row = _write_project_header(ws, project)
    next_row = _write_din4030_section(ws, next_row, samples, din4030_results)
    next_row = _write_din50929_section(ws, next_row, samples, din50929_results)

    _set_column_widths(ws)
    last_row_used = next_row
    _apply_page_setup(ws, last_col="S", last_row=last_row_used)

    safe_proj = re.sub(r"[^A-Za-z0-9_\-]", "_", (project.projektnummer or "").strip() or "Project")
    xlsx_path = os.path.join(output_dir, f"Aggressivität_{safe_proj}.xlsx")
    wb.save(xlsx_path)
    pdf_path = _convert_xlsx_to_pdf(xlsx_path, output_dir)
    return xlsx_path, pdf_path, per_sample_results
